# ============================================================
# NWA India — Admin Portal (Full Version)
# Features: Login, Dashboard, Detail View, Notes,
#           Delete, Export CSV, Filters, Enrolled Page,
#           Change Credentials
# ============================================================

from flask import Blueprint, render_template_string, request, redirect, url_for, session, jsonify, make_response
import json, os, csv, io
from datetime import datetime, timedelta
import config

admin = Blueprint('admin', __name__, url_prefix='/admin')

DATA_FILE     = 'data/enquiries.json'
BOOKINGS_FILE = 'data/bookings.json'
CREDS_FILE    = 'data/credentials.json'

# ============================================================
# HELPERS
# ============================================================

def load_enquiries():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r') as f:
            return json.load(f)
    return []

def load_bookings():
    if os.path.exists(BOOKINGS_FILE):
        with open(BOOKINGS_FILE,'r') as f: return json.load(f)
    return []

def save_bookings(data):
    os.makedirs('data', exist_ok=True)
    with open(BOOKINGS_FILE,'w') as f: json.dump(data, f, indent=2)

def save_enquiries(data):
    os.makedirs('data', exist_ok=True)
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def load_creds():
    if os.path.exists(CREDS_FILE):
        with open(CREDS_FILE, 'r') as f:
            return json.load(f)
    return {'username': config.ADMIN_USERNAME, 'password': config.ADMIN_PASSWORD}

def save_creds(creds):
    os.makedirs('data', exist_ok=True)
    with open(CREDS_FILE, 'w') as f:
        json.dump(creds, f, indent=2)

def logged_in():
    return session.get('admin_logged_in') is True

def get_stats(enquiries):
    return {
        'total':     len(enquiries),
        'new':       sum(1 for e in enquiries if e.get('status') == 'new'),
        'contacted': sum(1 for e in enquiries if e.get('status') == 'contacted'),
        'enrolled':  sum(1 for e in enquiries if e.get('status') == 'enrolled'),
        'dropped':   sum(1 for e in enquiries if e.get('status') == 'dropped'),
    }

# ============================================================
# SHARED CSS
# ============================================================

BASE_CSS = """
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Playfair+Display:ital,wght@0,600;1,500&display=swap" rel="stylesheet">
<style>
  *, *::before, *::after { box-sizing:border-box; margin:0; padding:0; }
  :root {
    /* NWA Arctic Navy palette — matches index.html exactly */
    --mocha:       #163240;
    --mocha-dk:    #0E2230;
    --mocha-lt:    #2F3873;
    --accent:      #AEB9DB;
    --soft-cream:  #E7EAF2;
    --soft-cream-2:#EEF1F7;
    --dark-cocoa:  #0F1B2A;
    --cocoa-2:     #1A2638;
    --off-white:   #FAFBFD;
    --line:        #D7DEEB;
    --muted:       #6A7387;
    --success:     #6FB890;
    --warning:     #E0A94F;
    --error:       #DC5855;
    --info:        #7AB0D4;
    --font: 'Inter', sans-serif;
  }
  html,body { height:100%; }
  body { font-family:var(--font); background:var(--off-white); color:var(--dark-cocoa); font-size:14px; }
  a { color:inherit; text-decoration:none; }
  input,select,textarea,button { font-family:var(--font); }

  /* ── SIDEBAR (dark navy — same as NWA navbar) ── */
  .layout { display:flex; min-height:100vh; }
  .sidebar {
    width:240px; flex-shrink:0;
    background:var(--mocha);
    border-right:1px solid rgba(174,185,219,0.12);
    display:flex; flex-direction:column;
    position:sticky; top:0; height:100vh; overflow-y:auto;
  }
  .sidebar-logo {
    padding:24px 20px 18px;
    border-bottom:1px solid rgba(174,185,219,0.12);
  }
  .sidebar-logo img { height:44px; width:auto; display:block; margin-bottom:8px; }
  .sidebar-logo p { font-size:10px; color:rgba(174,185,219,0.5); letter-spacing:2px; text-transform:uppercase; }
  .sidebar-nav { padding:16px 12px; flex:1; }
  .nav-label {
    font-size:10px; letter-spacing:2px; text-transform:uppercase;
    color:rgba(174,185,219,0.4); padding:8px 12px; margin-top:8px;
  }
  .nav-item {
    display:flex; align-items:center; gap:10px;
    padding:10px 12px; border-radius:8px;
    color:rgba(174,185,219,0.65); font-size:13px; font-weight:500;
    transition:all .2s; cursor:pointer; margin-bottom:2px;
  }
  .nav-item:hover { background:rgba(174,185,219,0.1); color:var(--accent); }
  .nav-item.active { background:rgba(174,185,219,0.14); color:#fff; }
  .nav-item svg { width:16px; height:16px; flex-shrink:0; }
  .sidebar-bottom { padding:16px 12px; border-top:1px solid rgba(174,185,219,0.12); }
  .logout-btn {
    display:flex; align-items:center; gap:10px;
    padding:10px 12px; border-radius:8px;
    color:rgba(220,88,85,0.85); font-size:13px; font-weight:500;
    transition:background .2s; cursor:pointer; width:100%; background:none; border:none;
  }
  .logout-btn:hover { background:rgba(220,88,85,0.12); color:var(--error); }

  /* ── MAIN CONTENT (light — same as NWA page sections) ── */
  .main { flex:1; display:flex; flex-direction:column; min-width:0; background:var(--off-white); }
  .topbar {
    padding:18px 32px; border-bottom:1px solid var(--line);
    display:flex; justify-content:space-between; align-items:center;
    background:#fff; position:sticky; top:0; z-index:10;
    box-shadow:0 1px 0 var(--line);
  }
  .topbar h1 {
    font-family:'Playfair Display', serif;
    font-size:20px; font-weight:600; color:var(--mocha);
    letter-spacing:-0.01em;
  }
  .topbar-right { display:flex; gap:12px; align-items:center; }
  .content { padding:28px 32px; }

  /* ── STAT CARDS ── */
  .stats-grid { display:grid; grid-template-columns:repeat(5,1fr); gap:16px; margin-bottom:28px; }
  .stat-card {
    background:#fff; border-radius:12px; padding:20px;
    border:1px solid var(--line); text-align:center;
    transition:transform .2s, box-shadow .2s;
  }
  .stat-card:hover { transform:translateY(-2px); box-shadow:0 8px 24px rgba(22,50,64,0.08); }
  .stat-num { font-size:36px; font-weight:700; line-height:1; }
  .stat-label {
    font-size:11px; letter-spacing:1.5px; text-transform:uppercase;
    color:var(--muted); margin-top:6px;
  }

  /* ── CONTROLS ── */
  .controls { display:flex; gap:12px; margin-bottom:20px; flex-wrap:wrap; align-items:center; }
  .search-box {
    flex:1; min-width:200px; padding:10px 16px;
    background:#fff; border:1px solid var(--line);
    border-radius:8px; color:var(--dark-cocoa); font-size:13px; outline:none;
    transition:border-color .2s;
  }
  .search-box:focus { border-color:var(--mocha); }
  .search-box::placeholder { color:var(--muted); }
  .filter-select {
    padding:10px 14px; background:#fff;
    border:1px solid var(--line); border-radius:8px;
    color:var(--dark-cocoa); font-size:13px; outline:none; cursor:pointer;
  }

  /* ── BUTTONS ── */
  .btn {
    padding:10px 18px; border-radius:8px; font-size:13px; font-weight:600;
    cursor:pointer; border:none; display:inline-flex; align-items:center;
    gap:6px; transition:all .2s;
  }
  .btn-primary { background:var(--mocha); color:#fff; }
  .btn-primary:hover { background:var(--mocha-lt); }
  .btn-green { background:rgba(111,184,144,0.12); color:#3d9e68; border:1px solid rgba(111,184,144,0.35); }
  .btn-green:hover { background:rgba(111,184,144,0.22); }
  .btn-red { background:rgba(220,88,85,0.08); color:var(--error); border:1px solid rgba(220,88,85,0.25); }
  .btn-red:hover { background:rgba(220,88,85,0.15); }
  .btn-ghost { background:transparent; color:var(--muted); border:1px solid var(--line); }
  .btn-ghost:hover { color:var(--dark-cocoa); border-color:var(--mocha); }

  /* ── TABLE ── */
  .table-wrap { overflow-x:auto; border-radius:12px; border:1px solid var(--line); background:#fff; }
  table { width:100%; border-collapse:collapse; }
  thead { background:var(--soft-cream-2); }
  th {
    padding:12px 16px; text-align:left; font-size:10px;
    letter-spacing:2px; text-transform:uppercase; color:var(--muted);
    font-weight:600; white-space:nowrap; border-bottom:1px solid var(--line);
  }
  td { padding:13px 16px; border-bottom:1px solid var(--line); font-size:13px; color:var(--dark-cocoa); vertical-align:middle; }
  tr:last-child td { border-bottom:none; }
  tbody tr { transition:background .15s; cursor:pointer; }
  tbody tr:hover td { background:var(--soft-cream-2); }

  /* ── BADGES ── */
  .badge { padding:4px 10px; border-radius:99px; font-size:10px; font-weight:700; letter-spacing:1px; text-transform:uppercase; display:inline-block; }
  .badge-new       { background:rgba(122,176,212,0.15); color:#3d7fa8; }
  .badge-contacted { background:rgba(224,169,79,0.15);  color:#a0731a; }
  .badge-enrolled  { background:rgba(111,184,144,0.15); color:#3d9e68; }
  .badge-dropped   { background:rgba(220,88,85,0.12);   color:#b83b38; }

  /* ── MODAL ── */
  .modal-overlay {
    display:none; position:fixed; inset:0; background:rgba(15,27,42,0.55);
    z-index:100; align-items:center; justify-content:center; padding:20px;
    backdrop-filter:blur(4px);
  }
  .modal-overlay.open { display:flex; }
  .modal {
    background:#fff; border-radius:16px; border:1px solid var(--line);
    width:100%; max-width:680px; max-height:90vh; overflow-y:auto;
    box-shadow:0 20px 60px rgba(22,50,64,0.15);
  }
  .modal-header {
    padding:24px 28px 16px; border-bottom:1px solid var(--line);
    display:flex; justify-content:space-between; align-items:flex-start;
  }
  .modal-header h3 { font-size:18px; font-weight:600; color:var(--mocha); }
  .modal-close { background:none; border:none; color:var(--muted); font-size:24px; cursor:pointer; line-height:1; transition:color .2s; }
  .modal-close:hover { color:var(--dark-cocoa); }
  .modal-body { padding:24px 28px; }
  .detail-grid { display:grid; grid-template-columns:1fr 1fr; gap:16px; margin-bottom:20px; }
  .detail-field label { font-size:10px; letter-spacing:1.5px; text-transform:uppercase; color:var(--muted); display:block; margin-bottom:5px; }
  .detail-field p { font-size:14px; color:var(--dark-cocoa); background:var(--soft-cream-2); padding:10px 14px; border-radius:8px; border:1px solid var(--line); }
  .detail-field.full { grid-column:span 2; }
  .notes-area {
    width:100%; background:var(--soft-cream-2); border:1px solid var(--line);
    border-radius:8px; color:var(--dark-cocoa); padding:12px 14px;
    font-size:13px; resize:vertical; min-height:90px; outline:none; transition:border-color .2s;
  }
  .notes-area:focus { border-color:var(--mocha); }
  .modal-footer { padding:16px 28px 24px; display:flex; gap:10px; justify-content:flex-end; border-top:1px solid var(--line); }

  /* ── STATUS SELECT ── */
  .status-select { background:#fff; color:var(--dark-cocoa); border:1px solid var(--line); padding:6px 10px; border-radius:6px; font-size:12px; cursor:pointer; outline:none; }
  .status-select:focus { border-color:var(--mocha); }

  /* ── TOAST ── */
  .toast {
    position:fixed; bottom:24px; right:24px;
    background:var(--mocha); color:#fff;
    border-radius:10px; padding:14px 20px; font-size:13px; font-weight:500;
    z-index:999; transform:translateY(80px); opacity:0; transition:all .3s; pointer-events:none;
    box-shadow:0 8px 24px rgba(22,50,64,0.25);
  }
  .toast.show { transform:translateY(0); opacity:1; }
  .toast.success { background:var(--success); color:#fff; }
  .toast.error   { background:var(--error);   color:#fff; }

  /* ── EMPTY STATE ── */
  .empty { padding:60px; text-align:center; color:var(--muted); }

  /* ── FORM ELEMENTS (settings / login) ── */
  .form-group { margin-bottom:20px; }
  .form-group label { display:block; font-size:11px; letter-spacing:1.5px; text-transform:uppercase; color:var(--muted); margin-bottom:8px; }
  .form-input {
    width:100%; padding:12px 16px; background:#fff;
    border:1px solid var(--line); border-radius:8px; color:var(--dark-cocoa); font-size:14px; outline:none; transition:border-color .2s;
  }
  .form-input:focus { border-color:var(--mocha); }
  .alert { padding:12px 16px; border-radius:8px; font-size:13px; margin-bottom:20px; }
  .alert-success { background:rgba(111,184,144,0.1); color:#3d9e68; border:1px solid rgba(111,184,144,0.3); }
  .alert-error   { background:rgba(220,88,85,0.08);  color:var(--error); border:1px solid rgba(220,88,85,0.25); }


  /* ── COUNSELLING ── */
  .cal-wrap{overflow-x:auto;margin-bottom:40px;}
  .cal-grid{display:grid;grid-template-columns:80px repeat(5,1fr);gap:1px;background:var(--line);border:1px solid var(--line);border-radius:8px;overflow:hidden;min-width:600px;}
  .cal-head{background:var(--mocha);color:#fff;padding:10px 8px;text-align:center;font-size:11px;font-weight:700;letter-spacing:1px;text-transform:uppercase;}
  .cal-time{background:#f8f9fb;padding:10px 8px;text-align:center;font-size:11px;color:var(--muted);font-weight:600;display:flex;align-items:center;justify-content:center;}
  .cal-cell{background:#fff;padding:8px;min-height:56px;display:flex;align-items:center;justify-content:center;font-size:12px;}
  .cal-cell.booked{background:#edfaf3;border-left:3px solid var(--success);}
  .cal-cell.booked .b-name{font-weight:700;color:var(--dark-cocoa);font-size:12px;}
  .cal-cell.booked .b-phone{color:var(--muted);font-size:11px;margin-top:2px;}
  .cal-cell.free{color:var(--line);font-size:11px;font-style:italic;}
  .cal-cell.past{background:#fafafa;color:var(--line);}
  .week-nav{display:flex;align-items:center;gap:16px;margin-bottom:20px;}
  .week-nav h2{font-size:18px;font-weight:700;color:var(--dark-cocoa);}
  .week-badge{background:var(--soft-cream);border:1px solid var(--line);color:var(--muted);font-size:11px;padding:4px 12px;border-radius:20px;font-weight:600;}
  .booking-list{margin-top:8px;}
  .booking-row{display:grid;grid-template-columns:1fr 1fr 1fr 1fr auto;gap:12px;align-items:center;padding:14px 16px;background:#fff;border:1px solid var(--line);border-radius:6px;margin-bottom:8px;}
  .booking-row:hover{border-color:var(--accent);}
  .booking-row .br-name{font-weight:700;color:var(--dark-cocoa);}
  .booking-row .br-slot{font-size:12px;color:var(--mocha);font-weight:600;}
  .booking-row .br-contact{font-size:12px;color:var(--muted);}
  .booking-row .br-date{font-size:12px;color:var(--muted);}
  .del-btn{background:none;border:1px solid rgba(220,88,85,0.3);color:var(--error);padding:5px 10px;border-radius:4px;cursor:pointer;font-size:11px;font-weight:600;}
  .del-btn:hover{background:var(--error);color:#fff;}
  .empty-cal{text-align:center;padding:40px;color:var(--muted);font-size:14px;}

  @media(max-width:900px){ .stats-grid{ grid-template-columns:repeat(3,1fr); } .content{ padding:20px; } }
  @media(max-width:680px){
    .layout{ flex-direction:column; } .sidebar{ width:100%; height:auto; position:relative; }
    .sidebar-nav{ display:flex; flex-wrap:wrap; gap:4px; padding:12px; } .nav-label{ display:none; }
    .stats-grid{ grid-template-columns:repeat(2,1fr); }
    .detail-grid{ grid-template-columns:1fr; } .detail-field.full{ grid-column:span 1; }
  }
</style>
"""

def sidebar_html(active='dashboard'):
    items = [
        ('dashboard',   'Dashboard',    '/admin/',            'M3 9l9-7 9 7v11a2 2 0 01-2 2H5a2 2 0 01-2-2z M9 22V12h6v10'),
        ('counselling', 'Counselling',  '/admin/counselling', 'M8 7V3m8 4V3m-9 8h10M5 21h14a2 2 0 002-2V7a2 2 0 00-2-2H5a2 2 0 00-2 2v12a2 2 0 002 2z'),
        ('enrolled',    'Enrolled',     '/admin/enrolled',    'M22 11.08V12a10 10 0 11-5.93-9.14 M22 4L12 14.01l-3-3.01'),
        ('settings',    'Settings',     '/admin/settings',    'M12 15a3 3 0 100-6 3 3 0 000 6z M19.4 15a1.65 1.65 0 00.33 1.82l.06.06a2 2 0 010 2.83 2 2 0 01-2.83 0l-.06-.06a1.65 1.65 0 00-1.82-.33 1.65 1.65 0 00-1 1.51V21a2 2 0 01-4 0v-.09A1.65 1.65 0 009 19.4a1.65 1.65 0 00-1.82.33l-.06.06a2 2 0 01-2.83-2.83l.06-.06A1.65 1.65 0 004.68 15a1.65 1.65 0 00-1.51-1H3a2 2 0 010-4h.09A1.65 1.65 0 004.6 9a1.65 1.65 0 00-.33-1.82l-.06-.06a2 2 0 012.83-2.83l.06.06A1.65 1.65 0 009 4.68a1.65 1.65 0 001-1.51V3a2 2 0 014 0v.09a1.65 1.65 0 001 1.51 1.65 1.65 0 001.82-.33l.06-.06a2 2 0 012.83 2.83l-.06.06A1.65 1.65 0 0019.4 9a1.65 1.65 0 001.51 1H21a2 2 0 010 4h-.09a1.65 1.65 0 00-1.51 1z'),
    ]
    nav = ''
    for key, label, href, path in items:
        cls = 'active' if active == key else ''
        nav += f'<a href="{href}" class="nav-item {cls}"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round"><path d="{path}"/></svg>{label}</a>'
    return f'''<div class="sidebar">
      <div class="sidebar-logo"><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAgUAAADoCAYAAAB/2Zd4AAAPnklEQVR4nO3d25acNhYAUPDK//8yeUiXTcoUBUKXI2nvl8nydAvddZCAXpdlWbZt25ZlWdZ1XRcAoCuvdfypNWdihxcQaADAJSXX45ezdTlpwW6daQCIrvRaWWKdPNwpKLkg56gkAQMAEWTbtg+yrq1R7/qf5itKBQPQrwg3sjXW6Zfqi/XHjFysNMECADm1XPijHTFUWyBTC363QCnXESgAzOPuOtH6Tr/mGhViMSxdYbUCEgBieLKu3Jn7Ixwv5BQmI0dKNaogAWAcNQKA0Rb/T8Jn8JNS2z930u2hgQFGVPIIYOYbx+4LsJf7eYKZOwZABCXn4VpHDD0ZslAvEYKEUTsOQAkRg4CZ5vFpCvpSYsvpapozdSyAq0od29Z+y2AE01dAzrt/OwkA50rNk+bfPFTIjgABIL8S86GjgDJUzhc5t7UcMwCzKHEk4O2w8lTaDbnOp5xzAaOJshtgvnxG5SVqsYNwJS2AWlo+uH0nTa5TmZnkWtgFCEBkuRdtQUAsKjeznB1cgABE0DoQuJImeajkwnI9XOghRaC23DcmbnTiU+mVeIsBiK71MwLmrPY0QAOePwAiabkjcDVN6tAQjeW46xccACly7zraxeyfhgnC7gFQQ6tdAfNNHzRSMIIDIDcPDHKVxgrMmwvAE3YFuEvDdcDuAXCVXQGe0IAdqb1zcCUtIAbBADloyE5dGbC5Br4BD3F5g4CcNGrnPHcAc/JH2ShB4w6i5s7B1bSA/AQDlKSRB+NYAcbU4pjAGJ+PBh+UYwXon4cHqU2jD87OAfQpZ0AuGOAqjT+B2pOLSQXStQgGjFledITJ5FjU3XVAXr48SBQ6xITsHEAcdgaIRMeYmIcRoR3PDBCRzjG5Fu88m5SYmWCAyHQSfvOmApTjc8T0QGfhfzxvAHkJBuiJTsMhn02GZ7xRQI90Hk45UoD7fJKYXulEfOUtBfiuxEN/xgy16Uhc4l1q+MxRAaPQobgl5zamLVF612p3wLigFB2LJLUfRLyaHtTgrxcyKp2LZI4UmE3L5wbupAmpdDAec6TADBwVMAOdjSxqv6FwJS3IwVsFzESHI6vauwZ30oO7PDvAbHQ6shMY0Du7A8xK56MIfwmOHnmQkNnpgBRl14Be2B0AQQEV2DUgMrsD8IfOSBWt/nzsnTSZy50+tCz6JnPQIamqxXHCnTSZQ6m+47iA3umYVGfXgFZaBwN304XadE6aEBhQU6mjgjtp63f0QCelqdyfeRUc8K717oB+Rk90Vpqza0AJdgfgPh2WEFoGBnfSJb4IwcDddCEKnZZQHCfwREp7b9u2+fPe8B+dl3DsGnBX6TYWEDALHZiQSkzCdg3GczcYWBZ9Bs7oxIRV6u7MRD8GuwOQn85MaK2/S383bcqzOwDl6NCE1/o985S0KSPK7kBK2tADnZpuRDhOSEmf52q0keMCEBTQGYHBXGoc9dgdgD90cLokOBhfpN2B1PShNzo53RIYjKnWg6ACAvibjk7XogQGKdfg/yIGA6nXgF7p7HSv5ANiFpDyagZgdgfgnE7PEEpO9nYNyhEQQCw6PsMo+T2DdV1XwUE+KXW5LHZ7oDSdn+FEOk5Ivc7I7A5AXAYBQyq9GAgO7qm5M5ByvZnbBvZ+tc4AlFDyc8d3039ynREICKAfBgNDq7F9bNfgWO1g4O41Z2gDuMugYHjRvoz39FrRPdkREZhBW44PGF6Nyb/mYhZZDwHB+iPlWjA6A4OpRD1OeHK9CFqV2fMDkJcBwnRqnTvPEBy02BlIvXZP9QqtGCRMSWDwzNNjj5p1Gr0uIRKDhWnVXFxGCQ5aBgMp149WfxCdAcPUeggMnl43h96CgRzXhBkZNEyv9t1nb7sGAgKYh4EDP2p++KaHXYPWwUBKHgQD8IwBBDu97BrkuPYnPQYDua4LszOI4E1PgUGO64+QDwEB5GEgwYEW29Y5vm7Y4g81CQZgHAYUfNBqkaq1SLcKQnLlRUAA+RlU8EWPuwZneYgUDCyLgAAiMbDggl53DV75yPmHl1o/wyAggHIMLrio58Agh9a7A7nzAPzNAIMbWm51twoOci/EjgsgLgMNEswQHEQIBkrkA/jMYINEre94SwUHJRZhAQH0wYCDB1oHBql5OBLpC4mCAWjDwIOHIix6UT5AtGd3APpj8EEGEQKDlHxEyMOegADaMgAho16Cg0hHBcsiGIAoDETIKNqi+J6faMHAsggIIBKDEQqIFhyUJCCAcRiQUMgMgcEMZYSZGJhQ0KiLpt0BGJPBCYWNtoCOVh7gDwMUKul910AwAOMzUKGiHhfWiB9GAsowWKGBXnYNegxigHQGLTQSfcGNnj8gv1+tMwCzerJwlvgLia80tx+p6QgIoF8GLwQQ5TjB8wMwNwMYgmh5dy4YAJZFUACh1A4MchxDCAhgHAYzBFPrrt3uAPDOoIaASt7B2x0APjGwIbCcd/O53lgQEMC4DG4IrsTrhykEAzA+3ymA4NYfrfMQJTgByhH5Q0dqL8ytgxGgLgMeOlQ6OBAMwJwcH0CHSi7aAgKYl6AAOlVi8RYQwNxMADAAHyICcrBTAAN4sqgLCIAXkwEM5uqugWAAeGdSgAGdBQaCAeATkwMM7D04EBAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAIazLsizbtm3ruq6v/93/25PEj9LYtm1blmV5mnYJr7x9EzHvL+/1e9S2vYrcd1Icjbf3fzv6vX359z8ToZ0/Xf9JvlJ+t0Q+UnyaA2u3Uam5eJ/Gvg8/TbekEeb5I7nKtX5L6MlALpFuCVcr812kMixLejnORCpjT33qmxJttdeiLq6UKWVxv/t7JfKR4iwftQOTbz8z+jz/dLxFKceRJ2U7KtfhHceVX7zqKN0oFXzlTuzK70Qpz7J8L1PvAVBKm0V1VpYcQUOkwCDHopMaUOTIR6oIeTjLy17OwCDKeMw1x1/5vdpS6v1K2S4FBVcudjUTESr1aSeOPAiW5XmdRx4U0ev+rhzjI1qglLONcgYFUYKk1sc7n/6/XDd/UcZjiXn+bhqllJrj13Vdf11NtPc7zDN387j+2P/b9iNvzto4Kh9xjdpeo4ynXoxU30eLZo55/ijt2nLtIn6qj19H/1gyMy0dLdxPJtOIHaa01uUbcfHLJUpwECEP7yLmKaLW4zuH3DtE778frS/lWsNe/30YFKSet8woWgfJaeSyjezVbpHG6t28HP18j7uVUcfQqHN8qbxHbccSPu4UjNZpap7t9Vg/R0Ypx4xa7xoYW/GNNscfGW0xrxH0nB4f5K7Q0RoIuK7mQjPKolbaSIFB6fy2DrRLeS/TrWcK9nrqMDU6S+1rQnSpE2jEJ9lHNkJg0Es+S8hd9q9BwVl01HNDmGzSqLeYenn75U4ec8w7+us1I9bTTGXKOfYv7xT0Ghi0zF/0uvnk0wIz4iAbQa/9rJRXfeiv95zN8fpYfNuBlHSSjw/eM5MjHdr51pFMsPGMOFmPVp7epB4ltJwfZuszd+o6JUj4JyUzR4lvW/9/cGdUqYNGe9Y32gS3rn9/svnKXPH+/vTVNEarvxaO6vsb9V5X6mvHV57XuRUUXLmghWQc7x1O21JSzoVFX33mU2Awwxzf0zMrKQHcy6cAISkoOMvIDJ2mN3fbQ9u2k6utIkndLcidBnkc1fuTxalXUfpe6q7B3r5Nk3cKZuwEs9h39qOJ+P1naCfHhFDDnfkix6uMPNfLcXHutehs/jv6mShyvQzw6EFDT6uOr9e3TmYTcZL6Zt+HHB30p1T7RTJCX1p3zn7u1YaP3z6IXmkt8xelbp4O2Cjl4Nzo7eTBwjauvJEwet8bxZXgIMsrie4mYxtpwI7ep0Yu35V++PToIFpfH6U9rwQG0UTNVwRn7ZklKDi7SNSGyZmvo7SiTU6jUK9jcdTYj8iBwQzzQu46/lRn2YKCyGboMGdaD9icRipLLZHq7Gwspu4kRCrf6CLPpZHzlkuNvv47KMhxsSvnFa3UnEyi1gHnRrxrHa08Z4y76570i57qeZT+X7McU+wUlDRKpzujjH0asUzL0teiNKqobTDDTlKu8nxK5zAoGPVp9dwdZoZnCXp6TzfFiLsDyxL7zw8f5edpHqOVkXZyzvNR54YS+XrV26/SF4gmR4c5WkiiHJ2UfoByWeK1bWr7tQ7qPuU7pQ23N89zRymt2uf9uqPe/C3L53n+Tpmjz3+pY/3T+vX7v18/dPTLOQof+Y7lbrl76CSlr9GyrKOVb7Ty3HE2KaWkEbGcZ+1bO7+l566obZHaBhHn+ivzRY4yff08ZM7AIFJn2UudnCOVp+QCE6GcIy2gNe8UI7TdJ0/nhcjzytMJvLe8RA0K9p6Mu4hlKlWeqndFESv2XY0gCRg7KJhRL+0RKWDLydoFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMbW2dAeaxbdv2+u91Xav2vf21P6mRp9p1cKXce6XzdDc/y5I3Tznq/5VGrnylpNdyLDG2X60zACVtP67+bOn8RHenvmZUsm7UOxH80zoDzOF9wtu2bat9h3N2vVf+auardh18u9a+jWrkrfUdbos+eGRd1/VOQGCXgJLsFFBVzUnszuS5/qiRn/11It0dvtdBpLxFUKM+7lxDQEAJggKK+zTR1dqKvTN5lppo38saeUKPnLfcogQ+V4OxKPllXIICqnlNfJG2zGuLlp9PXvm0CP2nRj3cOd7ppR/RH0EBRbVYVKItZN/yEy2/s7ga+LRYjM/yJCCgJEEBVXyayEouiNEnz+j5m0mUV1ZbjBPY8/YBxXy725ltovv29sNsQUKEhfhbP8z9TYI7jvrEbH2E+uwUUNyds9LRfCtb1En+/fXElnmpJUo53x86jJIv5mCngCKu3gXOMuFdWfwj7Rbs26ZUniKWda9l3zzKU5T6YmyCAqq48jBX7kmv9SI74912r16L8KvPRHvSP0IemIPjA4qrvSCaQMmlZV/Sj2lBUEB2qUFAieCh1R166tZvlB2FKPloYeayg+MDimm5EO63gO8cI5Q4crj60aYoi1G0rfNaIh4bQG12CsgqZWGr8b39K880lApO7v5Oy92N2RfFGcsMe3YKKOLu5FriTYT3NGsttqnXSd3dKJGvGovj1fzUXqgFBszMTgHZ5Fp0cy7e64/cP/vJ0zvtCAtShDzsRTlWqW3WctNWqMFP/57e4dZ4jbD0Fnm0Oth/b+C97C1e27y72EV7VbX09xverxUtSAPgInd3AJCJRRWgnX8Brcv1VTDaQbkAAAAASUVORK5CYII=" alt="Noble Wings Academy" style="height:48px;width:auto;display:block;margin-bottom:8px;"><p style="font-size:11px;color:rgba(174,185,219,0.55);letter-spacing:1.5px;text-transform:uppercase;">Admin Portal</p></div>
      <nav class="sidebar-nav"><div class="nav-label">Menu</div>{nav}</nav>
      <div class="sidebar-bottom">
        <a href="/admin/logout" class="logout-btn">
          <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" width="16" height="16"><path d="M9 21H5a2 2 0 01-2-2V5a2 2 0 012-2h4"/><polyline points="16 17 21 12 16 7"/><line x1="21" y1="12" x2="9" y2="12"/></svg>
          Logout
        </a>
      </div>
    </div>'''

# ============================================================
# LOGIN
# ============================================================

LOGIN_HTML = """<!DOCTYPE html><html><head><title>NWA Admin Login</title>{{ css|safe }}
<style>
body{display:flex;align-items:center;justify-content:center;min-height:100vh;}
.login-wrap{width:100%;max-width:420px;padding:24px;}
.login-card{background:#fff;border:1px solid var(--line);border-radius:16px;padding:40px;box-shadow:0 12px 40px rgba(22,50,64,0.1);}
.login-logo{text-align:center;margin-bottom:32px;}
.login-logo img{height:52px;width:auto;display:block;margin:0 auto 10px;}
.login-logo p{color:var(--muted);font-size:12px;margin-top:4px;letter-spacing:0.5px;}
.login-submit{width:100%;padding:13px;background:var(--mocha);color:#fff;border:none;border-radius:8px;font-size:14px;font-weight:700;cursor:pointer;letter-spacing:1px;text-transform:uppercase;transition:background .2s;}
.login-submit:hover{background:var(--mocha-lt);}
</style></head><body>
<div class="login-wrap"><div class="login-card">
  <div class="login-logo"><img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAgUAAADoCAYAAAB/2Zd4AAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAADE5klEQVR4nOx9d5hV1dX+e87tdfoMM7SBoUgXpIggAipgRbEbgdjLJ5+NGBWNv0ii0ZjoZzcqGmNBRRARjUCogggIgnSkDW2G6bf39fvjsvbsexkE4zADut/nWc/Mbefss88+e/W1tJqaGkSjUYrH47Db7RoA6LoOo9EIo9EIg8EATdMAAEQEGfz+kUBESCQS4i8fS9f1lO/E43EkEgmYzeaU38fjcQCAwWD40fMoKCgoKCicCEgkEgCQwucAIBKJQNf1FJ7KICKEQiHBe2Wex/zRZDId9hv+jIhQV1dH8Xgc0WgUwWAQoVAIsVhMjCUcDgOA4MEGgwFGoxEmkwm6rsPlcsFoNEKrrKxEJBKheDyOmpoaxGIxRCIRhMNhQXzwcDgsDmg0GmE2m2EymWAymWAwGJCfny9OYrFYYDabYTabNYPBkDIZuq6nkKZp0DQNXq8XJpMJZrNZTChP8NEEkGP9joKCgoKCwk9FulJ8LN9JJBKIxWKIx+NwOBzivXg8LiiRSIjvHToGRSIR+P1+1NXVwePxIBgM4sCBAwgGg/D5fAgGg4jH4zCZTLBarTCbzbDZbDAajbBarbBYLLDZbHA4HHA6nbBarcjOzhZ82GQyaSaTSQgguq4jHo8n+bGszbN00tCFsQQSi8UQCoUoGAwKaSQajSIej6O6uhrxeByRSAShUAiBQACBQEAIFbIEZTKZYLfb4XQ64XA4YLFY0KJFC7jdbmRnZyMjI0NjywFPmtFo/NEbki6ZKSgoKCgoNAaOJhTE43Gh9BIRwuEwvF4v1dTUwO/3Y/fu3QiFQvB4PPB4PPB6vYI3EpFg0DJvdLlcgj86nU5YLBa4XC643W5+rbGybTQaxfmZZwKA0WiEruvCUsHESOf/WmVlJcLhMEWjUaGhG41GjTV+liZkLZyIQERCwz/SBEajUSEwsPARDocpEAjA5/OhsrISpaWl2LlzJ8rKyrBz505UVlbC7/fDZrMhKysL+fn56NChA4qLi9GtWze0atUKRUVFGgsIsnvCZDIdNjZ+DSSFhmg0KiZQvtGapiGRSBwmWMjHY0nqpwofsvtEQUFBQeFw8D77Y+8lEokUviO7mGOxGGS+IH+PP2NeJO//RHSYS5v5iswfjEYjotEoiAgmkwmapiEajaKiooKqqqqwfv167N69G1u2bMGOHTtw8OBBBAIBwdDbtm2LwsJCtG/fHm3btkXLli2Rn58Ph8MBTdNgs9k0PpdsgWc3AfMoACnMXeZRPH5GQ7wq3a3Pv4/FYsnr2rt3L7G27/V6UzR9pkgkgng8LtwH7Iswm83CdGE0GtG6devD3AsWiwVWq1Vjhm02m1OYI1sh5Auora3Fjh07aOvWrdi2bRu2bNmCvXv3orq6GsFgELquIzc3F8XFxejYsSNOOeUUtGrVCgUFBTCZTHC5XHA6nZrRaBTCgsFgQCgUgtVqFRMTj8fF4ohEIjCbzYhEIsIq0ZDvR57YY3nvxwQnBQUFBYVUMKPifTP975Fi23hPlxU3YRJPU/rC4TCISPAD5m/Mu9icfkiBpUAggOrqalRUVGDbtm1Yv349tm7dioMHDyISiUDTNLjdbuTn56N9+/bo0KEDSkpKUFxcjMLCQrjdbs1isRwWU5d+bvn8/Nkhazvx9cViMeGS4NfxeByVlZXiNVv15ZgDi8UiLBHMmzkEQNd15OTkwGKxQJN/xIEMLEHI/g42h0SjUQoGg8I1EAwGhTWgrKwM4XAYfr8fPp8PgUAA4XBYuA5sNhtsNhsyMjKQnZ2N/Px8FBUVoaioCJmZmVo4HIbT6QQA4V9hph0IBBCPx1FRUUFlZWWoqKjAgQMHsGPHDpSWlqKurk5MisViQX5+Plq2bIlWrVqhuLgYLVq0QElJCXJzczWHwyEmVo5fOBI4zsJms6VYHdIXckPvNfRdBQUFBYXDwbynIebZ0HeZ2bN/XX6fFVxd12GxWITiBySZLSunkUgEVVVVVF5ejn379mHz5s3YuXMnysvLUV1djerqani9XrjdbrRo0QLt27dHq1atkJ+fj7y8POTn5yMrKwstWrQQjD8Wi7H2nzJmZtR8bcFgEHV1dRQIBLBnzx54PB5UVlaiurpa8DwA4noMBoOIFTik/MJqtcJoNMLtdqco63a7HXa7HRaLRWPFmJXUhoh5mCZLXXKkYkMaLpsZ0k0rPGiWrmRhQgq0oN27d6O0tBTr1q3Dhg0bUFpaisrKSng8HoRCIXTv3h1FRUXo1asX+vbti86dOyMvL0+z2WzCPGSxWMRY2SSfSCSEq8Lr9dKuXbuwYsUKLFmyBN999x0qKioQi8VgtVrhdDrRpUsXnH322TjzzDNRUlIirAosRRERAoEAAIjgkPR54GuW5+VIizjdZ6OgoKCgkIr0fTX9s1gsJhibYGDSdyORiHB1p+/H/FksFsMhBkzLly/HvHnzsHr1alRWVsLn88FgMMDpdKJNmzY49dRTMXDgQPTt2xdt2rQRPELOzJMzAAwGw2HW5UQiAZ/PB6/XSzt37sSOHTuwdu1abNmyBXv27EFNTY2wWLdu3Ro5OTlo164dTjnlFHTq1AktW7ZEZmYmzGazxtfG1vr0+ACeP+ZH6RaV9Ow++TcMTdOgRaNR8YYcyCcfmElOiTiS/0c+OINNGmaz+bCb6fV6sW/fPqqoqMDq1auxY8cObN26FWVlZfB4PACAnJwcFBQU4NRTT0VRURHatm2LvLw8ZGVlITMzEzabTdM0DXa7PcUdYTQaEQqFsHfvXiotLcXSpUuxa9cubNu2Dfv370dVVRV0XUenTp3QvXt39OrVCy1btkT79u1RWFgIp9Opmc3mlNgEvvF8wwGk3BR5YafHYSg3goKCgsKxgZWpdOYPHB7UJ1sBJGs1VVVVYefOndizZw92796N1atXY926dQiHw8jKykJeXh5at26Nli1b4owzzkBxcbFQRjlmQA4C5Fg5Hk8oFEJNTQ15PB74/X4cPHgQu3fvFrRv3z7U1dUhGo2K4MGioiJ06NBBxBa0adMGBQUFmtvtFtcMIEWLB1J5SEO8hPmR/Ln8PbZepB9XnlMAqZaCn4sfY4p80nTzO0uAbP6RJ7uyspL27duHffv2oaamRsQW7Nq1CzU1NQCA7OxsFBcXo1WrVujYsSNKSkrQuXNn5OfnC1MOn9tgMHBKB3k8Huzfvx+bNm3CunXrUFpaijVr1iAYDMJgMCAvLw9FRUVo3749TjvtNHTv3h1dunTRrFar8EOxFYRvliyJKSFAQUFB4b9DQ9bXSCQiNGT+TjweRzAYRDgcpvXr12P79u1Yv349du7ciQMHDqCyshKBQABEJPz8p512Grp06YKCggJkZGQgKytLy8jIQDgcPqxGgOwW3rJlC1VUVGDnzp344YcfxDk41i0ejyMzMxNt2rRBSUkJioqKkJubi6KiIuTn5yM/Px8Wi0Xj1EFWYGXrBoDDlMkfUzQZ/w2vkY+dkn3ATA1AgxJZQziW7xxNQGgog4EZN0tjAMSEsZQTiUQQiUTg8Xhoy5YtWLBgARYuXIitW7fC7/cDAKxWK4qKitC9e3cMGjQIQ4YMQZcuXTT5hnOgCQsOzODXrl1LCxcuxNdff43vv/8ee/fuRSQSESmTp5xyCs466yyceeaZ6NKli5aVlSWuKRgMijoNjPQoWAUFBQWFwyFrycyL2AoQi8WEK5eIUFNTgw0bNtCiRYuwcOFCbNmyBTU1NQgGgwCA3NxcnHLKKRgwYACGDh2Knj17Ijs7W7Pb7QDq48Tk/ZrPFwgEUFdXR6WlpVi1ahWWLFmCDRs2oKysTPAfo9GIli1bom/fvjjrrLPQrVs3lJSUwGq1ag6HQ7ihmW8ZDIaUwHY5o6AhzV1m1jw2+S9/9mNZgOluhKO5r1lp/0mWgqNpv8eSUtLQe3IdgiONR04J5Atg/000GoXf78eOHTvo+++/x7fffosdO3Zg3759KCsrQygUgslkwmmnnYaSkhL07NlT+Guys7M1m80mcjt50USjUdTW1tKuXbuwcuVKrFmzBmvWrEF1dbWIOC0oKECfPn1w+umno2PHjujRowfcbjcyMzM1OaqVzVxKMFBQUFBoGLJQwMrhodR1OiQEYNOmTfjmm2+wceNGVFVVwWQyIS8vD3l5eejcuTN69uyJgQMHomPHjsjOzk6pdSPHHHAxvtraWtq/fz/Ky8uxdu1alJaWCktDMBiE2+1GQUEBsrKy0L17d5SUlKB3797o2LEjsrKyUvz8cmofW75ZCIjFYikphDL/4s9/LNsNOJxvpgsMjaWANqr74HjjaGONRqMi8DAYDIobvnXrVuzatQtLly7F5s2bceDAATidThQUFKCoqAg9evRAz549cfrpp6NFixZadnY2gHrTkZyaUlNTg02bNmH16tVYtWoVvv32W1RVVSEjIwO5ublo0aIFunbtigEDBmDAgAHo1KmTZjabEY1GYTKZRDCnHKSSXgdCjjiV3z/SzVbWCAUFhaaEnOoHpGqtDPkzroPDFmDWzjk7gIMAjUYjKioqaOXKlVi2bBk2btyI3bt3o7KyEvv374fZbEZeXh66du2K/v37o0+fPmjXrh0yMjKQn5+vyWOSlTHOntu2bRt9++23WL9+PbZt24Zdu3ahvLxcRPq3b98ep556Knr06IEOHTqgbdu2KCgogNPp1DIyMn50Tn4pWWa/KKEAqDcBsWTGQkI4HIbNZkNFRQWtWrUKn3/+OebMmYOtW7cCACwWC8LhMNq3b49hw4bhwgsvRP/+/ZGXl6fxAk4kEgiFQrDZbGLheTweLF68mD788EN88cUXIpoUSC78Ll26YMyYMbjgggvQuXNnzeVyibHyGHnBclRrQwWU0i0r8kOoshoUFBSONxpKsU4vBieDM8TS0wB5v+N08FAohH379tFXX32FGTNmYM6cOSK2izXu3NxcXHrppTj77LMxfPhwLTMzUxyTrbzxeDwlriuRSGDPnj305ZdfYu7cuVi8eDE8Ho/oAQAA+fn5GDp0KIYNG4aLL74YLpdL47R4PgZf99H22V+KUnZSCQVHA2v2LJlysaT06E3OhohEIjhw4AB9/fXXWLFiBTZt2oSdO3eitLQUiUQCeXl5OOWUUzBw4ED07t0bJSUlnPmg8TGAekHE5/PR2rVrsWLFCmzcuBEbNmwQsQ4WiwUdO3ZE9+7dccYZZ6BHjx5o166dyG1NH5+cXytH1x7JdyT7qxQUFBSOF9JryAD1AgPvYbLgwHslV5T1er20Z88efP/99+C9d+fOnfB4PDAajSgoKEBJSQm6dOmCXr164fTTT0e3bt20dIWJ/fVc/OfAgQPEgYYbNmzAmjVrsHXrVsTjcWRlZaGwsBA5OTno1KkT+vbti9NOOw3t2rXTHA6HUCTl4nq/1mDxX5RQcCTITSlkP4/8OQBUVlaS1+vF1q1bsXDhQsyfPx+bN29GIBAQDZ+Ki4vRtWtXnHXWWTj33HORn5+vsTDCFRNjsRjKysrI7/ejqqoK8+fPx+eff44NGzakpFm2a9cOvXr1wtlnn43TTz8d7dq1E70oABz2cKWniP43JZcVFBQUGgOsjDS0D8k+9FAohNraWlqwYAGWLVuGr7/+Gvv27UNtbS1CoRDMZjMKCgowYsQIDBw4EAMGDEB+fj5sNptg2F6vFy6XK6UGTm1tLX3//feYN28eVq5cidLSUvj9ftTW1iISiSAzMxN9+vTB8OHD0atXL7Rq1QotW7ZEbm6uxrVtuKhPOvg8sjvi17LX/qKEAtauGbJPSU4tkYM65CwHLqHMbodIJIItW7bQzJkz8eWXX2Lz5s3weDyigJPVasXZZ5+NG264AYMHD4bVatVcLpdYRHKrTCLCmjVraPbs2Zg5cyZ++OEH+Hw+Ma7MzEyMGDECF198MUaMGKFlZGQIyVXTNAQCgZRCSnIAI79WbgQFBYXjCY7u50p9bNXkIHG5x0tZWRktWrQIH3zwARYuXAifzyeOo2kaWrdujWHDhuHqq6/GsGHDDgva48w4tqRGo1Hs37+fli1bhpkzZ2L+/PmoqKgQ4yEiZGdnY+jQofjNb36DQYMGadxXgMfU0L4pxx00VBPg14ZflFBwNPDi/akSXygUQjgcxoYNG+jbb7/F8uXLsXHjRlFembtanX322ejQoQMGDBiA3r17i8ZNckXDaDSKuro62rlzJ1auXInFixdj1apV2Lt3L8LhMAwGA1q1aiVSXQ5lNmiZmZnw+XywWCwpjT3Se2wrKCgo/Dc4Gi+QmWQoFIKmaSkVZisrK2nNmjVYtmwZvvvuO6xduxZ79uwBkGTaLVq0wKmnnoohQ4agb9++6NSpE/Lz8zUgmSLOpfCB5F7t9Xqxd+9e4oJDK1aswPbt21FeXg4AsNvtKCoqEq6APn36oHPnzigsLNQikQiISBzP4/GA47kaqvjHwkh6dUAWFn5N7tlflFDAAX5Aap5rQwWT0k1D7J/i9znIha0KzIgNBgMCgQCqqqpo586dmDlzJj799FMcOHAAfr9ftLnMyMjAkCFD8Jvf/Aann366xpJsetGKWCyGXbt20bZt2/Duu+9i4cKFOHDgABwOh7Be9OjRAxdccAHGjRuHvLw8ze12i+vlMSkrgYKCws/B0XgBK1Xs3jSZTKioqKB///vfmD17NpYtW4ZAIAC/349QKAQAKCwsxAUXXICRI0eiX79+cDqdmtPpPKxjIe9f0WgU33//PX366adYsGCBKFTn9XrFPtetWzeMHj0aZ599Ntq0aQOXyyXcDHLcVSQSAVBf3lfuophuWf1vFcZfIn5RQkE60os7NOQDa6hyFr8PIEWwkF0QnGKjaRrq6uqwfPlyevPNN7F8+XKUlpbCarUiFAqBiNCpUydcfPHFuPrqq9G6desGfVq8SPft20ezZ8/Gv/71L6xdu1ZE1rLpbcSIEbjmmmswZMgQtGzZUgQpKveBgoLCz8GxWAoikQjq6upow4YNmDZtGqZPn44DBw6ITnuapiEnJwf9+/fHmDFjcM4556BFixaa3KuGz8UCRnV1NZWXl+Ozzz7DrFmzsHTpUgAQKYxutxtt2rTBhRdeiFGjRuGMM87QuDotZ4Klp3zLWr3cXjk9gyIUComeNzLkuK5018IvHb9ooaCxcSxztX37dlqxYgWWLVuG5cuXY8uWLfD5fKKLVZs2bUSQYt++fVFQUKBxx0ZenLFYDMFgENu3b6elS5di0aJFmD9/Pmpra0WwZHFxMfr06YMhQ4ZwRS0tEAjAbrcfZorjB4b/AsdWLENBQeHkQUMpg3INkyMVkmNmzfsDH4djmgBg06ZNtHz5cnz77bdYtmwZtm/fLoKmMzMzUVBQgDPOOAPDhw9Hnz590LVrV43PDUCkGFosFsTjcZSWltLq1avx1VdfYdmyZdi2bRs8Ho9Q2uSGRAMHDsQpp5yC3NzcH92s1F7WOFBCwU/A0eZKloQTiQQ8Hg++++47eu+99zBz5kwcPHgQbrcbwWAQsVgMbdq0wejRo3HbbbehS5cuGv8OSD7Y3HbaaDRi7969xJL5ihUroGkaHA4HIpEIcnNz0aVLF9x1110488wzRQ5vKBRCPB6Hw+FISddkgUDX9cNMbAoKCic3OEBP7hOQXuCMAwRZcWCtW84a8Pv9mD9/Pv3zn//EwoULEYlERGG4RCKB3NxcDB8+HFdddRXOPPNMWK1Wzel0ihgqzvriKrHxeBxz5syh9957D0uWLEFZWZlw2wLJuIPzzjsPV1xxBfr164eCggLNarWCiMT4fgxKKGgcKKGgEcG1BBKJBAKBAMxms3hdXl5OU6dOxdSpU7FlyxYAyQfV6/VC13WcccYZGD9+PEaOHImMjAwtkUiAmTsfNx6Pw+/3Y9myZfT222/jP//5D2pqaqDrujCbderUCbfeeiuGDh2Kdu3aaW63WxQY4UIiRyuXqaCgcPJB9pmnZwLInwUCAVgsFvFZIBAAM999+/ZRbW0tZs+ejX/+85/YsWMHYrEYbDYbQqEQHA4HevfujTFjxmDUqFFo3769lm565/4BDocDtbW1WLduHX355Zd45513UF5ejnA4DIvFAl3XYbVa0aNHDwwYMADjxo1DcXGx5nQ6hbXUbrcrt2gTQwkFxwHsw5LTIeUOkcuWLaOFCxfiyy+/xNq1a2E0GsWDmZ+fj8GDB+OKK67A6aefDqfTqclNO/hYoVBIFF765JNPsGTJEng8HhFB26JFC3Tv3h0jR47ERRddhLZt22osDHBFMY5naCjwUkFB4eSCLOzLGU8MuaRwIpGA3++H0+lEOBzGli1baO7cuViyZAmWLVuGyspK2O124a7s1asXLrroIpx++uno3bu35nQ6UywPdXV1sFgssFqtCAQC+Oqrr2jx4sVYunQptm/fjj179iAzM1PUG+jbty/OOeccDBgwAF27dkV+fr7GXRDThYB0AUfhOCO9II6in0eRSER0YGSKRqOi4Ibf7xdM2ev1YtWqVXTllVcSAAJAFouFHA4HAaCCggKaPHkylZaWEscJxGIxcTw+X1lZGW3atIl+97vfUU5ODgEgo9EojmOxWOiOO+6gbdu2UW1t7WFjY3Mjj1Euo6xIkaKThzhjil/HYjH4/X5RX4WIRCthIsL06dNp2LBhZLPZCAAZDAYymUxks9moRYsWdP3119OKFSsoHA4jHo+L/YKrCPp8PnD6X01NDZ555hnq3r07ORwO0jSNNE0jAGQymSgjI4Puu+8+2rRpE3G3QR4HV6DlgET5ffl7io4/KUtBI0I20QEQi5kjc9MjZDlqNpFIYPfu3fTmm29i3rx52LhxY8pD0bp1a4waNQpjxoxBz549UVRUJMosy3nCRISKigp69913MW3aNGzfvl1oB3V1dXA4HLjqqqtw5ZVXokuXLsjKytLkrpB8LIayGigonDw4WmO0YDAo+gx88cUXePfdd7FlyxZhzUwkEqIU8GWXXYaxY8eisLBQuAe4YiuQ3C+MRiPq6uqwaNEimjt3Lj755BNUVVUhFAqJgOdWrVqhZ8+enDWl5eTkiLFGIhFhrWQXKX8mWzpUmmDTQgkFjYj0zmEyOFCG/WnpAYUWiwWhUAjbtm2jbdu2YcaMGfj0008RDAZhMpkQDAaRnZ2Nzp07Y/To0bj00kvRrl07jdMeuTAH5/P6fD5asWIFZsyYgS+//BL79u0T57Pb7ejWrRuuuOIKjB49Gu3atdOO9OApwUBB4eQCN4DjOv6BQAAVFRX02WefYcWKFVi4cCH27dsnWghbLBa0bNkSV155JQYNGoQzzjhDdIolIpGKyK4Hv9+P77//nv79739j7ty52LRpk4iNikQicLvdwnV5wQUXoLi4WByPrahWq1UII8eS9sd7pEIToLlNFb9E4qJCrM2z1MtxBewGkM1kbFWIRqMiZmDt2rV0++23k9VqJaPRSCaTicxms3At/O53v6M9e/YQUdIkyAKC7FqIx+PYunUrTZw4kaxWK+m6Tmazmex2uzDr3XDDDbRx40aS3QfKlaBI0clFbC3kuv1EhD179tCjjz5KLVu2THENGAwGAkDdu3en5557jioqKoizlYgIfr9f/O/z+cQ55s+fT9dddx1lZ2eL/cNut5PBYCBd1+n888+n+fPnCzdlMBgULtN0dwG7DOSxy8T7ZHPP66+NlKXgBIJcR4DNaUSEFStW0GeffYYZM2bghx9+QCQSgdPpFIFCl19+Oa677jr06NED2dnZGhc8ikQisFqt4lh79uyhDz/8EFOnTsXGjRvFQwoALpcLo0aNwvjx49G3b1+43W7NZDKJyotyARA2+8lSPac5Aql91OXXCgoKPw6ihuvvR6NRYUqXtWauxMfadl1dHaLRKH333Xfgdu579+6F2WwW7s22bduiY8eOuOWWWzB69GgtGAyKNsZsPZDPu3v3bpo3bx7ee+89rFmzBsFgUGj6uq6juLgY5557Lm688Ua0bt1ay8jIQEPuzV9LmeCTHUooOEHBWjo/SMFgEAcPHqQFCxbg7bffxqJFi8TnXFlx8ODBGDNmDMaNG6dxISS5CiOQrAFeXV1NGzZswGuvvYYvv/xSxDZwgaWOHTvi2muvxWWXXYbMzEyNiIQvUQZL/NxwRPn+FBR+HuQ6AeFwGACQ3lqdv8d/uSDQwYMH6ZNPPsGsWbOwfPly1NbWCsXCYDCgU6dOuPzyyzFy5Ej06NFDs9lsInWaS6prmiae9fnz59Pbb7+Nr776CqWlpTCbzaJGgc1mw9ChQ3Hddddh8ODBKCws1EKhEDgrIf1a5LErnOBoblOFolRKj7SV3QzBYFAw4unTp1P//v0JAJnNZnK73QSAbDYbFRYW0syZM4mjhaPRKMLhcIq5jv//9NNPqWfPnmSxWETWQkZGBgGgdu3a0dtvv01sxvN6veJ3kUjkmE17ckqmIkWKGqYfc9dxxgD3FSCqN7mXlZXRE088QdnZ2cItYDabKTc3lwBQ69at6W9/+xtxBgLvBXKmgpyRsHjxYhoxYoTYDzIzM0nXdQJATqeTrrzySvr666+JKOmqYEGBr0F+1tPdpIpOfFKWghMMfGMApEQSB4NBUfCD+xzU1dXh888/Fz0X+EG32WzweDzo27cvbr31Vpx55pno3LmzBiQjiHVdh9lsFrUKiJKpSR999BFWrlyJXbt2AUhWGAsGgygpKcE999yDiy++GLm5uZrFYoGmacKlwO4F1gjSx87XpTQFBYVjRygUgtFoFI3T/H4/MjIyAAA1NTXYtGkTvf/++5g+fTr2798Ph8Mhiga5XC506NAB1113HS699FLk5ORoXOSMrY9sRfT7/aioqKC5c+fi/fffxzfffINAICBqmJhMJnTq1AkjR47Eeeedh/79+2tOpxPRaBTBYBBut1s89+x6kIsoAcqFeFKhuaUSRfWUnl+cXu+AKBn0w9/jsqM1NTVYvHgxXXHFFZSZmUkWi0UEE+q6Th07dqTJkyeTrA1UV1cL7YA1j5qaGixcuJBuvPFGslgspGka5ebmkq7rZDAYqHXr1vTYY4/Rhg0biI/j9XoPG6McaKm0BEWKjp2CwaCwCMokBw9PmzaNRo0aRUVFRWSz2UjXdcrMzBSWwiuvvJLmzp1LBw8eJLbssfYeCARSjv/DDz/QY489Rh06dCC73U5ms5mcTqeoWTBgwAB68sknadeuXeJY6VZHHteRnnUVsHxyUbMPQFE9pRceISJRrCgYDKaY5fx+f0pWAxc3WrBgAfXo0YMyMjJI13WyWCxiw7Db7fTOO+9QZWVliptCPi8LIl999RUNGTKENE0ji8UiopYBUF5eHv3973+nmpoaUdAkvSgSH1cJBooU/XSKxWJCWCdKCt/z58+nwYMHEwCyWq1kNpvJarWS1WolAHThhRfSt99+S0RJ5YEZN7sPObsgHo+jtrYWf/7zn0UWktVqFc83AGrfvj397W9/S8kikLMQ+LnnPSldKODPm3seFf10Uu6DEwzsk0uvJ86QMxSi0SiAZF9zOf8XAN5//32aNm0aPvvsM5GtwFkDgwYNEgFHLVu21MLhsGhaIgcbeb1efPbZZzRjxgwsX74clZWVIgo6Go2iS5cuuOaaa3DRRRehZ8+eWjgcTumNLhbZjxRUUVBQSIIrBeq6Lp5xbqU+a9YsLFy4EOFwWAQh6rqOtm3bYsSIEbj66qvRr18/zWg0puwD7CJgt8GqVavon//8Jz7//HPs2LEDBoMBDodDdDy84IILcPnll2PEiBEoKirS0vejhrII5N4s7ELkOiwq4+AkRHNLJYrqKV2rlssZc7nPdGk8/X9uJEJE8Hg8mDt3Ll1wwQWkaRqZTCaRW2y1Wql79+40ZcoU4o5pLFjI9RUSiQTq6urw1Vdf0fDhw0nXdbJarSKgyW63U6tWrei+++6jqqqqFPeGCjBUpOjYKZFICO26qqoKL7/8MvXp00fUKJE1+YKCArr33ntp06ZNFAgEhDWAj8PBwZFIRBQbuvrqqykvL0+4CYxGo6hdcNZZZ9Hs2bOpvLycOIhYtlbwc8znkAMI062b6ZbBhiygik5cavYBKGpckh9afqDLysrotddeozZt2pCu62S320XMgKZp1K9fP/r888/J6/UiFAqlPPCRSAR1dXXi+P/617+oe/fuZDAYyO12i+MYjUbq2LEjvfnmm1RdXd1gpkP6/6o4iaJfGjFDbsh1xu/xa9k9EAqF4PV6EQ6H8d5771Hnzp3JYDCIrCBm5BkZGXT11VfTd999R0T12UosvBMln33+/4cffqAJEyaQw+Ego9EoeqFwhlGPHj1oypQpxO5IVbxMkXIf/MIguxfSi5x4vV689dZb9NZbb2H9+vXgCOJwOIw2bdpg4MCBuPnmm3HaaadpDocDRCRKm8rHLS0tpU8//RT//Oc/sWrVKhgMhpTubEOHDsW1116LCy+8EG63W7PZbOB1JtdZZ9OifGwFhZMVcv3+dPCzmF57gKuXer1emjlzJt5//32sXr1aBAFzqWKbzYaxY8dixIgRGDJkiGa324VQYbPZhFm/trYWLpcL27ZtoxdffBEzZ87E/v37Rc0SLj7Wt29f3HvvvTj33HM1m80m+rOkQ7n9foVobqlEUeOSrHmHQqGUcqW8iZSWltIjjzwi8piNRqPQIlwuF11//fW0efNmIkpqIhygxIFGwWAQgUAAfr8ff/3rX6moqIh0XSeXyyXMm7m5udS1a1eaMWMGyb+RxydHMSuNRNEvgaLRKAKBgLAShMNhsebltc9BeuFwGG+//TaVlJSIoD+DwSDKmWdkZNAtt9xCW7ZsIbnTIVGqe44DC0tLS+mmm24iu91ORqORLBZLSlnj/v3700cffUThcBjBYFBYGni8ykKgqNkHoKjxKL0FMr8fj8dTeiMwLV68mMaMGUNZWVlkMpnIaDSK/gh5eXn0pz/9iQ4ePEhyaqRcwIiP9/3339MVV1xBGRkZ4jgAyO12k8lkonHjxtH3339PLGBwQSU5A0K1R1V0slNDhceY+cvuhEAggNraWsybN4/OOussYdJn4dzlclFOTg4NHz6cPv30U5KPycK5nAkQCARw4MABevHFF6l169akaRq5XC4yGo1ks9nI4XBQfn4+Pf7447R7924iSk0l5mMpgUARkRIKflEkM+8jBfiFw+GU3uoVFRU0ffp0GjhwoPAzAiCHw0Eul4tOPfVUevnll0UAUnp8ADc7ISLMnj2brr/++hTBgIMbW7duTbfddhutXLmS0sfn9/tV+pKiXwRxYF96EDAH79XW1mLmzJl0wQUXkM1mS6kLwI2FzjjjDProo4/I7/en1CtpKP5m165d9Pjjj1Pfvn2FcMHWBhwKSHziiSdo+/btlN6EiIhS0gqVMKCISAkFv1hizSQSiQgNhoUBbpbEEcZESW3hqaeeIpvNRi6XS5Q1zcrKIofDQeeee65wKRClahocjBiLxeDxePD2229T27ZtxeaUkZFBmqaRw+GgoqIievzxx8nr9SKRSDRY/EiRopOROKXwSO8tWbKERo8eLWoMsABuMpnIYrFQQUEBTZ8+XZj2+Rgys66urhZBxNOmTaMzzzyTNE0jt9sthAyLxUJ2u51uuOEG2rFjBxEln2+Z+Xs8nhSBXgnlipiafQCKGpdkIaAhkosgcd1y/h0RYfPmzXTeeeeR2+0WlgPWOrKzs2ny5Mn0ww8/kFzrXD4+azWVlZWYOHEiFRUVpWgwNpuNNE2jPn360LRp00T6U7prQ5Gik43kZ4FN8vF4HP/5z39owoQJZDabhTBgNpspLy+PzGYzdejQgR544AEhDDCzDofD4jj8fAaDQaxatYrGjh0rjsUCvN1uJ5PJRIMGDaIvvviCeExyRkT6mGOxWIrAoEhRsw9A0fEjua6BbCbk13JFRK5vIJs4zzjjDBEbwPnMZrOZTj31VPrHP/4hqp1xCqNcK4EDrFasWEE33ngjmUwmMplMZDAYyGq1ktFoJKfTSZdccgl98cUXpGIKFJ3slEgkhNCdSCSwcuVKmjBhAhUXF4vqojLzzszMpPvuu492795N/FuiZKyOXAOAn8/ly5fTb3/7WyosLBRWBj6WwWCgfv360VtvvUUc9xMOhw+rLCj/L1sjVGqwIqZmH4CiE4d4Y+CNraqqCnfddZfIUODCRbwJDRkyhNauXUtEqZqRLCBwhPWsWbOoTZs2YiNjs6nRaCSHw0GPPPJIShe39LLLMskaGQcvNvfcKTr5SWaM6TEB8mey2Z2F2aSFLgSiOGKxCJ588gkqKMgjo1EnTQOZzcZD2QTJDJ1zzhlO69Z9R/F4VPyOj5XeuTCRSODpp5+mnJwcUXKc435wqETxww8/TBxMLAv8TCpeQNGxUrMPQNGJRTKD5Y1xwYIFdO6551J2drZwAXCaU15eHj333HNUWlpK8gZJVB/DwJvcvn37aOLEidSqVSsyGAyk6zo5nU4hJHCFxdra2sMsGfLYGnI3KEuDosagRCIhGGv6Z7KA6vf7hfDM63vv3lJ6//13qWfP7mQ06mS1mg8xbxsZDBo5HDbq1+80+uST6RSPRxEOB+H3exGLRUCUKoRwi+QPPviAunTpQhaLRaT82mw2slgs5Ha76be//S0tWrSIiJLPQLpFUL6u5p5bRScHNfsAFJ1YJJsUq6qqxIa3f/9+ev755ykrK0v4MDVNo6ysLAJAQ4cOpenTp1N6r3fejFhACIVCWLBgAV122WVktVpJ0zSy2WzCamA0Gql///40Z84ckgMRj2Te5M25uedN0clPDa0xNsNHIhGxztIFhkAggA8++ICGDTuLXC4H2e1WMhp1Mhg0MpmS8TidO3ek//u/Z6iqqoK83jr4/V4kBYEkRSIh+Hw+hMNhhMNhLFq0iK655hrKyckhs9ksYnH4/yFDhtAHH3xAHOBL9OOMXwkFio6Vmn0Aik48CoVCKYxWLnBSXl5OQ4cOJYPBkBLtbDabSdM0euihh0guShQIBIRrQS6XTER49tlnRRAj/97lconj3XfffcS/YZcCZ0405FZQpOjnULrL4EhuKc6yISKUl5fTDTfcQLqedBNoGkjXQRaLiQwGjQDQmDGX0L59e4hdC9FoGCwMxGIR1NXVQLYUPPnkk8IqYDQaSdM0IYibzWaaMGECVVdXg6heQFExAYoai5p9AIpOHOJMgPT3OVgpGAwikUiguroaf/nLX6ikpIQAiDxrTdNEo6VPP/2U2CeabtoPhUJCUFi3bh1dfvnl5HK5yGQykaZpwl9qMBioZ8+e9N5771F5eTlxe+iGxq3cB4oag+SYGHlN8VomSjLivXv30iOPPEKFhYUigNBiMZGuJ90FFouJBg0aSPPmzSGiODyeWoRCAXi9dWCBQP6/svIgzZgxg/r37y8aH+m6LqqOWq1Wuuqqq2jp0qUi80d+juQeCKoIkaKfQ80+AEUnFslWgvSNkYhQU1MjNKiNGzfStddeK4IHOTbA5XJRbm4uXXvttbRlyxYiqjdfyow9GAyKtKtPP/2UOnfuLDq3cSCizWYjt9tN559/Pq1bt47YnSCnbilS1FgkW6DYKsX/ExF2795Nr7zyCnXr1o0MBoMI9nO5XKRpyYyAkpJ29Nprr1J1deVhbgKmYNAPn88Dojg2blxP1157tYjZYaHYarUSAOrUqRP985//pH379hFRMqhXFlKY0isnKqFA0X9DzT4ARScOyRuN3OGNzao1NTUpr7mM69NPP002m01UMeSN0mKxUP/+/WnGjBlCu2FzpxwQFQ6HEQgEUFNTg1tuuUVYDOSqiBzY+OSTT1JDqVuq+IqixiDZhZBe3nv58uV02WWXpaQDyuZ9u91KY8f+hqqqKogojurqSqRbBAIBH2KxiAgunDLldbJazaTrEM+MzWYju91OmqbRuHHjqLa2NqWtcnp7dab0/gXNPZeKTk5q9gEoOrGIU/zSeyfw/3K6YTQaFVaDefPm0fDhw8ntdqe4FABQy5Yt6fe//72wGsgxAekugQMHDtD06dNp2LBhKTXh+VgWi4WGDx9O06ZNI66roCwGihqD0mttcAzLihUr6He/+x0VFBQIIdVqtQqLlt1uPxRoO41YEIjHo5AtAywERCIhhMNBLF68kK644jJyu52k6yCTyZBS6bBPnz70zjvvUF1dXYPFh+TsnCMVEku/tuaeX0UnBzX7ABSdnJTeNz4ej6O6uhqTJk0SZk+2GJjNZtJ1nfr27ZvS4EWuR8CbFpc/5oqITqeTNE0jTdPEJsylk8ePH0+lpaXEaWREDQdeyRYQtm7IG6naMH8dJFuT0vt4yOuYXycSCUyePJmKi4tTzPlWq1VkA2RkZNAf//jHQzEvQRDFEQoFkC4MBIN+JBIxVFSU0yOPTKLMTLdIWdQ0HMpWMJCmaTRhwgQ6cOBASq8CVYtDUVNRsw9A0clL6YIBb1zTp0+njh07kslkEhYDrm1gsVjoz3/+syhUxK4AOf1Q/v+TTz6hdu3aiUBENquy9aBt27a0dOlS4na1/HtO7Tpa+VaZESj6dRJ3ESVKMt+qqip88803NGDAAGH5YuIuogaDgQYPHiz6gSSZd71lIB6PIhoNi7+BgA/ffbeazj337EM9RTIOHc9MdruVbDYLtW/fnj766CNiCxwLryqIVlFTUrMPQNHJSz9WV33t2rX029/+NkUokLMKLrzwQpo3bx5Fo1FwuWQm3gRZyNizZw/dfvvtIsqbj6PruhAOxo0bR8uWLSMigiwg8DGOtLEqoeDXQ7K2zRq43P+DiLBmzRqaMGEC5ebmpsS0cIqg2+2m7t2704svvkhcI8Pv9yOZhptMNZQrGxLFsX//Xnrwwd9Tixb5IkshaUEzimyF668fT6tXrxaxN3JrZGUlUNSU1OwDUHRyU3pgk5wiVVtbiylTplB+fr7IuWZ3gtVqpfz8fHr99deJqL77W3qwI7eiTSQSWLx4sWgRazAYUny6JpOJTjnlFHr66aeJUyh5M+VaB8pVoIjo8BoEoVAIoVAITz31FJ166qmkaZro8yGvW7fbTTfccAPt3buXQqFQA+m79QJBIhFDJBLCvHlz6OyzhwlBwGo1k9VqJpvNQpoGatOmFc2Z829KChCp8Ts8LiJVh0BR01GzD0DRL4+YiRMlN7lly5ZR7969yeFwiMqFvMkCoNtuu422b99O6Xnhci8G7ua2Y8cOGj9+PLlcLtHFkYUMLvByxRVX0Lp168jn84kgRDkYsSHrgBIWfh0kF9Dyer0Ih8NYvHgxnXfeeaKiJsevcEyMw+GgHj160OzZsymRSKSsbyIS1Q6j0TBisQgCAR/KyvbT44//idxup7AOaBrI6bSL19dddy1t2bKJQqEA4vGoCG6UG5kxqewaRU1FzT4ARb8cSiQSKZq+vAFv3LiRbr31VsrOzk6J3GbG3qdPH/roo4/I5/Ol1J4PBoOHmf7D4TBmzpxJvXv3Tsnt5qhwANSqVSv6y1/+Qvv376cjjVV+rVwIvy6KxWJYt24dTZgwgfLy8kS8gNywi9fRU089Rdu3byei+nLdRKlNwIgIgYAPRHEsXryQTj+9PxkMGtnt1kMCsFNYC3r16kFTprxOBw7sI7nMsTw2/v9IBcUUKTpe1OwDUPTLIbnSGr8n+0Y9Hg/eeecdkVrIm7DT6RQNXyZNmkScgcACBqcvhkIhISCEw2Fs27aNrrjiCmEp4GOZzWbRSW748OH0zTffUCAQSLE+8HjTazEo+uUSr6O6ujq8+OKLomun2+0WtQbYOuB0Oqlfv360YsUK4rUntzPmdc3H9Pv9iEbDeOCB+yk3N5ssFhNZrWYyGDRRg8Bg0OjMMwfR2rVrSE5RlBsiycKqvFZlYUSRouNJzT4ARb8skgP8eIOTW8BGo1GsWrWKTjvtNLJarcKVYDQaBSMfM2YMbdiwgWTBQGbaPp9PbJjRaBTPPfccdezYUdQ0YCuEXG3u0UcfpbKyMpLdCHIuenPPm6LjT6FQCEuXLqXRo0cLt5NcaIvXYvv27enhhx8WPTyYIad3IOS1WVdXh7Vr19Lw4UPJZDKQwaCJVslsLcjNzaYHH/w9VVSUE2cncApjfV2D1KJhRMk1qgRWRU1JzT4ARb8u4g12165ddNNNN1FGRoZg4iwc6LpOp556Kr322muUSCREbQEiOkyz57St7777jq6/frxoQsPama4n/bgWi4nOOON0euutKRSN1keKp+aTp461oaqOipqPuE2xzDRlaw8H+cXjUcTjUUQiISQSMcRiEWzevJFuu+0Wys7OFI2LTCYDGY3JSoQmk0FUJPzmm69JrkTIhYj4bzDoF66CcDiIJ598glq2LBTdPlm4ZUGjT58+9Pnnn5MqQ6zoZKBmH4CiXx9VV1cLs+vjjz8uSru6XC4R8c3Nle69915R1pi1fI5VkCsjxuPJTfzdd/9F2dmZZDYbKTOzPsecNbfc3GwaP34sHTxYRrFYRGhryf/ry9rKgV1q8z4xiZt01cecxFN6DbDg9/rr/6AWLfJF5UD27es6hBBps1nolVdeokDAB4+nVhyDjxeNhoXQwbRjxw80cuS5ZLNZxPriRkacPnvNNddQdXW16kug6KShZh+Aol8PhUKhw2rLh8NhfPzxx9S7d+8UYYA1LbfbTeedd57w7fLv5f72RFzwKKnZffXVYho27Cyx8dvtVjIYNFFBDgCdfnp/+uST6RQM+hGLRYT5Vq52yGNWkd8nBkUikRTGKn8WCATAFiCmJUsW0Zgxl5DNZjnkTrKJ++902slg0KigII/GjbuO1q37jmRrQDQaThEOWMAIBv2ora3Ge++9Q927d02xSrGrymKxUElJCT333HOU7npQvQkUnejU7ANQ9OsjNgPz67q6Oqxfv150XNQ0TaQrctBg586d6f/+7/+IqD64izfW+jiBOHw+jygn+9prr1JRUQvh27XZLGQyGUSOeEaGiy6++EJav34dcaBXeulb/l8VkDnxiBtp1Weq+MEa/C233ER5eTkiBZCzAHgdGI06tWpVRK+++jJxWWKZ8ctZAfx+LBbBDz9spRtvvF6sK645YDTWWwdGjBhBCxcuJLnOAK9XJRAoOtGp2Qeg6NdL6RXm/H4/7rzzzpQocKPRKCwHZrOZJk2aRERJF4R8rGRud0z4ksPhIOLxKBYtWkBt2rQ6lLJoOHRMXQSD2WwWys7OpBkzPhbpkHxMn8+nSsyeYMQNu+T3OLefKI4vv/yC2rZtnVJKWCYWEq688nJixr9v3x5Kb2KUSMTg93uFyyAUCmDDhu+puLiNVPK4vruhy5Vcr7fccgvJxbKISDXsUnRSUbMPQNGvh5jBysFiHN3N1QuJCK+88gp17txZlEc2GAxkMplI13Wy2Wx02WWX0bp164go1SXB8QG8uXNRmIqKcvrf/72TcnKyyOm0pwQhctpYstTs9bRx40YiIlRVVYlxy24KRc1HshVHFthqamqwfPlyuuqqK4RFgF0GmpaMJ7FazeR2O0WNAI4VqKmpQnrPAg4ijEbDSCRi8Hrr8Oijj4haAyaTgcxmIzmddhGwWFzcht58802qra1NWefpwqsiRSc6NfsAFP16SDbHHymav6amBkSEL774ggYOHCgsBNxBjovLlJSU0NSpU1M6yfEmLveyZ02vtrYaU6e+R126dCZNS5qRMzJcgnHwcVu0aEGPPfYYcUMm5TY4sSi9QdCOHTvowQcfpIKCAsGsOXjQ4bClxBPcfff/0u7dOyndRcBZBHKaIK+blSu/oYsuuoAyMlxksZiEpUG2Og0ZMphWrkzWwuBxyUGr6WWVFSk6kanZB6Do10Xyhh6NRkVsgfw+b6A7d+6k7t27i9Qu9tnquk6aplFmZia9+uqrVF+3oD4dTbYYyBHpdXU1uOKKy6QqiAah/WmaRjabjYxGI1100UXCasACgqLmJ9kUP2PGDOrXr5+oZskZAGwt4MDS3NxsWrRoAbHA6PN5xPqoLxxUT9FoGOFwEJ9//hm1a9dWrBOjUU+pPWAyGeiqq64gj6f2ULvk+h4evL5V/ICik42afQCKFMkk1yQgStYz+P3vf082my2lnz3XqHc6nXT99dfT1q1bSRYImJgBhEIBof1VVVXQ//3fM8JqwDEGuq6TyWQiTdPIYrFQx44d6emnnxYlblkTlEvPpmcm1Lsywg0KPIrqSWaacttrjtIPh8MpbiaipID29ddf0x133EF2u10EohqNRiEUcPfBnJwseuihB2jv3lKKxSKiSVG60ChbBqLRMMrK9tNDDz0g4gTkQkQcrJqdnUkvvfQCyUGIzT2fihQ1BjX7ABQpaojk1rYbNmygt956KyXoUK6AaLFY6KyzzqI1a76lQMAnNntOUZS1QVk7XL16FV122aWkackGNVzRjtvk8rEvuugiWrVqFfG45EJKcpU7OS6CKZFIKNOxROlCn0xHmj+58+af/vQnateuHem6Lu6TyWQSBbCYkQ8Y0I/mzv2SuEFRujUgHo+Csw7YleD3e7F69So655zhZDTqwlVgNhvJbreKokclJe1o2bKvqLq6UqylpPDZ/POrSNHPpWYfgCJFTOkxBzKDiEQi+OKLLygnJ0eUp7VYLKTruigWU1CQR//+9+fk9dahrq4GrPlxRkIw6Bfaohxo9uSTT1BWVgZlZmYSt2U2m83CXcGtmf/yl79QaWkpyUyNrQGyYJB+PcqEnEosTHHWAPvf0wM6OYg0FArhiy++EG2znU6nSF3le8T+/YwMFz388EPE958ZP7sOEolYSv0BFg5isQhmz54lCl5x5oKuJ0nTkgLH5ZePIb/fK47LmS6c0qpI0clOzT4ARYqY0uu8y1kK3EthyZIlNHLkSGElYEbOlRBtNgs99tj/o9raaiEENOQ3ls3HXm8dvv56KQ0fPlzEFPDxs7KyhBZqs9loxIgR9OWXX1JNTY1gYtzqVm7JnC7UyIFnv1ZKT0Ft6DvBYBChUEjEDmzYsIHGjh1LhYWFKfeZU1a5oVF2djZddNEFtHDhfGJhUBYIGloDLBzs2bObfvvbceR2O0XvAl5LnGFQWFhAf/7zZAqFAkKg5GPWH7/551iRop9LzT4ARYpkkqvVpTeGCYVCCIfDqKysxCWXXCKyEurjDHTh8z333LPpwIF9FAz6haWAhYBAwIfa2mrI/uREIoaqqir87W9/E8GMrJGmBzk6nU66//776cCBAyQHS8qNmhq6ruae2xOF0gv4cNdBrlhIlBQOnnjiCcrLyxMCmqZpZDabhZuH37fb7fT8889TTU0VEolYikAgxwvIlgP+O2fOv+nMMweJNFVeP3KgYo8e3WjJkkUUDPqFq4EzXQIBn7IUKPpFUbMPQJEimY5kKWAKBALgJkn33nsv5eTkkMFgOJSyyHEAJrLZLNSqVREtWPAfOlKUOTOH+uY2yQp5c+bMoSFDhpDVaiWTySR8106nkwwGA9lsyXK5gwcPpoULF5Lf70/RguXaCbL1QFGqNSgSiSAQCIjXsVgMwWAQc+bMocGDB4vKllarVbTaZsHAZDJRbm4uXXzxxbRy5UpKHj+pvcvNrjjNUL7/1dWV8Pk8eOedt8nhsInUVC6HbTTqoiTyiBHn0Natm4l/yzErXChLTmls7rlVpKgxqNkHoEiRTDIDjUajQnuMRqMpjJdN80899RTl5uaKgDCDQRMEgFq1KqK//OVxikbDCAb9KRokByLWb/b1596yZQtNnDhRCAB2u13EGyQj3c3i9XXXXUfr16+n2tpaECW1XDlnXcUU1FND95AomVWwatUquvHGG8npdAqBTK5GKLtx2rZtS6+99hqx68bj8SC9+2UoFEiJJWDhb9u2LXTttVeLWAE5w4CzDAoK8uiRRybR/v17KRDwIRIJibUjBy6mCiHNP7+KFP1cavYBKFIkk8yY01vkEtULCkTJngnRaBTLli2j9u3bi+53nE/ucNhI10GZmW667LJLqaamSgSGycIBb+qcBkeUjAGIRCJYtmwZdevWLcWfbbfbRTElm81GFouFWrVqRVOmTBGCgWwpUPXu64ndK+wOIkrWo/jb3/5G2dnZonKlruuikqXBYCCLxSLiCG6++Waqrq5GKBRKC06MCy1etgYlEjFhLfj666V0yimdRACh3OGQ+xg4nXZ68cXnKb3IUUMpjPVWAiUUKPplULMPQJGiYyXZXy/HHtTU1GDLli3C5CwzcE4rNBqNNHLkSFq7di2lxyzI6Y9E9VHwnEe/d+9euuSSSygnJ4c0TSODwUButztFc2X3xeWXj6Fly76i9Fa7LHjIpmw5AC753eaf459D6Ro0/y8zVhbG4vEoamur8ckn0+nMMweJmAG55wXHDrDVoEOHDjRr1iyqq6sTgqEcwMnnDAb9KdUtfT4PPJ5aPPnkEym9CtiyZDTqZDTqZDabqWfPnvTZZ58RUdKqwemQqtS1ol8LNfsAFCk6VkoXCuTP/H4/Nm7cSOPGjRNMxGQyUVZWFmmaRpqWLE7UqVMnYXYmqt/s08vner1eITRUVFQQEeGDDz6gU089VTAwjoLngEQOVCsqakEPPHA/7dmzm8LhoGCKsgYbCgWEhln/fvPP8c+jpPDDhYI4yDO9WiBRHMuWfUWjRo0QwaFJ5mxMcdXwX6fTSU888QRt3bqV0tcC3yOuaCkLYpFICLFYBHv27KbzzhtJTqedsrIyRAChwaAJ94HVaqaxY8eKc1RWVorr4vbHihT9GqjZB6BI0bFSujshXTAIhULwer24+eabCQBlZGSIjAEORrRYLFRUVER/+tOfUjIH+NiyYJB+vkQigd27d9NVV11FBoNBMC1mXFzght0Y/fv3pQUL/kPMDLkNL1dZZC22PhCy+ef455BcECi9UFQsFgGn8917791kt1vJ6bSL3gQGg0ZWq1UUp2KLQdeuXenrr78mvt/yPUtPYZVTEVkQmzdvjihlzXEmybXhIl1P3itdBz3yyCTh+uFj/9haU6Tol0rNPgBFin4OcYXBRCIhtP6qqio899xz5HQ6yWQykdVqJbvdTkajkXRdF5r+hAkTaMeOHcRxChwc2FDGgJxuGI1G8eyzz1JJSQlZrVbSdV1q72wkTYMoggOA7rprAm3ZsolkTTkY9AtLwS8hUC0pQNULA1w0Kh6Pwu/3oqqqgt599180aNBAMU/sy2chiu9LRkYGtW7dmh5//HGqra0VvSfSqyFGIpGU+y8LJJWVB2ny5D9Sfn4u5efnpvS64JRDk8lAHTuW0GuvvUpckZAFD5U1oujXSs0+AEWKfgrJsQRM8mtmGsFgEJ999hmdcsopKRHsbrc7hQENGzaM5syZQ3wc/n16WiFRkinFYjFRWGft2rV01VVXUXZ2tghsZFM0+665s1737l3piSf+TFwwhyvupfrdm39+fyqxBSWZJhoR1QHZbRCNhjFz5gw699yzRaMiu92aEtzH88YWl0svvZQWLVpEwWAwxZLDLh4mHgPfF3bLrF69ioYNO0u0TebzsPDBDbBOP70/LV++TDRKkhtfsRWChYTmnmdFipqKmn0AihT9VGpIMCCqFwgCgYAIEPvqq68oNzc3JZXQZrOlmP5LSkpo1qxZdKTGRVxch6heAAkEAkJzfeaZZ6ioqIiAZBOeegZnF6Vy+b1zzz2b1q37joiStfZl7ba55/WnkiwQyJo6Z3hUVh6k++67J6UMcTKgr75xEc+Rw2GjvLw8evbZZ4mDB+ViRuluHWbY4XAYwWDwULZICEuWLKLCwgKyWEzkdjuFEMB9C9hSMGbMJSQ3QeJAz0AgkBK8mBqz0PxzrkjR8aZmH4AiRf8NMVOo11KTm7bcWpcLHa1fv56GDRuWkvfOqW5c0ljXdfrLX/5C5eXlJDcxkjsiyufkz4iSEfALFy6kMWMuEUFszOyYGdpsFtL1pMaan59L998/kfbt20PxeFSKzm/+eT1WShcI2H0QCPhQWrqLnnvuWSopaScsA+zTdzrtwsXCqaOFhQX0m99cQxs3biQWvvx+f4oARnR4aWS55sH+/fvpj398lDIyXCJ4kI9vtZpF58Tc3Gx66KEH6MCBfSRngSTjIeqPJ3doVKTo10TNPgBFio6VjqatyWZ+olST8549e2js2LEitsBsNov4As5UyMnJobFjxxJHntfU1IAotcVvQ2Nhv3ZFRTk988zfqG3b1kL7ZSFB15NMigMRzWYjnXZab3r33X9RfZGd5p/jY6GGBIJEIoG6uhq8+urL1L9/XxHAx2l/MpPWNAgh6bTTetMHH7xPlZUHiQW69FoV6U2nWBhjAe27776jiy++WFgBdB2ihwG7LEwmA7VvX0wfffQBcaAn1zSQYzq4GJJ8nfI5FSn6pVOzD0CRouNJvLEzQxk/fjyZzWZyu90pQYdsRbDZbFRQUEDff/89hUKhFObPjEFOX5RdC6x5rl27hoYMGSxSFB0OG2kaR9jX10/gbIVzzz2bfvhhK3HXQBZC4vG48Jen11Lg/49XMJx83XKApTwXrFkHg0GsXbuWhgwZLLR0u91Kup68ZhaI+LXNZqHMTDc98sgkYcJn8326hh6NRg9rJhUKhUSDrNmzZ6fUieC4AbYMsCDSq1cP2rp1M3HPi/QyxUnBoPnXqyJFzU3NPgBFio4nyZosm6FfeuklUV+AswfsdjvZbDYhKJSUlNDbb79NREnztSwE8LHT/w8EfCnFiP72t79Sp04dUrRk9nGz1YA/KypqQZMmTaINGzYQUcOBjvI1sfDA78lllX/OXDV0Ps7OYOFETglcsGABjRs3jlq0aCHSC9k9IDNn/puVlUGjR19EX375BTFDrg+6TBU8ZHM+3wMWwGpqavDggw+Sw+Egl8t1qLqkRQhibB3Iysqgq6++kkpLd5FcXCm9QqESChQpSlKzD0CRouNJqT7vZGnkSCSCadOmiTRC7m8gt2PWdZ3y8vLovvvuIzleIZ1hBoNBqdpdktGwcBCNhrFlyyYaO/Y3ZDYbhWWAzdxGo05Wq1nk0dtsNurTpw+98cYbJLdljsfj8Pv9DeTly5H3jTNf6Q2pGtLcE4kENm/eTPfccw+1bdv2sB4FHEPAQgGb8Lt160L/+tc/qbz8APFcpXYtTJ6DBRy57DQLCKFQCOvXr6czzjiDjEajyCbhLpZslQFA2dmZ9PTTT1FNTVVKbYiGyiGfTO4bRYqOJzX7ABQpOp6UmjKXylRXr15NAwYMIJvNRjabTbRINhqNoq+BwWCgSy65hHbs2EHycUOhUIqfOfk6aQpnZhONhuH3exEOB/H88/9Hbdu2Fjny7FuXKyFyrX8AdMEFF9C6deuI6PASu/F4Mn0u3azemBSNRlOsDyyUlJWV0fPPPy/6QbAwZbFYyO12pqUX2kVw3+WXj6Ht27eRXIY4vfyzfM54PC7ml2MNfD4f3n//fSGIsEDAJa05oNDtdlJWVgZNm/Yhcb0Euay03DRJCQWKFKVSsw9AkaKmIq5DwMwmGo1i27ZtNHr0aFEGmZkyCwXM4EaMGEFz586luro6ENVrrtFoVMp4qK/5X98kp95/vWLFcho79jfUokU+mUxJH7jL5RCZClwu2Wg0ktPppIKCAvqf//kfWr9+PclMUqZ4PI702IefQyzspFsH9uzZQ2+99RadddZZwsWi6zpZLJaU9M70lMMLLzyfPv/8M+LSx+lzIpvwWWALBAJCEGIBaMeOHXT77bdThw4dxPlZIHE4HJK1x0Bnnz2MVq1aQcGgH+kug4aEgfqxNf8aVaSouanZB6BIUXOQrH2XlZXRvffeKwQCLrErMx0AVFBQQFOmTCFmXrKmnjSzJ5kepxh6PLXCPF5bWy0Y4Kuvvky5udnC1M7MjIsrcQ8Atlp069aNJk+eTLW1tcKv31AufWOSXMRp6tSpNGrUKCEEMFOWrSp2u10UagJAnTt3pBdeeI52795JrJnHYhHU1dUgEgmlCAj1vSFSr4OFrU2bNtHAgQOFu4fvk8vlkoQRI9lsFrrllpto37494px8rtR6EA0JBCdXSqgiRceLmn0AihQ1FQWDwZRgOtb22WQ9efJkyszMpIyMjJR+CSwo2Gw2MpvNdN9991F1dTWI6oMNk0w6lekwY+IqhnLfg507t9NNN90gmKjNVq9xs6VA7tcAgM466yx6//33yev1gjMV5EI+P3d+uFQ0xxLMmTOHLrroInF+ZsIWi0X0lWAmzYJMdnYmXXPNVbR69SpixhuNhg8z2YfDwcO6KiYSCXD/gXA4jHg8jqlTp5LL5SKLxSKaULEwwi4fu91O2dnZ9Nhj/49Y8OI+Ew3FETRkpVBCgSJFSWr2AShS1Fh0LBpzepMj1ojD4TDC4TDeffddateunQhAZC0YgCh2ZDAYaPTo0bRmzRoiIhERXx8wF0+JL5CZT9J3Xi8sfPLJdDrnnOHi+MyAmfmxVs7m8ZycHLrwwgvp448/Jh67XLDp585fIBDArFmzaPTo0ZSTk0MGg0EISNw7gl0tRqORrFarGOeYMZfQp59+IuoApDP+RCJ2mJWAGXVyjpLjiMfj2LZtG91www2C+fMY7Ha7+J9dLl26dKEvv/yS2GUjd6Vk+jEXQr3bp/nXsCJFzU3NPgBFik4UYm1769at1LNnT8GQWHtnTdVut5PVaqWSkhJaunQpcZQ8M6V0DVRmjCwQyCbt7du30V//+iS1aJF/mF+eUxetVrPI8TcYNMrMdNP114+n9evXkSyMyEyRz3d4lH09I+TPE4kYNm/eTFdeeSVZrVZR4CmZ/19fx8Fut4t54J4PeXl59I9//IN8Po/ofdDQ+ficcrwFdzXkmIJwOIzvvvuO+vTpQ5qmkcFgIKvVmjL3LJgAaDAIVJEiRf89NfsAFCk6UYg171gsho0bN9KZZ55JFouFzGaz6JnAZnJZe3/22Wcp2Uyn3kLAuff11QrjSGfe/F0WHObM+Tedc85wMpuN5HDYyGazpGQocKU+uZdCUVEL+t3v7qPdu3fSgQP7SGa+8rn9fq84H6dNxuNRRCIh/Oc/c+nOO++Q0vrqA/csFosw3cvZEQaDgWw2G91+++1UXV19KH0wmCIQHMmd0tD4iOI4cOAAvfjii6RpGlksFmENYOsMB2JarVbKzMyk22+/nbjqZHqGhiJFiv47avYBKFJ0IpEcxb9161a66667RKChrDGn+tGzaezYsUJrb0hLZu1d7iAYDgdTtHoWJv7xj1eoZ8/uIoKf2zDrOsjtdoriQFwd0Ww2Uteup9D/+39/oB07fqB0i4TMjINBvzDj/+c/c+n668dTmzatyGQykMPhIIPBkFICmv+azWbhRsnJyaEbb7yRli5dSsFgEKFQ6FAMQLqPPn6Y+d7rrRP/swUlFotgxYrlNGbMGCGQWK3WBgM/LRYLde7cmaZMmUJ+v18EXKpOhooUNQ41+wAUKTrRiHPyiZJ+9scff5wyMjKEEMDNlGRmZTQaqWXLQvrqq8XExYvYX+73J60IzDTj8WiKRs1NkeLxqGDgZWX76bbbbhHH5/x/uVqgxWIiq9UsiiBx1P8zz/xN+PVlawQHPG7atIGuu+5aUYKYgx1ZM5ctIWw9YIGgS5cuNH36dOK+A1yKOTl3qcIAB/M15NePRsNCQJg5cwa1adMqZQw8x2wd4KZVPXr0oKVLlxJRfYBoY1RzVKRIUZKafQCKFJ1IlN6Zz+PxgIjw7LPPUnFxMeXk5JDVahXMi+MLkpYDnQoLC+jDD6dSVVUFMQNkBp0ebZ+uVfPnbNoPBv2YMeNjuuSSi8lsNgrXgd1uTSkSxCWUMzJchxirRuedN5LeemuKSAmsqCinb775mq6/fjy53U5RepiPwb91OByUnZ0tmLCu62Q2m6m4uJiefPJJ2r9/P8lNioLBoFT6OFUgSBcG2G3B87F//1669967RYtjjh3gYkiymyI7O5suueQS2rJlCxEdHlypOhoqUtQ41OwDUKToRCG5wQ9RfYoeFw1asGABnXrqqUJzdblcogpiMkvBKroC3njj9bR580aRL59eRIcZZiQSSgm8k4UGtizs3LmdXn75Rera9ZSUaoEsKDgcNjIYNNGN0OGwiTLDw4cPpUcemURXXHEZ5eZmi0BFbk7EVQDTzfRsqi8uLqaJEyfSnj17SK6N0HCxpIaDCxui2bNn0ZAhg8lo1IXFQ9M0ysrKSqlSyGP405/+ROXl5cT3hSi1WVJjZWAoUvRrp2YfgCJFJwpxfj5RahMkomTKYigUQkVFBfXq1Uu4DJh51dfgry/gM3DgANq6dTPJwYUys284LS6e4kZIJGJCoNi/fy/dccdtouCRwaCRpiXLJet68twsGHB7Zv4ev2cyGVLaGet6MjaBhQi32y18+ZdffrnQzNMtJ0T1jLi+AuKRhQIWgIjieOmlFygvL0e4Q2RrB1sKTCaTCHZ87733iM8ZCASO2ChKkSJFP5+afQCKFJ1oJEeyc0Mi9p2Hw2F4PB7cdNNNlJOTQ7quCw2bOwFyMSKjUaf27Yvpww+n0sGDZSQ34UkkYgiHg4cJCHLxHRYS5LTCcDiIuXO/pCuvvJyyszPJYNDIZDIIJm+zWUjXkdJ4yWo1k65DjM1o1MlutwpGbDBoZDQm6w64XC4aM2YMzZ49m1goSi/yJFtP4vG4FOSX6jrg17FYBJFICN9/v5bGjv2NOK/NZhECSvI6TMKNoOs6jRo1itatW0fcFVLuXSG7C7jQUXOvG0WKfgnU7ANQpOhkITnS3ev1YtKkSaL0LzNbo1EXDJq14KysDJo06UGqqCgnZvJy1gEzUBY+GmJwLKiEQiEkEgl4PB68/fbb1KdPHxEU6HI5Upg8j4GFArYGyAGLHFtgMGh0/vnn09tvv01yVUH5/A01YJKtK2wJYHcICzWhUAAzZ86gHj26CUElyfyTQZK6DiGUmEwmcjqd9Nvf/lbUH6gvDqVIkaLjTc0+AEWKTiZiBsiMc9q0aYcYf1aKvz8ZhGgVjBoAnX/+qJROgYenJSbPIZdi5rLDRPUaO1ce5HH8+c9/FlH6fF6bzUIWi+kwa4DL5RDWBR5nt25d6IUXniOPxyMsI5FIBKFQqEEBhS0DsraeHE99tgNfU01NFSZM+B8yGnVh0ZDrL7BAknRtmMhsNtNf//pX4uMm6z9QgwKJIkWKGp+afQCKFJ0sJLsVotGoYI7/+c9/6JRTTiGr1SwyBNhKwEJBdnYmAaBevXrQhx9OpVAocFj9/UgkIo4rM+P0wjyJRAJ+v1+Y8D0eDzZt2kT33ns3tWvXNiVYkKsjylUS2XTfr99p9Oyzf6c9e3aTLJSk5/yz2T4Wi4FN+SwQhMNhwbgjkVBKGubnn39G5503UsQysAWFKzLyODgu4swzz6TPPvuMwuEw/H5/yjgaqwukIkWKfpyafQCKFJ1sxEyQiIQm/91339HAgQPI4bCJmgGsiddH85tEtP99991D1dWVYDN7eu19mSEyA2a/evp4uB10KBTAkiWL6Prrx4tgRLYMsIneYjFRnz6n0jPP/I1KS3dRJBISBY18Pl+D5yWqZ8psoeCSxPw9bgjF6Y/PPvt3at++mCwWU4q1pL4mgibcCBaLia655iratm0byTUHWEhq7vutSNGviZp9AIoUnUwkp74dngYXx9ixvyFNQ0p0vd1uJbvdepjZfujQIbRu3XfEzDQajaZYBWTzfboZnzX2cDgsLAtsuo/FIpg27UMaNGigcB9YrWZq3bolPfroI7Rz53ZKD2yULQVESQuAnH2Rfn62HMiBhpFICLt376QLLjgvrQhSMuiS6xFoGoTwlJnppgcf/D1VV1cK9wjR4ZYBeSyKFCk6fqQRERQUFI6OWCwGo9Eo/gIQ5n6TyQSTyYDy8nJ699138cQTTyDZEwAwmQyIRuMpx7LbrYjH42jbti3uvvtujB8/XrPbnYL58vH5XJFIBGazWfj8rVYrNE1LG2GywqDBYAAA1NbWYtq0abRy5Uq0bt0aV155JTp16qQBQDAYhM1mE1q+zWYTYzQYDNB1XRyVN4tYLAaz2Qy/3w+HwwEAYlzl5eX09ttv4cUXX0RZWZk4RiQSQSQSg9vthMfjg9GoQ9M0RKNxDBw4APfeey8uv/xyLZFIQNeN8Pv9MJlMMJvNCIfD0DQNZrO58W+mgoJCw2huqUSRopOR5Hx51mK5tkA0Gsbs2bOoV68eIogOh0z5mlbvR+cGRwUFeXTTTTfQDz/8QESpLZ1l7Tm9nK9sSUj6+WPweutEqmMkEoLf7xXlhBOJmCh1zFkB9fEMoQavM91lIcc7cM2C3bt307XXXkvZ2ZmHNW2yWs3CQsBzYTBodNlll9LatWsoFArA7/c22LpYtk6olENFipqGmn0AihSdLMSMWvZzy1HxzIjZJL9q1QoaNuyslIA6ucgQpzACSUZ69tln0/fff0/hcPiwjAO5g2M4HE7x9ycSiZTiQekMPxDwpRQPisUiKSmR9d8l6VpSX8disRR3CcdVfPTRR5SXl0e6rqcEEzocNnFtFotJXHteXg7deecdIp4itZV0fXaFXI9AuQ4UKWo6Uu4DBYVGAlGSeRuNRui6jng8jn379tFrr72GJ554Qpjk2d2QSCRgMpkQCoXgdDoRCgVQWFiIW2+9FXfddZfG5n2j0Yh4PA6j0YhQKASr1SqOQ0SSeV0/wsjqx3e4y6Ee8XhUuB7YvQAA4XAYFotFnFPTNHz77bf0xhtv4LPPPsOBA+WwWEwwGEwIBAIwmUzQNE3MBwCYTCZ069YNDz/8MC6++GJN0zQYDAZompbi8lBQUGheKKFAQeE4Iekn17F9+3ZasmQJ7rzzTlGciJlhPB4XsQJGo45YLAGbzYJRo0bhtddeQ05OjhaJRGAymYRgACBFOKiPcfhxoeBoCIeD0HUdRqNRjC8ajcJqtSIUCsFoNMJgMOCzzz6jhx9+GOvWrYfJlGTmHI+QlZWFcDiMQCAAXdeRkZGB2tpanHnmmXjllVfQsWNHjWMk5LgMFjoUFBSaF0ooUFBoJMhBgmz2ZsYdCoXwxRdf0P3334/du3enWBQSiQRsNhsikRB0XUcsFoPBYEBGRgaeeuopXHfddRpbGerq6mCz2VKC7+o1+Z8nFAAJYcUIBAIwm80pgZU7duygSZMm4aOPPgIRweFwHMpSiMFut0LTDPD7/dA0DRaLBQaDAYlEAg888ADuu+8+jYMT4/E4dF1PsSaoYEIFhRMDSihQUGhEyKZw4aPTNMRiMZhMJixbtoz+7//+TzBWXdeFIAAAug4QJclut8Jut+Pqq6/Gvffei3bt2qXY/kOhkGDASfw8oUB2HwDJzAJN0+DxeOjjjz/GCy+8gG3btonrjEbjsFhMsFgs8Hh84loMBgMikQg6d+6MF154AWeccYaWFHoi4nOg3pLyYy4NBQWFpoUSChQUGhGJRAIAUlL6GKzR19TU4JFHHqHXX38dZrMZXq8XFovlkEBAwq3gdCbT+EwmA1q3bo133nkHffr00dikzxYJtjbouvEnj1d+/oniQkDx+/3IyMjAvn37aOLEifj4449TNPxoNJ6SXmixmJBIAEajEcFgEJdccgnefPNNEReRvL7UNMdEIiEEBLZQKCgoNC9+rr1RQUFBAlsHOCOASwMDgNlsRiKRgNPpxLPPPqtNnjwZubm5MJlMiEQiCAbDh+oeJGMHPB4fbDYLotE4SktLMXr0aLzxxhtUXV1NMjMF6oWRnwNm2mzteOutt+iss87C1KkfingGjjkwGLQUrZ9/l5eXh8mTJ+Ott97SnE4nLBZLSuAlp1FyoGH6uRUUFJoZzZ3+oEjRL4XkNEH5PRYM+D2uN5BIJPDOO+9Q//79yWBItjg2GDTKyckSZYkBUEaGS6QtGo06nXfeSJoz59/EqY8NpRT+lDEzVVVVEFEcK1Ysp/PPHyWqMfJ5Od2Quy7y/9xxceDAgfTFF1+kpFRyGmNdXd1h545Go6LrY3PfO0WKFCVJuQ8UFBoJcjwBZwywv5zN/PwdoiTDdDqdiEQiuOaaa2j69OkwGJLBd8kI/RjMZmP9w6pp4hg2mw0PPvggHnzwQQ1g8/tPj+Dn55+IoOvAY489Rs8//zzq6uoOVSSMwmg0IB6vr3YYi8Wh6xpsNht8vgAsFhMuvvhivPTSK8jNzdVisZiIL5AhpzbyfPC5OeZCQUGheaGEAgWFJoTM4DVNE9ULY7EYXnzxeXr88ccRCiWzEILBMIxG/RDD1kWaYCJB0LRkMOJ5543Cww8/jFNPPVWzWJJ1BepjDJK/DYfDsFqt4txyvQSj0Yjq6mqsWLGC/vKXx7F48VeHnZOZPL/HaZTxeBzFxcWYOHEibr75Zo29kSpwUEHh5IUSChQUmhDpQgEA0VwoGg1j6dKl9NBDD2Hjxo0iVS8YDMNsNgpfPDN9q9WKQCCAli1b4pZbbsH//M8ELSsrC6FQCBaL5ZAAkTh0jOCh/gZRkWaoaRp27txJL7/8Mt577z34fB6Ew2HRa4HDFEwmg8goSAYiJmAwaDj//PPx8MMPo3///odqKSStAEooUFA4eaGEAgWFJkRDQgHA5vTk/xs3bqS77roLixcvRiQSg9Nph88XgMvlgNfrh8ViEgzfZDIhHI7CYNDQo0cvvP322+jRo4fGtQWi0aiI/mehIhwOQ9d1TJs2je69914cPHgQRqMRnH0QjcZhtSaDInmcoVAEBoMGk8kEu92OiRMn4u6779YsFguI6JBLJHk9SihQUDh5oYQCBYUmBj9z6cwzFArAaDTCaDSiqqqKXnjhBbz11lvYtatUVDvMzHSLvgNy3EHScmCH1WrFDTfcgFtvvRVt2rTRuGoguxN8Ph9WrVpFTzzxBObPnw+r1Sr8+eFwEJFIsl5CRoYLdXVeMTaXywG/349zzz0X9913H4YNGyaOzV0TlVCgoHDyQwkFCgpNDPmZYwaatB6QYN6s6c+fP59uvfVW7Ny5U3w3FksIv7/BYIDT6UR1da2oEGgymXDWWWfhnnvuwahRozTW9nft2kUvvfQSXn/9dcRiMcRiMQSDQeFOMJvrqxcmMyaAzEw3PB4PrFYr/ud//gcTJ05EXl6exnEKchtpJRQoKJz8UEKBgkIzQI76r0d9ISKO9o/H4/B6vTR+/HgsXrwYgUAQZrMJkUgUmgZhLbDbrQiHo7DZbNA0DV5vUsu/9957ceedd2LBggWYNGkSysrKYLVaRQvkjIwMRKNRBAIBAIDDYQMRIRAIweGwIRAIolWrlnj++ecxevRoLRgMwmQyCUEASAYsJpsbJTMvlFCgoHDyQgkFCgrNhPRnj336Pp8PTqcTAFKaJj3zzDP00ksvYefOnbBYLAiFQjAYDIfiC5LHYKHC6XQeymAIIicnB36/H4FAQPQj4OBDIhINjywW0yHBwoJgMIz8/FzceuutuPHGG9G2bVuturqasrOzNdn9wZaC5Hsq+0BB4WSHEgoUFJoIHGB45M/jIu0vHA6LOgccHGg2m7Fq1Sp69NFH8cUXX0LTAE1L1g4wGAwwGs3w+Xyw2+1IJBKisyELEHx+LhZkMpkO1SKIwGKxIBIJweVKxhJ07FiC559/HkOHDtUMBoOoOyDXK+Dj1Pd6UO4DBYWTHUooUFA4ScDVEX0+H5599ln6+9//jlAoBLvdfqhyYEIICdFoHCZT8q/TaUc0GkU4HIWu45CpXxNuB3YlWK12xONx3H777fjLX/6imc1mEYQoF2ZSUFD45UIJBQoKJwnkioDhcBhfffUVPfroo1i6dKnoR5D8LAqTyXCoOVFY/N5ut4ryw5x2GApFxN/Bgwdj0qRJGDZsmGYymZBIJGA0GkV2gYKCwi8fSihQUDiJwEWHgKTGv3PnTnrqqafwr3/9C8GgXxQXstuTcQJcRhkAQqEILJZkkKLdbhM1DGw2GyZMmIA77rgTLVu21CKRCMLhMFwuF4BUYURBQeGXDSUUKCicJGDmzKZ8n88Hq9UKg8GAf/zjH/TXvz6JiooK0WQoFkuIuAN+zG02K4LBEIiSmQbdu3fHpEmTMGrUKM1ksqSkGIZCIWiapgQCBYVfEZRQoKBwEoH7EACp7Yaj0Si2bt1Kjz32GGbNmoVgMAiHw4F4PI5oNCpqHxARHA4HnE4n7rjjDtx2223Iz8/XAIjARDndEIBIkVQNixQUfvlQQoGCwkkCTlWUg/68Xi+cTqdorGS1WjFjxgy6++67UV5enuIiYAFh2LBhmDJlCtq2batxh8JkoKEVQDJLIpFIiH4HRqPxsI6HCgoKv0yoJ11B4SSB0+lEMBgUAkE8HofL5RJpgVarFR6PB5deeqn20Ucf4dJLL0VeXh50XUcoFELHjh3x3HPP4T//+Y9WWFioJRIJBINB6LouBALuj5CsaJisZ6DrOjweT3NeuoKCQhNBWQoUFE4iRCIRmM1mAEmXAZv02UXA9QN0XUddXR2mT59OU6dORa9evXDTTTehffv2GtcrkPsisGDBVgj+GwqFhMCgoKDwy4cSChQUfiE42rMsF09SBYYUFBQagnIfKCj8SqAEAQUFhaPBePSvKCgonBxQTF9BQeHnQQkFCgq/EBzNEKA8hQoKCkeDEgoUFH4lUN4DBQWFo0HFFCgoKCgoKCgAUEKBgoKCgoKCwiEooUBBQUHFGygoKABQQoGCggJUvIGCgkISSihQUFBQUFBQAKCEAgUFBQUFBYVDUEKBgoKCgoKCAgAlFCgoKCgoKCgcghIKFBQUFBQUFAAooUBBQUFBQUHhEJRQoKCgoKCgoADgkFAQjUYB1PdjTyQSSCQSh/3/UxAMBsUxiQjxeBzxeFx8nn7c/+YcPxV8fUSEWCyWcs5YLJby3UQigVgshng8jkQiIa5D/j7PW3MjfWyMWCwGv98PoP56jnWej3TMEwHyvMvr9Eg4Ua5FHkM8Hk9Zc7zOIpFIym/4vvF3GroOIhLXz88dzxE/c011/XxeHg+/lvcYfp4aQigUSnktz8ex3ke+5vRjRaPRlOf5eCAajYpr5/smnysSiYhxAEA4HG70MaTvrfL5eU7SPz/afBBRyr2Vxy2vY97rmxPyPi8/M/I1JhKJlHHKn/F8pL+X/j35s6bgX4z0ccv7g/x++vjT5+JIvF2LRqOIxWKwWq0pByYiGI31/ZL8fj8cDsdRBxyJRGA2mwEkFwsRwWQypXwnGo2K92KxWMr5jydisVjKNSUSCei6njImXuwWi+WIx9A0DQaDAUD9TdF1XbzX1JBvfjAYhNlshqZp0HUdWgNVaXjB8Gf8NxqNit/xvDDi8XizXd+RxiE/BD82Np6fhuaiqREKhVLWeiAQgM1mSxkbrykiEs+S/Dt5E5LvcTQahdFoTDkWP4/yWj9ekO8Nny/9eWkIvB7l8fHzKB+Hj3Gk+yjvUfL1sjBsNptT9qdAIACr1dpo8xKJRGA0GqHruni+IpEIiAgWiwXhcBgmk0mcj+eLN2d5b/pvEI/HD3vmWQCxWCzw+/1wOp3iu/Ief6zPdyAQgNlsFr8Lh8Pif4PBkLKvhEIhWCwWaJrWJOuPwUKffD18fQ3tBTzmn7JPyIID75nHG+nrR94TwuEwLBbLYfcVQIoQ3NA9Tn/+NNk6EI/HBbNOfyiPBfzAsUTJxwqFQtB1XTyMDSEYDMJmsx3TeX4O/H4/iEg8HOmCgoxIJCI+T99YG1p4zYUjLWYiQiQSSRFwjsRQ0sEM98e+09RIZ4zpzM9kMh1xrOnfby7U1dUhIyMDQKoA7fV6oes6jEZjyv3y+XxwOp3iYW9IYANSH2wWUnmNpgvlxxO8bgwGg2DkvOHyPTjSfZCfL97EZcaZLsQe6fy8FoxGo/idpmnC8mc0Gg97bgOBAOx2+8++ft4309cb7yMsuLHmZjKZfnQP+m8RDodhMBgOO248Hhfjaoi5/BhkZY6FAXkeeT2zNYa/29A+dDwgP0/pSFc+Wbvm+ZGFZ1638rXJ185oDmWDLYw/Npfpa09+zRYfFkIbWncaP0D8QPJfr9cLl8slDniswgF/nwfPC1N+X5bMjqZFNCbSJWnePHVdT2Eq6YsaqF808m/4eoHm1UJlk1AwGITD4UiR/FliP9o406/tRAKvFd5UY7GYuD+sIZ4I9+JokBmAPO70MfNG3dBaBFJNxPzbeDwuNsX057WhTa0xIbsJ08/Dzx2PVRZM5Wc/XWORNXz+nI/RENK/LysawWAQiURCzDePiZ+XxkC60MoWH1mhkDVWeQ9sDKH1SIIIj42IYLPZIO/5zGCO5fzy9cn7i9frhd1uP2wf57Wbbr1qChzJ+sL3XraQsKYvPzOyu4TXU3NbS2XBVebL4XAYNpstZfyRSEQ8C7Lwx5YbBgtH8r6vRaNRRCIRcTJ5o2Kmruv6MW0q/NCyKYMf8EgkAoPBgFAo1KB573hvWDJkH0o0Gk2RptnUma6RcDwEEYmHQrasEBHC4XCTuEAaQvpmyZtBJBJBJBJJsYrwzU/37bKGma5ty76p5hIWeAzRaBQWi0XMvXwt8vyfqEIBa62JRAJ+vx9utxsAUFtbi8zMTASDQRiNRphMJng8HjidTui6Lp6naDQqrrkhqxBvhCaTCX6/HyaT6Uetc40Fvj+BQECsNbaIsHbO901eQzIjkt0E6dqbPHfA0QVbVkKICKFQSLjTmPnKwhOjMfYgWXEADtdCj/RXvv6fi3Rtma+f93BN01IYJc9HOrP7sTnmtRaJRITQxecNhUJCy+Y5ZR7QFGuRx8fjZ2tvIpEQDFEWUIH6uB1ZWEtXjmSFkvdN+TlsKoGBx8pWRdkSxnOfLoSxlYSfQV5r6eMVlrxoNIodO3ZQy5YttUAgQC6XSwuHw6ipqaGCggLNZrMJE+axgn0f27dvp5YtW2o8oSaTCbt37yan04n8/HyNJXmDwSA2vuMJFgLSmR5P7J49e6isrAxlZWXweDwIhUKIxWKw2WzIyMhAUVERioqKkJ+fr7Ew0Nhmv/8GzLS9Xi/8fj/FYjHk5uZq0WgUfr+fnE6nFgqFyGg0aukbK1tqDAaDEHDSfWXNKR3L49m/fz8ZjUYUFhZqQP0DHw6Hoes6XC7XMTGO5kK6NW779u2Ul5fH10JWq1Wrq6sjm82mGY1GBAIBkq1sDE16QVJAidVq1fx+P7ndbo01BxagjtezJd8fn8+H6upqatWqlRYOh5FIJGAwGBCJRFBeXk4tWrTQbDZbg2NhZs6bq8ykDxw4QIWFhcLV+WMur0QigVAohL1791KXLl20uro6mEwmscf88MMP1L17d40ZmNVqTXFz/Fywhma1WlFdXY1QKESFhYUax4h4vV54vV4qLi7WZGGvsZQi2fLA8x+Px7Fr1y7Kzc3V+JkOh8OIx+MUi8VQVFSk8W8ZDc0FXwMznUgkgtraWtI0DW63W6upqaGcnBytsrKSAoEA2rRpo5lMpiYTCI6FMft8PpSWltKuXbuwf/9+eDweweBdLhcyMzNRWFiIli1bIicnR7Pb7ULxkIWEhmLLmmLPkddKMBiExWKBruuoqqpCVVUVbd++HTt37oTP5xP3iq8rPz8f+fn5aNmypcYuzGg0ephb0rhhwwaaMWMGFi9eTC6XC3a7naqqqmAwGDBy5Ei68cYbtZ+yYOvq6mC327F792569tlnsWPHDsrMzITP5xN+pSuuuALXXnut8JUCaBLmyhceDocRjUbhdDpBRJgzZw7NmzcPS5YsgdPpRGFhIdq3b49WrVrBbDZj//79WLVqFTZt2gQAOPXUU+mqq67CgAEDNCD15jQnli9fTh9//DH27t2LWCxGmqYhOzsb8XicotEo7HY7yfEEvNg52NThcCA3NxcdO3ZEjx490LlzZ+Tm5mpA00jCR3uwvvnmG7zwwgsoLCwkh8OBuro64TO/6aabMGzYsBNPEpDAm24wGERVVRU9+eST2LVrF9lsNhxiolRcXIyysjLizYg39VAoJLSEcDhMctyOxWKBLESUlJTQ5MmTtczMzEZ1Bx3t/kydOpXmzp0Lv99PBQUFqK2tBZAUhoqLi/GHP/yB4vG41pDWn26+ZWzYsIHmz5+P22+//ajrj4XbTZs20VtvvYWNGzeSyWRCRkYG6urq4PP5cM4556BVq1aUn5+vyW6JxtrQ4/E4bDYbgsEgVqxYQVOmTIHNZiO2pBgMBlx++eUoLCw8zC3yc5EuXPA1VVVV0Zo1a/DKK69Qfn4+ampqhBBkNBrx97//XSiAPwYWUHkfj8fjmDVrFj799FNEo1FyuVwIhULk8XgwYMAA3HbbbVRcXKyxpfV47x/y8eWYCn7e3n//fSxYsAC7du1Ct27dMGDAAGRmZiIcDoOIsHnzZuzatQulpaUwmUw49dRTadiwYRgwYABat26tBYNBWK1WIQwcTYg6HuAYFF5n4XAY//73v2n69OlYsWIF2rZtix49eqBDhw6w2+2oqanBtm3bUFlZiXXr1gEACgoKaODAgRg1ahR69+6tsStJPJdEBI/HgwceeIBatmxJVquVioqKqKCggDIzM2nq1KnEpiJZK2iI5NS9cDiMYDCISy+9lABQcXExde3alb766itirYK/GwwGj3rsxiJeoLFYDF9//TVdeeWVlJ+fT926daOpU6fSzp07Sb4GeYyhUAivv/469ejRg7KysuiGG26gdevWET8oTXUNDc0736P58+dTdnY2tW7dmrKysqht27ZUWlpKNTU1qKysRHl5OR04cIDKyspo//799MMPP9CKFSto3rx59NRTT9FVV11FhYWFpOs6de7cmX73u9/R119/TU1xffL6kd/j9M9wOIxnn32WHA4HFRUVUVZWFhUVFdGcOXNIvrcNHedEIda24vE49u7dS4MGDaKCggLKzs6mtm3b0v79+6mmpgZlZWV08OBBKi8vp/3799O+ffvohx9+oE2bNtHy5ctp2rRpNHnyZLrsssuoU6dO5HQ6qUWLFpSZmUlFRUVUWlpKrAXI5z2e9ycWi+H1118X9yYvL4/y8/Np8ODBRESQx5NIJA57LR+X7+Xf//53Ou200+jgwYPE6/zHxsjrNBwOY8yYMdSiRQvKysqiFi1a0AMPPEC813i93qNe208l+Rnh/e2aa64hh8NBWVlZVFBQQOeeey5VV1eL+8Hnbaw1GwwGhauTrRY890uWLKGCggIqLi6mjIwMys3NpaysLLrqqquorKyM+Nk50ljk+8UB28FgEC+88ALl5+dTVlYW5eTk0OTJk0m+D421/n7Ks8X/l5aW0p/+9Cfq2LEjFRYW0uTJk2n//v3EY5fvA/+/b98+evLJJ6mkpIQAUO/evemee+6h3bt3E1sW0u9ZU+w5PP88t/v27aPf/va35HQ66cwzz6TFixdTVVVVylrg/0OhELxeLz7//HO69tprKTc3l+x2O5199tn0xhtv0M6dO4ktryBKCgWlpaXUp08fysvLo1atWlFBQQHl5+dT586d6ZNPPiE2Q8oTziZCeaD8WSQSQTwex7Zt26hv375ktVrpscceI7/ff9iD0FgMR/b9y4swfTJLS0vp/vvvp44dO1LLli3p8ccfp8rKSuF7ORLxscrKyuivf/0rtWjRgrp06UKTJ08WN4MflkAgkHLOplg0kUgEfr8fI0aMoKysLCosLKTi4mLiIKujEdc1WLduHU2YMIGysrIoNzeX2rVrR/fccw+Vl5cTR65y4FJTPfA892vXrqXTTz+dWrRoQTk5OXTFFVdQIBA4oQWB9HniCPlYLIaHH36YsrOzqUWLFtShQwfieZWvh9e0LKTKG/+GDRvoww8/pOuuu45atmxJLpeLZs6cSWxJSWfAx4t47T/yyCOUm5tLbdq0oby8PLr//vuJ7x/PgzwX8l++7ng8jurqagwYMIBat25N8+bNo3QGI88H/44/i8VimD9/vlgnN9xwA9XU1MDv9x/X9So/74FAACtWrKCMjAxq0aIFORwOeuyxxw5jvsf73vC5/H4/7r//fsrOzqacnBxq166dEFbuvPNOCofDKXMTiURSGIs8x3y/WFi/7bbbyO12U8+ePammpkbcm2NVKI+V5HvO6yR9/hKJBOrq6jBlyhQaOHAgZWZm0tVXX03bt28XTJ2fQfla2RXJ17Vnzx56/PHHKTs7m7Kzs+mUU06hJ554gmpqasQ8NNWzlX7tX3/9NXXv3p0KCwvpnnvuobq6OiGYy/s5k/ycxWIxrFq1ii6//HIhHPbs2ZOmTZtGwWAwWbzokM9cKy4uxjXXXAOXy4VYLAa73Y6DBw/i7rvvxsKFC4nNexxFy+mHRJSSfsJmDk3TkJubq3Xt2hUWiwWnnHKKiJIEGr9gEY+NYwdYMDlkWoXRaITH48HEiRPxz3/+E/v27cOkSZNwzz33aBzU9WPEechZWVna7bffrj399NOoq6vDyy+/jNtvv50qKyvJbrcjGo2KABw+d1OYlwwGA2w2G4qKigCkBi/xfTka7HY7evTooT366KParbfeing8Dr/fj08++QTPPvuscPPIfmE5YPF4wmg0IjMzE5mZmSKdJjMz87A8/xMVsmmcTd1ZWVnCjXM0cBQ7u794bXft2lW74oortKeeegqvvPIKNE3DokWL4Ha7m9Q9Z7fbEYvFcMUVV6Bly5bweDxivLyZAfXXDqTOCQfChUIhhEIhfPfdd7Rlyxb4/X78+9//htlsThEgZD+1HPDFrx0OB2w2G9xuN+655x7Y7XZYLJYGMx4aC7JbxGKxICMjA06nU8w/x70QHV4n5HjDYrFgwIAB6Nq1K0aMGAGfzyfM4e+99x4efvhhYiYbCAREWieAlEJuHHzOMJvNOOuss5CVlYUOHTrA5XKl1JRozCByOYgUQErMEwvUdXV1eOCBB+ihhx7CunXrcN555+GVV17RCgoKNDk2Tp53Tq9kt0A8Hkd+fr52++23a3/4wx8Qj8dRW1uLd999Fz6fTyjJ7IJpioJNvN43bNhAf/zjH+H1etGtWzf8/ve/h9vtPsxVyHPD94uIUFtbC4PBgO7du2tvvPGGNnbsWMTjcdTV1aF9+/bJuh08IW63GxkZGejXrx+eeuopOBwO1NTUICMjA/v378cf/vAHfPvtt8QBFpFIRAQUpgfuMeLxODIyMtCtWzfouo4uXbqIlD/5+439UPAY5MVjsVhw4MAB+u1vf0vz589HIBDAhAkTcNttt2n8OUdxHolYCDIYDHA4HLjooou0G2+8ETU1NViyZAnuuece7N69mzj6OxgMQtO0Rsl/Ptbr1jQNbdq0SclrlzehHyNZmszKysL48eORkZEBi8WC2tpa/Otf/8LOnTuJCzzJaWVNsbHF43E4nU6NY0Hi8ThcLheA1CqHJyrkOeKHNzc3NyW98MfAWpzFYhHMjTNMAKCwsFAbPHiwtmzZMpSVlaGurg5sRWmq+QmHw+jRo4fWt29fsSF98803IhhUvk5+Nll4ZebBqZjTp08XfuF58+ahpqbmMH+uLJCyosLnqqmpQSAQwB133AEOLmxIIGjMtctj42cvKytLPP/xeByZmZkp95uo6Wqd6LoOv9+PzMxM/PGPf0TXrl1TMo5efPFFvPjii8QpbvwbBjNkBmvUANCjRw84nU6ccsopoq4Ef4evvbGugbV0GfF4XAR3Pv744/TOO+9A13UMHjwYU6ZM0VwuFxwOR0oQZkP3nWN4LBYLTCYTMjMzcf3112s33XQTYrEYOnfujFatWmnpRfqaYv/j5/zzzz/HihUr4PP5MGjQIBQUFGhcC0hG+pii0Siys7NFFpfD4cBjjz2m9e/fHz169EDHjh01ANDTN6Oamhqcd9552gMPPACn0wm/349WrVph9erVeOSRR7Bt2zbiIAc555FvEgctyAPLy8uDwWBAdnZ2ymfMsORF9HPAjIILaHAAFjO7559/HnPnzoXRaMSgQYMwYcIEECXTDI+13CiPlysH3n///drZZ5+NSCSCzz77DM8884yoUmez2YT5pinAG2ROTk7KQj3WdCc5dauurg7t2rXTOnbsKAKLAoEAZs2aBSD1IW+qAEu58Ay7iuRiWyc60pkQp4se6/pIzwRhAcFkMiEajSIUCiEzMxPdunXTHnjgAVRUVJDL5UphuMcTnBceiURwwQUXwGw2w2w2Y926dfjPf/5DzDDTLYXyWmJtz+v10jfffAO32w2TyYSdO3fi008/JWa26b+V55Wf902bNiEnJwe33Xab5vP5UgSI4yEQAPUbMZ/HZrNpnE3BQixr3PL3mgLBYBC5ubkIBALo2LGj9uKLL6Jdu3bwer3IyMhATk4OXn75ZXz88cfCJefxeADUB/HJTF7eC9mq0Lp165TvyXnwPxfycyJXzpQZ/IwZM+iTTz5BZmYmiAj/+7//m5IqLgc9yseQBQ153/R4PHC73XjwwQe11q1bY+zYsSnn5WM2xR7IAs0XX3wBi8WCSCSCXr16we/3p1yj7HKQIQe28vizsrJwxRVXYNy4cSLDUOe8S2YizERvueUWbeLEieA6Bjk5OVi3bh1uvvlmVFdXk8zsuLqV/MABEMcym81wOBwih5dNUjypmqY1GuNMNyMCyQU0Z84ceuONN1BYWAgAuPDCC5GTk6OxMHAsDye7DxKJBGw2G3Rdh9PpxNVXXw2j0Qi73Y5Zs2Zh1qxZJFdm40pmxxtcRY7zUWVT2E8BazgVFRXkcDhSargfOHBAlIY9koXoeIGZouzSYWbX3JkfxwLZKgPUaxjHKrSlF1bhOWcTLae3xeNxdO7cWevQoYPGJvmmgK7rQhA/66yztOLiYkQiEVitVjz77LMij5rXo/yMMNjFt3HjRrRr1w7t27cXkeSvvvoqKioqiJk5MyYmdvGxlrt8+XJceOGFsNvtcDgcKX58/n5jzk1Dz4Bc0VHXdXBWSXqufFM8P3KUOQC0bdtW++STT1BQUICKigoEAgFUVVXh97//PRYtWkS6rov6GbKrQ2Y47NLhqq85OTkpwnpjukgasrSx5VbXdVRXV+Pvf/87vF4vqqurMWLECJxxxhkax7HJmVfpx+K8fwZbcHjd5OTk4K677kK3bt3g8Xhgt9uFcMf77fGGyWRCWVkZbd26VfDR6upqOBwOodSm7zHpAoLP5xN1K0wmEwKBAM455xyMGDFCY16hywdgUzfn9d52223arbfeCgCorq6GyWTCunXrcMMNN4jsgvSiCDyQdCnKbrfDarVq8gMhT2RjmND4QWd/I4/J5/PhhRdeEAEoNpsNQ4YMgdVqhc1mQ11dHaxW61EDPWTfFY85GAzi3HPPRUlJCTRNQ3l5OV5//XX4fD4EAgHhe2+KRSMvbjnl8Fj9puw35GhyAKiqqoLD4YDRaERtbS26du0KAIKZMZri+mRthQUEFtKaspRvY4GF4WOtBSELeFwUJhaLiShzPh5b3nhz9vl8TWatAuqFynHjxgmtbMWKFVi/fj0BR2YUzFwyMjIwY8YMnHvuuRg/fryI0fn222+xceNGAPXMNn3eeB6qqqro22+/xahRo0SsAmuHDZ23MdavvP/Jue18bu4ZIJeXTWeyxxMcr8HastPpRF5envbee++hXbt2MJvNyMrKQiwWwy233IIvvviCqqurhcVTrvMizyFfm8ViEWnmx0NhkDV7/svnCYVC+MMf/kAVFRXQdR02mw3nnHMOMjIyhLIrKxU8rnSFiQVQvi/8/VgshvHjx2tFRUUiLkF2YzVFTAGQtOSzoA0AixYtSikOJruJZbDLn+8Px+3Y7Xa0bt1ay8jIEPMp1BOeiLq6OtFZz2Kx4J577sGoUaOE2ctqtWLdunV48MEHyeVypZjn5ZxjeTDs/0wvWpIeFNFYkE2KiUQCCxYsoGXLloEvvKSkBPn5+eK7GRkZx7Rw2QzIwYQc2OdyubQuXbrAbDbDarVizZo1WLRoEdntduHzPd6FmQAI6wDfS77Jxypw8W/5Adq/fz8qKytFsGGfPn0wcODABq0eTVXNC6iXfvl6j+QfPNHBOcfHGmgoC70M1tD4fS5jy+/x5t8UgYb83HEU9JVXXql169ZNmDc//fRTeL3ewzZ3Fug4GvzgwYO0cuVK9OrVC6NHj9batm0r/KDz5s0T6YRyhTmOm+BiRNOmTUPbtm3RuXNn0YzoeK8R2ewMHN5oB0jtisiC2rFaihpjfJmZmSl7gs/nQ5cuXbTJkycDAMrKyoRF4ZlnnsH27duJhSpZ++S1xdYuXdeFYCDPQbrLqDFBUoD7li1baNasWaJhUygUwsCBA48Y0yW7HWShTB4nB6qzsMquZl1PlseWhfmmUkpycnJQUFAgxvz9999j6tSpFA6HxXjTr5Mhl3Y2Go0pJat5DWqaBp0vGoCo3ud0OrniFXJzc7VXXnlFO/PMM4XJIhAI4J133sH9998vJH9ZomItlTU4u92espAYvGAaM1CD/ZpyiebZs2dD13WEQiHU1tbitNNOQ0FBgcbaFE/M0QLx+Hp4UoF6X/bpp5+O8vJyOJ1OaJqGp556StQ9b6oFk95mVnYJHQtYk+Cy1M8//zyqqqqExnn99dejQ4cOGm9g/EA0lT8/nRHww8qb0okOOZATgFin6QF4R4Js9ubnE0j6inft2kVz584lACmCkmwyP96QMwpMJhOysrIwdOhQ2Gw2GI1GLFy4EHV1dWIgPH7+Hce0fPbZZ8jKykLHjh01t9uNIUOGQNd1uN1uLFy4EBUVFRQIBA6bF77GSCSCl19+Gddeey2Kioo0ufcHMyiZETSkWf03SLeAyumVzEADgUDKnsNjagokEgl4vd6UmJbMzExYLBaMHDlSe/XVV8GKXiQSwcqVK3HnnXfC4/GkmNwjkUiK5YXvGwuocsOnxnw2ZXcz/9U0DcFgEJs3b4bP50N+fj4qKytx3nnnoW3bthoRiX0xPZ6N+QO7opj/8XdZmWULHD+rsouqobiY44VIJIKCggKtR48eqKmpgcvlwv79+3Hrrbfir3/9q6gzIAs48rPPY+d7xAKq7OohIuiyOY3f5BrrrDFGIhG8+uqr2vDhw8XJHA4H/vGPf+BPf/oT+Xw+8T6bc4FkAwcAomEDS5MyfgrTOhbIpqFEIoEDBw7Q5s2bU3xfXbt2Fb5XDko8ljHw5PKmy5KqwWBAfn6+CPgym8349ttvsW/fPpIX5fEGB0zKWqK8UcqxAZzPyuCNy2g0YteuXXTXXXfRxx9/DCJCZmYmHnjgAYwbN05jgYHB0ctNtbGxpYYFHo/HA03TTopAw4ZKbKcLnVyOl7/Hm3j6/eLvAkB5eTmNHTsWBw8eFOZ6tuCxm6GpzJt8naxdXnTRRSI18IcffsB3330nhE/ejGStzePxYMqUKbj88svhcDgQDAZx8cUXi9S5Xbt2Ye3atSltYmWhAgAWLVpEsVgMl1xyiahCJwsAzAjke9EYQgGfgzVt3idkgYTz/oH6tN70qP7jhXTTP8d3sBB30UUXaffcc4/oW3OoMi0uvvhi2rdvHwEQpmrWLpln8LzK8Uzp7r6fCzkOgzVbXgMrV66EyWQS8Q89e/YUsXK8t//YGOSOjw1lAxkMBqEQ8jrz+/2HxYYcT/D4rr/+ehQXF4sx5uTk4KmnnsLgwYNp4sSJNG3aNFqxYgXt3r2bIpFIyjzJac28/vg6OGsnhRPyzZMXKTNtq9WKhx56CLt378auXbvEQd944w20atWKrrnmGk2efDZbMiNqqmYYsi+ViFBdXQ0u2wyA6wykSLD/jelblpw5X15edESEHTt2oE2bNg22Uj0e4C5azLhZYOFsEYYsiHHMhcfjoa1bt2Ljxo149dVXUV5ejjPOOAODBw/G0KFD0a9fP81qtYo0ODZXpwcUHU/I5sv0AK2TyX1wpLHabDYEAgG43W4QJQtgccolb8Yc1Guz2bB582aaPXs2li5ditWrV4t1xhu9vK6bwpLCG7Uc49GqVSsMHz4cM2bMgN1ux4cffogLL7xQbGjpzZ6+/fZbqqurw8CBA2EymWAymdCqVSsUFRVh9+7dCAaDmDVrFi677DKx0clMx+/3Y968ebjqqquE1Q5ovtbZ6cy+qYTnY0G6ayMUCuHmm2/WANBTTz0l6kvs378fkyZNwlNPPUUFBQWiTwILPYea6GkGg4EaO3hTBq8veS0nEgkEAgHat29fyn7cpk2bFJ5zrPefG3nxOpa1a6DemsrB8/y8NVWZfgAYPny49tBDD9HDDz+MiooKUY6/vLwcH3zwAWbOnIns7Gzk5uaiqKiIunbtio4dO6Jnz57Izc3VsrOzxV5CRHA4HCku7sOEAr5weQKZuXTt2lX7+uuv0a5dO/L7/UIjePDBB9GmTRsaOnSoxhXvWNrkoJrmMO9yYCH7MdmlkJeX91+ZtRpaVLzYMjMzhdWBBau9e/c2qb/bbreLa2U/bkVFBc4//3yqrKwUkc9AvWkzHA6jtrYWXO2tQ4cOuPXWW3HdddchLy9P83q9opsfb+Tp1pWf2jDr50KeU/7/ZBIKGOkBZpWVlbjrrruorKxM5NtbrVYEAgFYrVahXft8PuzZswfV1dVwu91Cs5FrhqQLTk0ttDHy8vK08847j95++23Y7XbMmTMHK1eupH79+mlAqsshGAxi0aJF6N+/P7p16yYG3L59e23gwIG0fv162Gw2LFy4EDt27KD27dtr6dYuj8dD69evxzPPPNNkigijIcsD3+N0Bvxjvz3eaGicmqbBarXCZDLhvvvu08rLy+mNN96AzWZDNBrFjBkzYLFY8NJLL6V01U0kEnC5XCgrKyPmE0217vj4kUgE+/fvT3HDtWnTpsHv/hg0TQM3CkokEqKtNrtTiJIpwPwsyn0QmgKaponaQBdffLF2ySWX4Oabb6YZM2aI4kVsCThw4ADKysqwadMm/Pvf/xYWlEGDBtEZZ5yBiy66CL1795Ybq4m/h4k3crR6esCgwZBsfzxjxgzcdtttKC0tFT2ab7rpJrz44os0dOhQzWaziTKSVqsVDoejScyXbCricct5p/LDyTEO8m+OBbJrgsEaDj8Q6WZsliSbYoPy+/3gvHTOX3c4HLjvvvuQm5srsibkwB9e/MFgEIFAAFu2bMG6detw7bXXIjMzk0477TT07t2bG0RpbIHw+/2w2WwwmUzCItRUHSPl+ypXsDsZIQsGRqMR/fr1Q1ZWFjIzMxEKhUR2DK8vLorl9/uxbds2fP/999i7d69woTQk7DaVJpN+D3hN9OvXD3369MGOHTsAAC+++CLeeust4ergefD7/bRu3TpMmDABmqaJKqQWiwUXXnghXn75ZVitVvh8Przxxhv485//LJ5tFg43bNiANm3aoHXr1hprek25NmXI8Qvye/IG3JzrtiHhgIXLRx99VAsGgzR16lRkZ2cjHA5j6tSpyMrKokmTJmnRaPSwNHSuVXC8hIL0dS27KThrgH3/drtdMNCfYr3weDyYPXs2HThwAFwe3uVyCaHgUCdadO7cGaeffjq6d++uyd09jzdYm2cl7LnnntPGjRtHS5YswZo1a7B//35UVFSI+8DWALaaLVmyBEuWLMGbb76Jbt260ejRo3HRRRchPz9fxNilPCnpWmS6mYZvQLdu3bSHHnqIxo4dC6vVKhjLH//4RxQVFVGvXr00DuJgi0JTRdfKYEuFyWRK8dMea6Gin3IuzjJgv5PX6xWm36a0FPBYOGYCAAYNGqRlZmampAqla9h8f4PBIKqrq6m0tBRvvvkmXnjhBdjtdrRo0QKjR4+miRMnapFIRFgPAAip+XiDNyD5Oji24EQyyx4N6Vqk7GO++uqrNa465vP54HA4hPbFsTkcaW+xWFBeXk4LFizA7bffjmg0Kt6X0ZRWOrlTH2+S7dq10+644w6699574Xa78cUXXwhrAce6WK1W7NmzB9FoFN27dwcAkeUDAKeddprWt29f2rBhA3RdxyeffII777xTtFTmTfnDDz/EsGHDUtZ2czBeWRGRGeeJDNnt43Q68dhjj6Gmpgaff/65qDPz5ptvwul00sMPP6yxb/+QRq0ZjcYG3QfHa/7lGC+2TAPJ9S4HVB7rGNh1N3jwYNTW1mLZsmWYMmUKvvvuO2RnZ8NgMKCsrAxDhw7FZZddhpKSEu1Y4hUaG3LwalZWFkaOHKmNHDkSlZWVVFFRge3bt2PDhg1Yu3YtNm/ejMrKSlFvKDs7GyaTCR6PBytWrMC6deuwaNEiPP3009SiRYukIP1TBsMPHpsvPvjgAzgcDvHZrl27cP3116O2thY+n0/krXKBheMNudIdAJGvarfbhX+IiFBWViY2rJ8q3aW7A1gSra2tFb4m3hCys7OFFaEp8viBZGAa+/5F3imnmuj1tbBZqpdfs38pPz9fGzRokPbqq69qDz74IILBIH744Qc8/vjj6NChA+3Zs4e8Xu9h0eNNAQ6SZMjBlCc6GvIvywyDtS3WcLKysoQ5EKiPEpb9qjabTbvkkku0p59+OqXXAQcM8/ebA8w0AGDMmDFamzZtRAvoGTNmAICIV9J1HdOnT0e/fv1gs9k02XoCJINob7jhBiH41NbW4uuvvxbKi8FgwI4dO2jVqlXo06ePMO1y6nBTIf0eN5T61lzr9Wjn53XCLlCLxaL94x//0Hr06CGyXRwOB/7+97/jX//6F7GbmGPIjqSRN9b1po+dz2swGLScnBxhKTOZTNixY0dKqu6xgOOuCgoKtK5du2rjx4/XHn/8cfTq1Uusw4yMDJx//vkYPHiwxm4FzsY43mDGzs8VB3yyApGZmal17txZO/fcc7X//d//1d577z3t22+/1aZPn47f//73GDFiBMLhMLxer5ibQCCAuXPn4p577qlX4OWTcjpDQyZImbFwHYNzzz1XmzRpkvBFERH27t2Lq666iqqrq4n7qbtcriZNW+MH0GAwoLCwEIWFhWLziMVi2L17N4BUjfNYIbtB+PcsaMjm7OLiYlHyE2gabSUWi8HlcmHv3r3C7MrR13J3sPQodrn0Jwd+BYNBGAwG3HHHHdqkSZNE5GosFsO4cePA7UdZu2iKDAveFNKFAv7sZEF6BoIsGFitVhF4x88MC55ycxp+Pg8VBcOVV16pde7cuUkDPxsCn5/vC0eAZ2Rk4IILLhClfufPn48dO3YQCzgHDhyguXPn4pxzzgG7H2WBL5FI4IILLkCnTp2EK3Pu3LnweDzCZDxt2jQMHDgQrVu3FhpcUwoEDaE5hYB0pFssGhqbz+eDwWAQQa4ulwuvv/46BgwYIOqVuFwu3H///Zg5cyZx5Ho8HqemsBDIgiYzfLvdjpKSkpSaCatXrxa/OVbXNY+XY6asVivOPvtsrVevXggEAoLvderUCUBSiOBS3o1hfT4aTCYTzGazqMXA78ViMTidTpHib7FYhIATj8fRp08f7e6779beeust7fPPP8eDDz6IwsJChEIh4fqdPXs2Zs6cSQBShQIOEpRNyg1Fd8snvPbaa7Wbb75ZaKiJRAJr1qzB/fffD03TRKvSptJWOKaAF0h+fr7WqlUrMf54PI7y8nIAqcVD/hvwph2LxVBZWSmKZhARunbtilatWmmyv/N4gwNtbDabkAjdbjdqamrI7XanVOqSTW2ye4fngjdmXddx6623an379hWV89atW4d33nlH1KwAmia7hMedLtT8UqBpmqg1L3eo43vF6032j3OtA5fLhc6dO2vsvgJSY2CagjHJmy/fF/Y7JxIJjBw5Ek6nE5FIBLt378ZXX30lKuWtWLEC0WgU7dq1S7HisVBhsViQm5ur9e7dG1arFX6/H6tWrRI1CwKBAD755BNcf/31IoaJm8Q0ZTXHdJwoAgHjx1wZzFw47ZeViJKSEm3SpElo0aJFSrfE+++/H8uXL6dDweQ/usE1xjyku92A+tLa7dq1E+/H43Fs3rxZrK1jFQrk78l9AvLz80Vwpa7ryMjIEPseFwxqiuJ0PD45Jk7OLGNhmVMOOe6D0w0tFgtOO+007d5779X++Mc/okuXLvB4PMKqNnXq1OSezwdmCYzzvlkKS9dqgHpmSpSsoHb33Xdrd9xxBzwej6gVPW/ePDz44IPEJg7ZxCKbsRpzw0ovTsERzePHj0d1dbUIuFi8eLFwb/B35QWVvoj4fZ54rjvPtQ98Ph927NghUjvi8TgmTJgg/PQ/xYT1c8CBjjU1NbDb7XA6ndysRJMfjoZKvfJrm82GYDAIoL4Ykq7rOOuss4QETURYsmSJqMkANE3xjkNSMvH4OY3mSGb59LUl/8+90xlyL3IAYg4aE+nahNFoFOZ0NoV6vV4xSN54WJCTi2ClF/6Sa4MwU+XzcZDU8Yaslctj5aCn3r17a927dxdpXlOmTEE4HIbf78f777+PMWPGIDs7WwNSLUCyYDFq1Chhmdy1axe++uorOBwOLFiwgDp06ICePXtqbPWSgxibAvJ+yZbJSCRCPp8PJpOJBXTxHMn3rbHH0RDz572rrq5OxFzwHgnUC5uytYktVAMHDtRefvll0WVP05LVb++++25s27aNuD4NV9aT0VjZQTxX8v5lMBjg8XhwwQUXiGqWDocD69atw6ZNm4h7gsg860jFvOQ4Jfk9jhXjoF9d10X8lhxY2RjgwEAGj4X7hsg9DtLXjhwLyIHm7L7ifZp518iRI7U77rgD8XgcwWAQLpcLBw4cQCAQSAoFsVgMfr8fdXV1CIfD4I5i8kOeHqQGQJg57XY7HnjgAe3GG29ERUUFsrOzYbfb8d5772HixIkkCxkNobEmtaHKcGazGcXFxbj22mtFHv/+/ftRWlpKLE3xGFjS4g2XGSm/7/V6ReAiT/6h79Bnn30mJnfgwIEYNWqUZjKZ4PV6wXPcFGDTPwtigUAgJTuC70NDle64CJXNZhOR7zz+goICURykqKgIO3fuFM010tfK8cIh87oG1EvyXG2NNcKGKlOyi4SIxEPF7hXWelgzZxMhX1c0Gm000yBvWtx2HKhv98vPAAeJAhBxOESpVdnkwjEcF6LrumjUIm+eXq9XXGdzw263Y/To0dizZw/sdju+/fZbfPzxx1RdXU2rV6/GhRdeKIrfyAV/gHo/d8+ePVFQUCDW8XvvvYdIJIKlS5fi4osvTplL/svWlOMNTdPEGmSwGToYDArTPF/X8bKgpgdZMrEVk6vZsbtYrmHyY8ccOHCg9uSTTyInJ0fETO3Zswc333wzVq5cKbrnptcHaCzBJ70ODMPlciE/P18bO3asUD5ra2uxYsUKBINBsW/L1gWZ5/Azxcdn4SgUCiEYDIreDwcPHkTr1q1FbQY+VmO6xuXOw8zA2a24ceNG4n3A6/WKfY/3aPm65Hnn+89KHK+FSy+9VHO73eJ1OByGCDRkydXr9aYUMZD9ebJ0xiZKZjyxWAyZmZmYOHEixo0bhz179girwTvvvIO3335bpMOlo7HN6jKT54lq2bKlduONN8LlcsHn8yEcDmP+/Pkwm81i4mOxmGgWIgd9yaV8OZuANTu/3w+z2YxZs2aJQitutxtXX321+A5LlE1R6pirpXHkLfuggsEgyRsVC2npD4fT6RSaq9VqhdfrFRuGx+MRjWk4P57T5pqK4YTDYYRCIZItHuxP5rS7hoQT2UUir225FDY/QHJwHgtTje0a4XKwsrWJLUxsreBnj7NmZKtWQ9k8iURCZIREIhH4/X5YLBa4XC4cPHiQToSGUUajEddee63GzXeICG+++Sbeeecd9OrVC7169dLSW43zdTLDbdu2rTZo0CDRzGXlypX45ptvqKqqCkOGDBHuBhkNxUgd7+sEINYmMyRZMUpnSo1hzTga87XZbPB4POB0O55T4NgsfQaDARdeeKH2yCOPgIhQVVUFq9WKDRs24KmnnkJpaalwMTZm6XpGuttQTk222Wz4/e9/r3Xt2hXBYBB5eXmYOnUqQqFQSrvqhpQhfqbC4XCKpm61WgWFw2FkZGSgpqYGXq9XWMSBJCM/En/7KeD252zZY54SCoVQVVWFZ555BjU1NcTXxAIQ8yUG8y85yFy2NLIrwWw2o7CwEA6HA9FoFCUlJcnAUf5xIBCgSCSCPXv2iEXMlJ73LxcC4qpO4XAYLVq00P785z9rw4YNExuSruuYOnUqnE7nYebc4+FnT3/oWKPq3bu3dvnllyMjIwPhcBjTp0/H5s2byW63C42Lo/ZlDZOPxxPK0n04HBaa3Ntvv429e/fC5/Nh5MiRuPrqqzWW6BqqXXC8YLPZhF9aTvVic6rsD0wHM0mOBOeAMG77vGXLFjEnuq6je/fuorgOP2jHG3yuUCgE7vfO45HXLG8YDWUmyG4g2boQi8Xw+eefU1lZGYVCocNaEzeGpsnVCPlB5dxqPpfH44HH4zmshSvPbbrAIIPHyEWruJbE559/Ths2bPjZY28M8DNz8803o66uDpmZmVi3bh3eeustjB8/HkC95QSob1UOQNxjXddxySWXCIsLADz66KMoKipCVlaWxpYx/o0clNYUSLcEsFbJAWl1dXVincpVRRvT586QmSA/J1yoLBAIECsODf32x85x+eWXiwDzcDgMu92O5cuXY8eOHcJKKT87x9M9LD/zVqsVL730EvLy8hCLxbBr1y589NFHojwzW4GPlCVhsViEps7PUjweR21trSi5nZGRgczMTCE8MDVGTIHJZILVak2xdnGmQFlZGa1duxbPPPOMEBR4/fOzIO/t6dfI98Dj8QiBIBgMoqysDJwQMGbMmGQ7Al4wBw8eRCAQwMaNG+HxeMQikgP0OF2KJ1C+GHYlZGVl4bnnnsPAgQOxe/duWCwWEahxvAUCIPWhlAO0bDYb7r//fpx99tmoq6vD5s3/v70zD46ySP/4t+eeZDI5zJJAQoIRWFFhVYiigICgiIigEVgUj2IVKJRFUGTBYzlqBVFRl6XENbAo6Iq6KhAKD/AgIocKcsgVDgMkXCHHJJPJnP37IzxNz8uEDORNQv3sT1UKCDPv291vv/2c/fQerFy5MmwgteEEKTwA4OzuDKDWGvB4PJg5cybftWsXbDYbcnJyMGHCBBiNRthstrCYfFMITWrniRMnwix8sljkl0KLbE1RXJo8Qfv27ePLly8Xmml8fDxGjBgh+tiY7mn5xSNFp6KiQiyyJ06cCEsokuOgsneL+qKNt5tMJrhcLvzwww983rx5oPgvvZjyM2wocuIgnVFRWloKg8EAm80WVpGS5iFt+ZLPjaeFVvZqhUKhcyqHfvvtt3zatGlo164d5AOEmgtSVB566CHhZYqNjUV8fDw6duwoEgqBs+eoAOGhS7/fj+zsbJadnS3CdZs2bcLtt98uFHuKb9NYNEVmOLVZjuWTB5bi0ABQXl4eFg+Wk8Eaoy2yR7Cqqgq//fYbDAYDdu/eHRaWOl9oVzaSjMbaA5Aef/xx9sQTT4hwZUJCAuUvneOx0ytvTFaWZSuY7mG329GxY0f21FNPCeNh0aJF+P777zntECAjQGswyMfFAxBrOGNMhLtsNhtKS0uhNfj08opQsST5udChTN988w0OHjyI999/H1999RWnonEAwsICdSk8jNXujCPPNa15tEb27NkT/fr1A4DaA5GMRiNKSkpQVVWFsrIyFBcXc9kypgGgwSfkIjmkNIRCIbRs2ZLNmjULWVlZYa557fcbA632bTSePQo4Pj6ezZgxA5MnT4bX68WsWbOQm5srcgtkLZEgq4XcgLLrOi8vj7/11luIjY3FbbfdhhdffBEZGRmM9vDLWdRNYan4/X64XC4UFRWJLSqVlZXweDwiIYYECFnHsvfgjMdIHDh0xnrlDz30EBhjYjvM888/jwEDBjDSpIHGCY9EShKsqKgQZ1kYjUYUFRWhuro6LNQDnHXdEgaDQSzSssX+22+/8Tlz5vC77roLPp8PqampYdeRw1ANhSqCyuEb2k8N1C4KO3fuFN4oeaGSSxgD4SfG0fXkLOgXXniBjxgxAjfeeCMuu+wyFun42KbGYKitG5KWlsb69Okj+nD77bcjKSmJ0bjLAg04uwbRO2ixWDBixAgxH3v27Ikbb7yR0bsOnBsGaqr+0ftF73xhYaFQ4Mxm8zm7dSIlBOrZHoKuv2/fPphMJvz4449hh9dFc39Z8TcYDBg5ciTGjh0rYviy9U7XpbCenkZgJI+zrOw//PDD7JNPPgFjDLt378bIkSOxadMm2iUh1g7Zw0m5YgBECOSMXOS03Zx2dsXExIj3jFz8eiiedGig7F0xm80oKyvDypUrxdhOnjwZ+fn5vLS0VHyPKusStMZrx4jqGlgsFnz88cc4deoUMjIy8Mwzz6BFixbM7/fX5hSUlJTw/Px8caznmjVrIJ98KOcWyJXr6Fhguhk9sNjYWHTo0IG9/fbbSEpKEolDTZEFLGu08u+AWq2rZcuWbOzYsfjb3/6GzMxM/OMf/8Abb7zBaSsYTRo5i5a0KbJkioqK+D//+U/+9NNPw2azYdSoUXjjjTdYamoqo/29NPHIQm0KbDYb1q9fz0tKSkQym9lsxsaNG0VOAAl/WkC1ygIJD4/Hg/z8fD5kyBAUFhYiEAggKysL//3vfzF06FBGcTW73d7o/ZPnzbp16+ByuYSyU15ejtWrV3M64ln7PeoXCYkzCbV8586dfPbs2XzIkCHIzc1FQkICevTogeTkZEbb2SjLGoBufaS+GI1GbNu2jW/btk2Eo4xGIzZu3Ihjx46JBYyEPSnfctIkPbfq6mqcPHmSb9myhS9evJhnZ2fzhQsXwmKxYMiQIbonQ10sdApbTU0N+vXrB6fTCa/Xi65du8Jut4eFfmQLkAQoeYC8Xi9uvfVWpKenw+VyYejQoWExZlIeaMFvyuJhpKCFQrUHjeXl5Yn6HmcKLKGgoIDL7WzMtgBnS5lv376dHz16FFarFZs3b0ZRURGn9TyadpDQJK9VYmIiGzNmDPr27SvkAhkbhN7jLucERJInDocDXq8XvXr1Ynl5eejduzcqKiowcuRIrFu3ju/fv5/L6zrNOc65CBdYLBahEHzyyScoLCxETU0NMjIy8MorryA7O5tVV1cLBR+ALuEDWbGg5FjGGBITE/Hkk09i4MCBcDqd2Lt3L0aNGoXc3Fy+b98+Tuu82+0OMxbkhHLabWUymXDy5Ek+depUvnz5cvzpT3/Ca6+9hu7du4uCYczj8WDdunV8/PjxKC4uhsViQfv27TFv3jx06tRJFAGhBy3Hbmmx0WqC5BoNBALYvHkzf+ihh5Ceno61a9ey87k4Ggq5DOVFgbwF5IKkMqyMMaxbt46vXLkSeXl5SE9Px+DBg3HjjTfi2muvZXQMMiWhhEIhFBYW8m+//RbLly/Hr7/+ip49e+KBBx5Ajx49GC1WVqv1nFhoU3H48GE+ZcoUrFy5Uri+AoEAMjIyMHr0aLRp0wYWiwV2u13kDlA8iqpllZeX49ChQ9ixYwe2bt0qFu3OnTtj0KBBaNOmTdiDoxPE9F7gIllQa9as4ePGjcOxY8fE7oGqqiokJSVhyJAhuOmmm+B0OuFwOMQ8Je9JZWUlioqKUFxcjF27dmHXrl2itgRQ64FYunQpcnJymHykr1y/Qa9nyXntNtY5c+bwefPmiZ0OFPOdOHEiOnbsiJSUFLFAkUeAvG5VVVVwu90oKSnBoUOHcOTIEfz66684ceKEKD5z3XXX4auvvmK0b7+5CvnI8WV6/w4dOsTvueceGAwGrFixAqmpqUxOhgIgTuOkynGy185gMOCxxx7j3333Hb777jukpaWJevx0L6owKq8Jjd1PMkhqampw4MABPnjwYJSXl4d5XF9++WU88sgjjKxouW8Nvb+sCMjhwKqqKj5hwgSsXLlSeAKnTp2KcePGMRJA9Vn08jXp8xTbHj16NF+zZg2+/PJLdOzYUcgNet56zD85PKjtN4CwJHkq2lZUVMQ///xzfPzxx9i5cyf69OmDnJwcZGdnIzk5WbRT9kyVlpZi48aN/Ouvv8aHH36I9PR09O/fHzk5ObjqqquYfOywPBZ6QKFFyp+S5+7Jkyf57t27sW7dOnzxxRfYv38/0tPTceedd6JLly64/vrrER8fzyjxkOQ0GRV79uzhP/30E9555x34/X7k5ORg+PDhyMjIYPK7xhYuXMhXr14d5gaurKxEZmYm7r//fnTr1o1dyGIoP3yyavLy8vjs2bOxfv36OpUCeUI3NmQ50oPdvn0737BhA95++224XC4YjUZkZWXhiiuugNFoxJEjR3D06FHs378fGRkZGDhwIB544AG0b9+eUcZ/U7ln5drytJh4PB789NNP/LXXXhMeDqoHTgKHkp5kL4HsWmWMwel0wm63o0WLFmjTpg3atm2L1q1bIzExkTVFcQ7g7AtGYSez2YzTp08jLy+PL1myRLSd5hhZJhQWkF3zQLhlQYse/b828WjRokWMYvuyparXvKTndfDgQb5kyRLk5+fDZDLB6XSK50ptkdsmvy9yEmWkLUdURMzlcqFv374YP378JRE6IKifXq8XEydO5F6vFwsWLGDRLqqy5fn+++/zjRs3Yu7cuYwWtEuFn3/+mefm5mLHjh2Ij48XntITJ04gNTUVw4YNw1133cX0LP8ue49kAfndd9/x5cuX49ChQyKZmPa933HHHbjvvvuQnJwc9SSXlR+g1vvmcrn4yJEjMXXqVPTo0YORlSuHnRt7fdcmmcuWMuccGzZs4P/73/+watUq+P1+XHnllWjfvj1SU1MRFxeH8vJy7Ny5E3v27IHNZkP37t3Rv39/XH/99UhOTmba6zc18rZJt9uNAwcO8HXr1mH16tXYsmULTCYTEhISkJmZiczMTJEDcezYMVRUVKC8vBw9evTA4MGD0bVrV0ZbS2WFCAAY51wcaEMvLO1vT0hIAICI3oDzNZwsa4qjlZaWYsOGDfyOO+5gcuKKTFNMGlkjl70IwNltGsePH+cFBQU4cOAAjh07hmAwiOTkZKSkpOCaa65BixYtGB2vqXXhNsX5DnIf5L+XlpYiMTFR5HfQC0kPmzRGOQZN0GSXPys/CxK8TaEYyEoPHRVsNBpRXl4Oh8MhBKT84tPLL88t7f/LfdL+Scj3I/RcCGSFh1zp9PzItadtX6T3RI7Zykq2rEjR9akMKhCdi7ixkYv2bN++nQeDQWRnZ7NorS3KszhTvIgXFxfjlltuYXp6choCJdsZDAZRII2K55BFHgqFwjwD2rVID2TjjNZ4OkyLxolOuKWdMPW933V5IqgAzvbt2/nll1+O+Ph4RglydJ+mODCtrjbLIfBAIICqqipeXFyMffv24ciRI6LgUSgUQlpaGtq3b4/WrVszOllQKzibSykAzp71Iudv0PkLR44c4cePH0dRUZGoY5CQkIC0tDS0aNECWVlZTJbjZFiQgUQwWkjIVa6dGFqtsD7kiaN1GWkLyzQn5FqmLGfaekexP/mlpcQTxs7uF6dCD9EU/tATeUzlyoqRiEagyUJT3vcrV4RrSmThQIqMttyvNpYo96+uvmoVnUhJjJGEit7WAc07CmHJ71ekNmn/LhPpXZKVKuCsZ6EpylBHA/VFroapbXNd0NyXF2n5d829pgDh80xOEqV/y3+n56LXeyYnQRPyrpVQKCSKmclCOtrwiqwIRLIw5c/RdYHGz52oD21SInBuwSI5UV4WugTJBfkazYmcUyGvAZQgLntS5f6QIiAn5dL1hDGiXXC0VpWcZHgxD5YELxUraW6FoK4XRxau2skPIKKGDCCi5t+YaJU0uSIfTXK5rVoLmpAzXOsSitpknqayxLR9DIVC8Pl8YiGr64WQXcuRhHl9804756P5zoVA3gBtlv3FKNza38ttpZwWOYQij19zQt4tzrmI+0a7tsjWrDwWdJ3mXqzpXdOuc/LziVQ3Qe9aCqTgkwUoe5C0XIgXSSsXtNvWZY+A9p7NrRjI1KWQyd43+lP2NDS3UqDNqYi0xmnnUjRGzTn9k5PigPB9s3UldkTTeABhgutS0OZlAVnX/8t9lT0JtCNBdslrX+zGFpp0D20/orVm6xt/ui5dq6mfl+wFkd3fTUVdygbQ+M+3IYsmtVvOk7lUkQU7WV+UQFlf/2k+UunZut6H5kYraEk5l3dEyFa0XE+jIWhd+mQsEG63W2x7k2loEqDsNaBaJlQKvimfSX3304ZIyItG+SjadzwaIdycaLeTRgpvyKETuS+0ZVv2fhPCUyC/YFphRxpufRNHG8e6VAYvEpSspRXy5LLWvqiRxoW+AzRNf+WFs66FsC63XqT2aV3U2jwDrcXc2P2rb3GSi9hoFSK5/XWFFyJ5xWTquveFhtDOh1z6luZbXbtxzhc2qO8elNdDCWeXQrydni8dMKMN50Q7v8gi1XrGmhvtLoL65o38jumRUxBpPEjwcc7DvCxkDF7oel2XYKTvaw3CppQD0dwrmndZmyRJO7MoYbe55hspMBf7Ltdl2GjHhMnWoVZB0Cab1TdpI91U3lN6KbiP6poUkVy52gxvee+oNmbTFH2TJz3FxGTtj4QMUd+ie77YfHMRyZ0ZyeppKvQOJWifg97zJ9LCobd7+mKRDQ+K4daXFyOjHTuqI6GXpa0XkUKO9Jzl7aXnywW5WGhnVV3ngMhtlO8frafqfNayNqFQG4Zt7mdUV+hGzk853xhcKp4Cmk/UlkiVP2XjUH4G2gTlSDDZMpYnxsWGDhojk1ZPtBMjUo5BNItoc00QWgi1kzca93Yky7MuRUGrUNSVVNRY0MSXiyvRn5Hc/HKORH3KT6TP0D0jeRX07jf1TfZSaeOZdbVRRrvQyv1vSmX1QqjLm3ihQkP2NlzsWqU32rWvvj7RTiG9yhxHynfSbs8FcI6QuJi1Wn4PtWMvh0suVc63fkeSYY2VZ3Qh1LfG12f41eeNFF7YulyUCoVCoVAofl80f6BRoVAoFArFJYFSChQKhUKhUABQSoFCoVAoFIozKKVAoVAoFAoFAKUUKBQKhUKhOINSChQKhUKhUABQSoFCoVAoFIozKKVAoVAoFAoFAKUUKBQKhUKhOINSChQKhUKhUABQSoFCoVAoFIozKKVAoVAoFAoFAKUUKBQKhUKhOINSChQKhUKhUABQSoFCoVAoFIozmJq7AQpFU8E5P+//M8aaqCXNQ339J/6/j4NCoagbFu1CoVAo/n9Da4FSChSK3y8qfKBQKBQKhQKACh8omhDOubJCm4lLYewb6pVs7vY3JpfC81EoAKUUKJoY5aJuWlR48NJGPR/FpYZSChTNQnNYRs2daNjU99fe7/cugBo6/no/vwt9Hs09fxW/D5RSoGg2mlox+D0tqpH6+nvq/6XO711BU1y6qERDRbOiFkf9UWN6aaOej+JSRm1JVOhOMBiE0WgEAFRXVyMmJgY+nw+ff/45X7NmDSZPnoxWrVoxxpj4rM/ng8ViadB9/X4/zGYzAoEADAYDDIazOu/hw4f55s2bsWXLFuzZswd+vx82mw1JSUno0qULunbtiqysLBYbG4uamhrYbDZUVVXB4XCc06fz4fP5YDKZwu4dDAZhMBjAGMORI0f4tGnTcO+992LAgAHM7/eDMQaTydTgfAv5+zSep06d4s888wxCoRCMRqPog8/nA+cccXFxSEtLQ0pKCm655RYkJSWxpKQkMMbEeF5I/+tj48aN/I033kBMTAwCgUDYWAWDQXDOYTKZ4HK5YLFYMGjQINx9992sIXNDHpd9+/bxBQsWYPz48UhLS2MmU62zlNrCORdjRWMYCARgNBrBOUcgEMCCBQs4YwwjR45kJpMJVqtVzJnzEQqFhHeM+uz1evHKK69wp9OJsWPHsmAwKN4Duia1TW7noUOH+JQpUzB58mRcd911TI/3R6EAVPhAoTNySIBzLhZKg8GAxYsXY+PGjUhMTMT06dNRXl6OhIQEeDwe2O32Bt+bBJjJZILP54PRaERJSQnPy8vD0qVL4XQ60b59e/Tv3x/x8fHw+/04duwY8vPz8dZbb+H222/nTz/9NEtKSoLX64XD4QgTVNH0nRZmzjmCwSBMJhOMRiOCwSACgQBWrFiBX375Bbt27UL37t1hMpkQGxsrFIeGIAtyg8GAYDCI4uJi7Nq1Czk5OUhISEAoFBJKCAC4XC4UFxejoKAACxcuRJcuXfjo0aPRoUMHYTDo9Xz8fj8KCwuxbt06jBkzBmlpaeCcw+fzIRgMwmazIRQKCUFsMBiQmZmJUCgUJhgvlmAwCJ/Ph88++wwZGRkYP348AKCyshJxcXEAIBQCr9cLq9UqfmcymcAYw7Zt2/js2bNx5513wmg0wmAwIBAI1KsQAAh7viTwTSYTunXrhhdeeAE33HAD79KlCwNqlTabzQa3243Y2FhUVVUhJiYGBoMBbrcbr7/+Onw+H9LT0wFAKQQK/eCcqx/1o9uP3+8Xf/d6vcI6ysvL4927d+fz58/nDoeDFxUV8UAggEAgAM65sFwb8hMKheDxeOD1esE5h8vlwsSJE3laWhp/9913eVlZmRDywWAQfr8ffr8fLpcLhw8f5n379uXjxo3jwWBQXEvuS333p77Q9QOBgLgf/a5z5878zTff5JmZmXzJkiWcPkP3pPG62J/q6uqwfxcUFPBrr72W79q1i1MbSQjLz8ztdqOgoICPHTuWt2zZkq9bt45TmznnYWPRkJ9Vq1bxlJQUvnPnTh4KheD3+8OeidfrFc/F4/GEjWlD5gWN69dff82zsrL4VVddxY8cOcLlftFnQqFQ2PhwzlFRUYGysjKMGjWKt23blg8fPpzXNebnezfoHnRfeu6TJk3iw4YN4z6fL6zP9F6EQiExBzdu3MhvvfVWfvToUe5yuXR7NupH/XDOVU6BovEgd3NlZSXmzp2LO++8E2PHjmVt2rRBbm6usLR8Pp+w8hsCY0xYmwAwf/58vn79eqxduxb33HMPS0hIEMqAwWAQlqfdbkfr1q3Zhx9+yB5//HEYDAaEQiGYzWZ4vV4A0Vli5BEIhUIwGAwwGo3w+/0Aaq3NRYsWcYPBgPvvv59NnDgR7777Lk6cOMEb6iGQIYtebncwGBRhEKPRKMba6/WipqYGBoMBMTExaNu2LXvllVfYjBkzMGjQIHz55Zfc5/MBQFSWcH243W4AtRa77Hmg/jPGYDabYTQaYTKZYLPZYDAY4Pf7EQgEGnx/AIiPj0dKSgo6dOiAyZMnh7nlCZofJIj9fj+cTify8/P5sWPHMHDgQAC18zoUCkXtRSFvA3B2PEkh/Otf/4qDBw9i4cKFnD5D74Xb7QZjDBaLBZWVlZg/fz7uuusupKWlMYfDITwKCoUeKKVAoRucn3Wz+/1+GI1GMMbw9ddf88LCQgwbNgzV1dWYNWsW1qxZg8OHD3NyDesFCZBPP/2Uv/3223juuefQokUL5nA4hKCXFRBy7/t8PjidTvzxj39k1Bej0Qir1SoEbLRjIOcEkGJ09OhR/uabb+LRRx+Fw+HAgw8+yMrLy7Fjxw6hxOgFhSJoLEiYuN1u1NTUCNe81WoVgjcYDKKmpgZmsxkPP/wwmzBhAqZNmwaPxwOXy6VLu+heRqMRbrdbxOoDgQBqamoQDAbBGANjTIwJKQoNDR0AtWGQyy67DJxzjB49GuvXr8eaNWu4z+cT96Xx8vl8MBgMYYphbm4u7rvvvlpr6kw/5HyIC4HzWq+N2WxGdXU1WrZsycaMGYM5c+Zg79693Ov1ij6bTCbxLN9//31+8OBBDBgwADU1NWCMoaamBrGxsQ0eH4UCUEqBohGgBQ8AqqqqkJeXh5ycHGRlZTGDwYABAwawpKQkrF27FkajETExMcKibggkDB0OB2bPno0+ffrgrrvuYvHx8cIiIxeZz+cT95Qt1PLycgC1C3FlZSUAiNhyfcixZ7KwKb/h+++/h8PhwPDhw5nBYEB8fDwGDx6MFStWwGAwgPOGJ/xSfzjnQvkJBoOIi4uDzWZDbGysiGOTG5s+T14WUhCef/555nA48NFHH3Gn06lL+4DaWDpwdswpv4HaJUNu/wsVuHVhs9lQUVGB2NhYpKenY9q0aXj55ZeFtwQI91oAtWPjcDjw6quv8kAggH79+ol8FfIQUDvrQw6VyMqj0+lEMBjE8OHDWXZ2Nl5//XVYrVZxXYvFArPZjIKCAj516lRMmjQJ7du3Z2azWSTMKhR6oZQChW7IFjIJns2bN/Pjx4/j4YcfFpYhYwzdunVDXl4egsEgPB6PLuEDoNYyX7t2LT98+DBGjx4NAOK+JITIFWs2m8E5DxM6CQkJQhmIi4sTwj3a/svCkwTF6dOn+dq1a9G/f3/ExcWhsrISgUAAQ4cOxebNm1FQUMD1ELrUH/JOcF6bC0EegpqaGhHXJmucrHMShpRoBwB33303NmzYIPJEGgp5akKhEKxWq8gb8Hg8ACByCUihpHbpseuBrkcxeqPRiEceeYTZ7XYsX76c07iQgkYCNxgMYu/evXzp0qV44oknkJSUxJxOJzweDyorK0WoIZr5S5+hfjHGhBfCbDbDZrNhxowZ2LNnD9avX8+9Xq94Tm63G4sWLcKgQYMwaNAg5vF4UF1dLcbzQuapQnE+lFKg0BVyaZKAWr9+PRITE9GuXTsGQLhjBw8ejJMnT+KLL77geriGAYhM8LKyMnTo0AFXX301k+PDspVGCX6MMSF0KJErLi4OHo9HWGlk3dYHZf9TPw0GA2pqanDq1Cns3r0bw4YNg8/nQ1xcHCwWC1q1asWuu+46TJ8+XRf3OLWBhB9jDE6nU+RPkCJEAom2KHLOhWAxmUxCIHbq1An79u3TLbwTDAbF9j2r1Qqz2Qy73Q6bzSZ2TYhkpzNtpBwNPZQSUg4ptOX3+zFs2DAsXLgQFRUVnPPaZEBSbml74MKFC9G5c2f07t2bUegjJiZGjGcgEIjK00V9k0MjVqsVVVVVYu60a9eO3XHHHZg1a5YYA8YYNm3axLdu3YqZM2eC89pdPbRjAlC7DxT6obYkKnSFXJmBQAAlJSX8008/xYsvvijcxDU1NTCZTLjiiitYr169+Jw5c9CrVy8hEMgykmO1FHeuT3CSEDl27BiSk5OFMKRryhYnXUsWQuQ6N5lMF7UFj/aKkyXu9Xphs9kwd+5c9OzZE2lpaUzuSzAYxNChQzFs2DBs2bKF035zq9UqhCFZl9HsgyfI0iUhR+OqdVvTZwGE1XWg79NWuEi1Fy4G2d19RrhxyumgZ0A7Eux2O+6991706tVLtz34JpNJxPFJiOfk5LBVq1bx6dOn47XXXjunvsaOHTv4Dz/8gFWrVjHyNJFiS+NJ3q/6kLeC0r8BiCRQymcYPnw43nvvPcybN49PnDiRcc4xa9Ys/PnPf0Z8fDyj3AfqD3nl9ExYVfx+UbNIoSskzEwmE/7zn/8gISEBt9xyCyNrk3YHBAIB3HfffSgoKEB+fj4PBALCgpJd2bRwRuNCrq6uhtVqhdVqhcfjgcViEa5p2b2q9RTQ/eSENnK9X8j+eNpfTwLYarUiPz+f5+fn44477hBChIrzOBwO3HzzzSw7OxvLli1DeXk5rFarsMxJIaiqqmqSuLHcz5KSEh4XF0c5EVyv8I7X60VFRQUSEhLQokULJCQkwOl0IjExEXa7HU6nEy1btkTLli2FgqUXVAuBEi2NRiMsFguefPJJbNiwARs2bOA09na7HT6fD3PmzMGwYcMQExOjWzvqIiYmBm63G5mZmWzy5MnIzc3F+vXr+auvvsr9fj969OghFAjte6FX3oVCoTwFCt2gvflWqxXbtm3j8+fPx9y5c0XSG1UIJKuvY8eObNCgQXzZsmW49dZbhWuVBBApF0B0Vf5iYmJQXV0Nk8kEj8cjCr+Qu5Wga2ktZrJIa2pqYLfbwxIMo7XEZKutsrIS77zzDnr06IGePXuGdYCETExMDCZNmoTJkydjxIgR3Gw2M4fDEVYwSK8zCSJdR/s7j8cDm82G5ORktmvXLm4ymWAymXRpAHlBWrVqhb/85S9o164d0yp75LmorKxEbGys7jkFNBeqq6uFd6Rz585swIABfMGCBejSpQtCoRBsNhtWr17NDxw4gOnTpzeJJR4MBsUuggceeIBt3bqVz5w5E8ePH8frr78udsbQLhHgbJKoXuEnhUJ5ChS6QVv4XC4Xvv32W6Snp6Nbt25wuVw4dOgQ93g8vLi4mO/evZsfP36cHzp0iD/wwAP49ddf8csvv3BKmpJjrkD0VhC5va+66iocPHgQx48f5+Xl5RGFodYdTpa97Dm40Fg6eSPIC7F3716+adMm3H777Th48CAvLCzkBw4c4AcOHOCnTp3ie/fu5QUFBfzqq69GSkqK2I1B7aP7N+V2M7vdDsYYPB4P9u3bhw4dOiAxMTHqvIrzQaGc0tLSMAtXLtxEMf3Y2FgxD6LN7q8PCiNRnJ4UEI/Hg8mTJ7PS0lJs2rSJWywWFBUV8b///e8YN24cUlJSWLQ7UBoCJTnSGEyaNAlFRUW455570LdvXyYnxdL4yUmZCoUeKPVSoSterxfBYBArVqxATEwMHnvsMcTExIjsbpvNJorXkOcgEAjg448/RnZ2tshKpyQrAOfkA9QFWUxXXnkls9vt/J133sGMGTNE/Je2ecnXls8eoGQ38mREOkPhfJC1RmGE1atXAwBWrVqFFStWiOQ/qn0glyM+fvw4li9fjieffFKU3SVhSAmUTWENyuV3V61ahb59+4o8i4ZC8XfaOirXtSBorOVzGvQKXQSDQeGxio+PFwmPTqcToVAIw4YNw0svvYQbbrgBy5Ytw+WXX457772X0bxpbCgPxWq1wu12IzExkV1xxRX85ptvBnB2N4vagqhoTJRSoNANcv9/88033Ol04l//+pdI7DIajSKXwGKxCAHscrmwdetWzJgxA6NGjeJXXHEFk3MKZAEeDbSoTpkyBc8++ywGDhzIr7nmGia74uVkO1noUwKZ1WoVAkv2GtSX7CYfcnP69Gn+008/YdasWejUqRNiY2NZdXU1p6x32hZJ2/DObJFDbm4uf/TRR5mcYClXX2wIZBlHSjSk31ssFni9XixdupQXFhaiR48eIgmwoVByn9VqDRt3eV6Qt4AOJ9ImnDYEuRYBzUuiqqoKAwcOxOLFizF16lReUFCAuXPniu+QAtGY0O4MAEJppu2IAMQ7RHB+9qwNvQ6sUihU+EChG4FAAHa7Hbm5ubjhhhuQkpLC0tLSWKtWrVhSUhJr2bIlS01NZSkpKczpdLKEhATWrl07dtNNNyErKwtLliyBy+UKy4IHoksyBM5auTabDSNGjGD9+vXDs88+i/3793OqmCfvMiDLlXMOt9uNiooKnDx5kgMQ5zYQ0WS/k9AzGo3Iy8sDYwzXXHMNMjIymNVqRevWrVmrVq1YSkoKy8jIYCkpKSw5OZllZWWxxMRElpOTg5dffhkulwuUeKlnHFsuVkR/0g9QG/6oqalBXl4ef+655zBmzBh06tSJyaGMhhAbGyvCM263W3iVDAaDGF/a5ieHcQB93OOkDFAFS1L6vF4vnE4nkpOT2bRp0/DZZ5+hU6dOuPzyyxlVHGyKREOai+RRIuWoVatWKCsrC9vSSEmy8ncVCj1QSoFCN8xmMz799FNeVFSEAQMGwOPxwGQyoaqqSmyRo0S8uLg4ESpITU1lgwcPxrJly3DkyBGutWjlsrfngzLLqfTrzJkz0apVK0ydOhWff/45p4NrgLNxf6A2pnzs2DE+ZcoU/sEHH4hT78hyjVYgUp36iooKzJs3D3369EGbNm0YnbhIhXmoVj1jDHa7HdXV1XA6nejXrx/KysqQl5fHyVtB99aroqBWIZB/f/ToUT5v3jz++uuvY+rUqRg1ahSjAkN6bAn0+XxiTlitVtjtdlEvgMosAwirKUF/6qEUUBiquroaDodDFDKSn3Pv3r3Zs88+iyFDhjS6ZyAS8tHJVqsVPp8Pp0+fRmJiIgCI0sqyEqBHNVCFglDhA0WDIeHi9Xrx73//G127dkWnTp2Yy+UKO4wnUtY/WUcPPvggW7BgAV+xYgWuvvpqcawyEP3RvXIBHIvFgpSUFPbSSy/xRYsWYcKECWjXrh2/7bbb0Lt3byQmJsLlcmHfvn34/PPPsWbNGgwZMgQDBgwQ+QUXk/luMBiwZMkSXllZiTvvvBPA2TLJsrtaTh4kK7Rt27Zs1KhR/KOPPsKtt97KExISGG3hjNZjIAtQqnjHOcePP/6I3377jTscDnDORQZ7KBTCzp07sWnTJqxevRr9+/fH3//+d3Tv3p0BZ4WQHpDwtVgs+Omnn1BUVMRtNhvoZEAKl1BeBx1G1LdvX0bP9mKQE1ZtNpuYr/KRzBSm4pzjscceE+dfALWufLLg7Xa7UGyp3gHVLtADqrVByi0pUOcLD+h1b4UCUEqBQkd+/vlnHgqFMH78eJSXlyM+Ph6MMVRUVCA+Pj7id+QY/6RJk/DFF1+gpKSEJycnM6ocF20hIbmMLBVJSklJYWPHjsXgwYP5hx9+iG+//RaLFi2C3W4Xwuf666/H/PnzccsttzBSRPx+vxDW0QrlQCCAyspKvnPnTjz11FO4/PLLGXkkLBZLvdYu5xwTJkxgY8eO5YWFhUhNTRXKEeU61IdcpTEmJkZUaHzvvfeEy14uJBUMBtGqVSu0bdsWH3zwAa688kqkpaUxuhYJooYIZYJO/fvDH/6AxYsXo6ysTJQIlr1IFF7gnKNt27bIzMzkVBGzIVDVxo4dO55TDCqakzqpdgEdqkQHEWkLbl0scpvO7NLg7du3VzsLFE0K08stqfj9QnOouroaZWVlPD09XaxiJEzkcw+036XvV1RUwO1287S0NFHZjz4TzQ6Euu5BuN1uGI1GVFZW8lOnTgEAUlJShCJAiVx+vz9MiJPAqG9xJsFw8OBBnpyczOggoWgXdRIKhw4d4klJSYwUKZfLBafTWe/3ZSFOCYqhUAjFxcX8TMiGud1uHgqFEBMTA6PRyILBILfb7UwumQtAuO4vZPdFNASDQZSUlHA6VIhyB6i+BCljcqJleno60yuRLhAIoLCwkGdmZjK6H3kQ6oM8VidOnOA+nw+tW7fWrdoiEQwGxXwzmUw4cuQIj4mJYZdddplu91AozodSChQNhuaQ2+2Gw+FAaWkpkpKSzrEuI2W/y/9H2eby5y500aU6+fKhQGStm0wmcQANnQgou38jVVSU/x0N5OWgbHWKiUdrZVPBJeCskhDtdsRIZaJJkNGZFKRoUW0AGm8aZ6q1QOND46CHYhAKheD1esO2o9J9qe1yASnavhptmesLQVbWGnJtOddFL+WJlDE92qdQXChKKVA0mPMJewBCANT3OaBWqCYkJDByb8uC4UKQyy1Huh8JP1I4tMoHWcq0GEcrGClhjgT7hSbJ0U4Is9ksEjEv1G1PVQm12y8jtVWuIKn9P+BsVvuFnL1wPrTjKAt98tRQnonebnP5GGatdyDa50tKJQltOsNCDy9GpJAGPR/tbgyForH4P5y4OrHVMi43AAAAAElFTkSuQmCC" alt="Noble Wings Academy" style="height:56px;width:auto;display:block;margin:0 auto 12px;"><p>Admin Portal — Secure Login</p></div>
  {% if error %}<div class="alert alert-error">{{ error }}</div>{% endif %}
  <form method="POST">
    <div class="form-group"><label>Username</label><input class="form-input" type="text" name="username" placeholder="Enter username" required autofocus></div>
    <div class="form-group"><label>Password</label><input class="form-input" type="password" name="password" placeholder="Enter password" required></div>
    <button class="login-submit" type="submit">Login to Dashboard</button>
  </form>
</div></div></body></html>"""

# ============================================================
# DASHBOARD
# ============================================================

DASHBOARD_HTML = """<!DOCTYPE html><html><head><title>NWA Admin Dashboard</title>{{ css|safe }}</head><body>
<div class="layout">{{ sidebar|safe }}
<div class="main">
  <div class="topbar"><h1>Dashboard</h1><div class="topbar-right">
    <a href="/admin/export/csv" class="btn btn-green">&#8595; CSV</a>
    <a href="/admin/export/excel" class="btn btn-green">&#8595; Excel</a>
    <a href="/admin/export/pdf" target="_blank" class="btn btn-ghost">&#128438; PDF</a>
  </div></div>
  <div class="content">
    <div class="stats-grid">
      <div class="stat-card"><div class="stat-num" id="stat-total" style="color:var(--mocha)">{{ stats.total }}</div><div class="stat-label">Total</div></div>
      <div class="stat-card"><div class="stat-num" id="stat-new" style="color:var(--info)">{{ stats.new }}</div><div class="stat-label">New</div></div>
      <div class="stat-card"><div class="stat-num" id="stat-contacted" style="color:var(--warning)">{{ stats.contacted }}</div><div class="stat-label">Contacted</div></div>
      <div class="stat-card"><div class="stat-num" id="stat-enrolled" style="color:var(--success)">{{ stats.enrolled }}</div><div class="stat-label">Enrolled</div></div>
      <div class="stat-card"><div class="stat-num" id="stat-dropped" style="color:var(--error)">{{ stats.dropped }}</div><div class="stat-label">Dropped</div></div>
    </div>
    <div class="controls">
      <input class="search-box" id="searchInput" placeholder="Search name, email, phone..." oninput="applyFilters()">
      <select class="filter-select" id="filterStatus" onchange="applyFilters()">
        <option value="">All Statuses</option>
        <option value="new">New</option><option value="contacted">Contacted</option>
        <option value="enrolled">Enrolled</option><option value="dropped">Dropped</option>
      </select>
      <select class="filter-select" id="filterProg" onchange="applyFilters()">
        <option value="">All Programmes</option>
        {% for p in programmes %}<option value="{{ p }}">{{ p }}</option>{% endfor %}
      </select>
    </div>
    <div class="table-wrap">
      {% if enquiries %}
      <table id="mainTable">
        <thead><tr><th>#</th><th>Name</th><th>Email</th><th>Phone</th><th>Programme</th><th>Date</th><th>Status</th><th>Actions</th></tr></thead>
        <tbody>
        {% for e in enquiries|reverse %}
        <tr data-id="{{ e.id }}" data-status="{{ e.status }}" data-prog="{{ e.programme }}"
            data-search="{{ e.firstName }} {{ e.lastName }} {{ e.email }} {{ e.phone }}"
            onclick="openDetail({{ loop.revindex0 }})">
          <td style="color:var(--muted)">{{ loop.index }}</td>
          <td><b>{{ e.firstName }} {{ e.lastName }}</b></td>
          <td style="color:var(--muted)">{{ e.email }}</td>
          <td style="color:var(--muted)">{{ e.phone }}</td>
          <td>{{ e.programme }}</td>
          <td style="color:var(--muted);font-size:12px">{{ e.submittedAt[:10] }}</td>
          <td onclick="event.stopPropagation()"><span class="badge badge-{{ e.status }}">{{ e.status }}</span></td>
          <td onclick="event.stopPropagation()">
            <button class="btn btn-red" style="padding:5px 10px;font-size:11px" onclick="deleteEnquiry('{{ e.id }}',this)">Delete</button>
          </td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      {% else %}<div class="empty">No enquiries yet.</div>{% endif %}
    </div>
  </div>
</div></div>

<div class="modal-overlay" id="detailModal">
<div class="modal" onclick="event.stopPropagation()">
  <div class="modal-header">
    <div><h3 id="modalName"></h3><p id="modalDate" style="color:var(--muted);font-size:12px;margin-top:4px"></p></div>
    <button class="modal-close" onclick="closeModal()">&#215;</button>
  </div>
  <div class="modal-body">
    <div class="detail-grid">
      <div class="detail-field"><label>Email</label><p id="dEmail"></p></div>
      <div class="detail-field"><label>Phone</label><p id="dPhone"></p></div>
      <div class="detail-field"><label>Programme</label><p id="dProg"></p></div>
      <div class="detail-field"><label>Current Licence</label><p id="dLicence"></p></div>
      <div class="detail-field full"><label>Message</label><p id="dMessage" style="white-space:pre-wrap;min-height:50px"></p></div>
      <div class="detail-field full"><label>Status</label>
        <select class="status-select" id="dStatus" style="width:100%;padding:10px 14px" onchange="updateStatus()">
          <option value="new">New</option><option value="contacted">Contacted</option>
          <option value="enrolled">Enrolled</option><option value="dropped">Dropped</option>
        </select>
      </div>
      <div class="detail-field full"><label>Notes</label>
        <textarea class="notes-area" id="dNotes" placeholder="Add internal notes..."></textarea>
      </div>
    </div>
  </div>
  <div class="modal-footer">
    <button class="btn btn-ghost" onclick="closeModal()">Close</button>
    <button class="btn btn-primary" onclick="saveNotes()">Save Notes</button>
  </div>
</div></div>

<div class="toast" id="toast"></div>
<script>
const ENQUIRIES = {{ enquiries_json|safe }};
let currentId = null;
function openDetail(idx){
  const e = [...ENQUIRIES].reverse()[idx];
  if(!e) return;
  currentId = e.id;
  document.getElementById('modalName').textContent = e.firstName+' '+e.lastName;
  document.getElementById('modalDate').textContent = 'Submitted: '+e.submittedAt;
  document.getElementById('dEmail').textContent    = e.email;
  document.getElementById('dPhone').textContent    = e.phone;
  document.getElementById('dProg').textContent     = e.programme;
  document.getElementById('dLicence').textContent  = e.licence||'—';
  document.getElementById('dMessage').textContent  = e.message||'—';
  document.getElementById('dStatus').value         = e.status||'new';
  document.getElementById('dNotes').value          = e.notes||'';
  document.getElementById('detailModal').classList.add('open');
}
function closeModal(){ document.getElementById('detailModal').classList.remove('open'); }
document.getElementById('detailModal').addEventListener('click', closeModal);
function showToast(msg,type='success'){
  const t=document.getElementById('toast');
  t.textContent=msg; t.className='toast '+type+' show';
  setTimeout(()=>t.classList.remove('show'),2800);
}
function updateStatus(){
  const newStatus = document.getElementById('dStatus').value;
  fetch('/admin/update-status',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:currentId,status:newStatus})})
  .then(r=>r.json()).then(d=>{
    if(d.success){
      // Update ENQUIRIES in memory
      const e = ENQUIRIES.find(e=>e.id===currentId);
      if(e){ e.status = newStatus; }
      // Update badge in table row
      document.querySelectorAll('#mainTable tbody tr').forEach(row=>{
        if(row.dataset.id===currentId){
          row.dataset.status = newStatus;
          const badge=row.querySelector('.badge');
          if(badge){ badge.className='badge badge-'+newStatus; badge.textContent=newStatus; }
        }
      });
      // Recalculate and update stat numbers live
      const counts={total:0,new:0,contacted:0,enrolled:0,dropped:0};
      ENQUIRIES.forEach(e=>{ counts.total++; if(counts[e.status]!==undefined) counts[e.status]++; });
      document.getElementById('stat-total').textContent=counts.total;
      document.getElementById('stat-new').textContent=counts.new;
      document.getElementById('stat-contacted').textContent=counts.contacted;
      document.getElementById('stat-enrolled').textContent=counts.enrolled;
      document.getElementById('stat-dropped').textContent=counts.dropped;
      showToast('Status updated to: '+newStatus);
    } else showToast('Error updating status','error');
  });
}
function saveNotes(){
  const notes=document.getElementById('dNotes').value;
  fetch('/admin/update-notes',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({id:currentId,notes:notes})})
  .then(r=>r.json()).then(d=>{ if(d.success)showToast('Notes saved');else showToast('Error','error'); });
}
function deleteEnquiry(id,btn){
  if(!confirm('Delete this enquiry? This cannot be undone.')) return;
  fetch('/admin/delete/'+id,{method:'POST'}).then(r=>r.json()).then(d=>{
    if(d.success){btn.closest('tr').remove();showToast('Enquiry deleted');}else showToast('Error','error');
  });
}
function applyFilters(){
  const q=document.getElementById('searchInput').value.toLowerCase();
  const st=document.getElementById('filterStatus').value.toLowerCase();
  const pr=document.getElementById('filterProg').value.toLowerCase();
  document.querySelectorAll('#mainTable tbody tr').forEach(row=>{
    const match=(!q||row.dataset.search.toLowerCase().includes(q))&&(!st||row.dataset.status===st)&&(!pr||row.dataset.prog===pr);
    row.style.display=match?'':'none';
  });
}
</script></body></html>"""

# ============================================================
# ENROLLED
# ============================================================

ENROLLED_HTML = """<!DOCTYPE html><html><head><title>Enrolled — NWA Admin</title>{{ css|safe }}</head><body>
<div class="layout">{{ sidebar|safe }}
<div class="main">
  <div class="topbar"><h1>Enrolled Students</h1><div class="topbar-right"><a href="/admin/export/csv?status=enrolled" class="btn btn-green">&#8595; Export CSV</a></div></div>
  <div class="content">
    <div class="stats-grid" style="grid-template-columns:repeat(3,1fr)">
      <div class="stat-card"><div class="stat-num" style="color:var(--success)">{{ enrolled|length }}</div><div class="stat-label">Enrolled</div></div>
      <div class="stat-card"><div class="stat-num" style="color:var(--mocha)">{{ programmes|length }}</div><div class="stat-label">Programmes</div></div>
      <div class="stat-card"><div class="stat-num" style="color:var(--yellow)">{{ this_month }}</div><div class="stat-label">This Month</div></div>
    </div>
    <div class="controls">
      <input class="search-box" id="searchInput" placeholder="Search enrolled students..." oninput="applyFilters()">
      <select class="filter-select" id="filterProg" onchange="applyFilters()">
        <option value="">All Programmes</option>
        {% for p in programmes %}<option value="{{ p }}">{{ p }}</option>{% endfor %}
      </select>
    </div>
    <div class="table-wrap">
      {% if enrolled %}
      <table id="enrolledTable">
        <thead><tr><th>#</th><th>Name</th><th>Email</th><th>Phone</th><th>Programme</th><th>Date</th><th>Notes</th></tr></thead>
        <tbody>
        {% for e in enrolled %}
        <tr data-prog="{{ e.programme }}" data-search="{{ e.firstName }} {{ e.lastName }} {{ e.email }}">
          <td style="color:var(--muted)">{{ loop.index }}</td>
          <td><b>{{ e.firstName }} {{ e.lastName }}</b></td>
          <td style="color:var(--muted)">{{ e.email }}</td>
          <td style="color:var(--muted)">{{ e.phone }}</td>
          <td><span class="badge badge-enrolled">{{ e.programme }}</span></td>
          <td style="color:var(--muted);font-size:12px">{{ e.submittedAt[:10] }}</td>
          <td style="color:var(--muted);font-size:12px;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{{ e.notes or '—' }}</td>
        </tr>
        {% endfor %}
        </tbody>
      </table>
      {% else %}<div class="empty">No enrolled students yet.</div>{% endif %}
    </div>
  </div>
</div></div>
<script>
function applyFilters(){
  const q=document.getElementById('searchInput').value.toLowerCase();
  const pr=document.getElementById('filterProg').value.toLowerCase();
  document.querySelectorAll('#enrolledTable tbody tr').forEach(row=>{
    row.style.display=(!q||row.dataset.search.toLowerCase().includes(q))&&(!pr||row.dataset.prog===pr)?'':'none';
  });
}
</script></body></html>"""

# ============================================================
# SETTINGS
# ============================================================

SETTINGS_HTML = """<!DOCTYPE html><html><head><title>Settings — NWA Admin</title>{{ css|safe }}</head><body>
<div class="layout">{{ sidebar|safe }}
<div class="main">
  <div class="topbar"><h1>Settings</h1></div>
  <div class="content"><div style="max-width:500px">
    <h2 style="font-size:16px;margin-bottom:6px;color:var(--accent)">Change Login Credentials</h2>
    <p style="color:var(--muted);font-size:13px;margin-bottom:24px">Update the username and password for this admin panel.</p>
    {% if success %}<div class="alert alert-success">{{ success }}</div>{% endif %}
    {% if error %}<div class="alert alert-error">{{ error }}</div>{% endif %}
    <form method="POST">
      <div class="form-group"><label>Current Password</label><input class="form-input" type="password" name="current_password" placeholder="Enter current password" required></div>
      <div class="form-group"><label>New Username</label><input class="form-input" type="text" name="new_username" value="{{ current_username }}" required></div>
      <div class="form-group"><label>New Password</label><input class="form-input" type="password" name="new_password" placeholder="Min 8 characters" required></div>
      <div class="form-group"><label>Confirm New Password</label><input class="form-input" type="password" name="confirm_password" placeholder="Confirm new password" required></div>
      <button class="btn btn-primary" type="submit" style="width:100%;padding:13px;font-size:14px">Update Credentials</button>
    </form>
  </div></div>
</div></div></body></html>"""

# ============================================================
# ROUTES
# ============================================================

@admin.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        creds = load_creds()
        if request.form.get('username') == creds['username'] and request.form.get('password') == creds['password']:
            session['admin_logged_in'] = True
            return redirect(url_for('admin.dashboard'))
        error = 'Invalid username or password.'
    return render_template_string(LOGIN_HTML, css=BASE_CSS, error=error)

@admin.route('/')
def dashboard():
    if not logged_in(): return redirect(url_for('admin.login'))
    enqs = load_enquiries()
    progs = sorted(set(e.get('programme','') for e in enqs if e.get('programme')))
    return render_template_string(DASHBOARD_HTML, css=BASE_CSS, sidebar=sidebar_html('dashboard'),
        enquiries=enqs, enquiries_json=json.dumps(enqs), stats=get_stats(enqs), programmes=progs)

@admin.route('/enrolled')
def enrolled():
    if not logged_in(): return redirect(url_for('admin.login'))
    all_e = load_enquiries()
    enrolled_e = [e for e in all_e if e.get('status') == 'enrolled']
    progs = sorted(set(e.get('programme','') for e in enrolled_e))
    this_month = sum(1 for e in enrolled_e if e.get('submittedAt','').startswith(datetime.now().strftime('%Y-%m')))
    return render_template_string(ENROLLED_HTML, css=BASE_CSS, sidebar=sidebar_html('enrolled'),
        enrolled=enrolled_e, programmes=progs, this_month=this_month)

@admin.route('/settings', methods=['GET','POST'])
def settings():
    if not logged_in(): return redirect(url_for('admin.login'))
    creds = load_creds(); success = None; error = None
    if request.method == 'POST':
        cur = request.form.get('current_password','')
        new_u = request.form.get('new_username','').strip()
        new_p = request.form.get('new_password','')
        conf  = request.form.get('confirm_password','')
        if cur != creds['password']: error = 'Current password is incorrect.'
        elif len(new_p) < 8: error = 'New password must be at least 8 characters.'
        elif new_p != conf: error = 'New passwords do not match.'
        else:
            save_creds({'username': new_u, 'password': new_p})
            success = 'Credentials updated successfully.'; creds = load_creds()
    return render_template_string(SETTINGS_HTML, css=BASE_CSS, sidebar=sidebar_html('settings'),
        success=success, error=error, current_username=creds['username'])

@admin.route('/update-status', methods=['POST'])
def update_status():
    if not logged_in(): return jsonify({'success':False}), 401
    data = request.get_json(); enqs = load_enquiries()
    for e in enqs:
        if e['id'] == data['id']: e['status'] = data['status']; break
    save_enquiries(enqs)
    return jsonify({'success': True})

@admin.route('/update-notes', methods=['POST'])
def update_notes():
    if not logged_in(): return jsonify({'success':False}), 401
    data = request.get_json(); enqs = load_enquiries()
    for e in enqs:
        if e['id'] == data['id']: e['notes'] = data['notes']; break
    save_enquiries(enqs)
    return jsonify({'success': True})

@admin.route('/delete/<eid>', methods=['POST'])
def delete_enquiry(eid):
    if not logged_in(): return jsonify({'success':False}), 401
    enqs = [e for e in load_enquiries() if e['id'] != eid]
    save_enquiries(enqs)
    return jsonify({'success': True})

@admin.route('/export/csv')
def export_csv():
    if not logged_in(): return redirect(url_for('admin.login'))
    status_filter = request.args.get('status')
    enqs = load_enquiries()
    if status_filter: enqs = [e for e in enqs if e.get('status') == status_filter]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID','First Name','Last Name','Email','Phone','Programme','Licence','Message','Status','Notes','Submitted At'])
    for e in enqs:
        writer.writerow([e.get('id',''), e.get('firstName',''), e.get('lastName',''),
            e.get('email',''), e.get('phone',''), e.get('programme',''),
            e.get('licence',''), e.get('message',''), e.get('status',''),
            e.get('notes',''), e.get('submittedAt','')])
    resp = make_response(output.getvalue())
    resp.headers['Content-Type'] = 'text/csv'
    resp.headers['Content-Disposition'] = f'attachment; filename=nwa_enquiries_{datetime.now().strftime("%Y%m%d")}.csv'
    return resp

@admin.route('/export/excel')
def export_excel():
    if not logged_in(): return redirect(url_for('admin.login'))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
        status_filter = request.args.get('status')
        enqs = load_enquiries()
        if status_filter: enqs = [e for e in enqs if e.get('status') == status_filter]
        wb = Workbook(); ws = wb.active; ws.title = 'NWA Enquiries'
        headers = ['#','First Name','Last Name','Email','Phone','Programme','Licence','Message','Status','Notes','Submitted At']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='163240', end_color='163240', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        widths = [4,15,15,28,14,22,18,40,12,28,20]
        for i,(col,w) in enumerate(zip('ABCDEFGHIJK', widths),1):
            ws.column_dimensions[col].width = w
        for i,e in enumerate(enqs,1):
            ws.append([i, e.get('firstName',''), e.get('lastName',''), e.get('email',''),
                e.get('phone',''), e.get('programme',''), e.get('licence',''),
                e.get('message',''), e.get('status',''), e.get('notes',''), e.get('submittedAt','')])
        out = _io.BytesIO(); wb.save(out); out.seek(0)
        resp = make_response(out.getvalue())
        resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        resp.headers['Content-Disposition'] = f'attachment; filename=nwa_enquiries_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return resp
    except ImportError:
        return 'openpyxl not installed. Run: pip install openpyxl', 500

@admin.route('/export/pdf')
def export_pdf():
    if not logged_in(): return redirect(url_for('admin.login'))
    enqs = load_enquiries()
    stats = get_stats(enqs)
    rows = ''
    for i,e in enumerate(reversed(enqs),1):
        status_color = {'new':'#4a7eff','contacted':'#f5a623','enrolled':'#3ecf8e','dropped':'#ff5c5c'}.get(e.get('status','new'),'#aaa')
        rows += f"""<tr>
          <td>{i}</td>
          <td><b>{e.get('firstName','')} {e.get('lastName','')}</b></td>
          <td>{e.get('email','')}</td>
          <td>{e.get('phone','')}</td>
          <td>{e.get('programme','')}</td>
          <td style="color:{status_color};font-weight:700">{e.get('status','').upper()}</td>
          <td>{e.get('submittedAt','')[:10]}</td>
          <td>{e.get('notes','') or '—'}</td>
        </tr>"""
    html = f"""<!DOCTYPE html><html><head><title>NWA Enquiries Report</title>
    <style>
      body{{font-family:Arial,sans-serif;font-size:12px;color:#111;margin:32px}}
      h1{{font-size:20px;margin-bottom:4px}} p{{color:#666;margin-bottom:20px}}
      .stats{{display:flex;gap:20px;margin-bottom:24px}}
      .stat{{background:#f5f5f5;padding:12px 20px;border-radius:6px;text-align:center}}
      .stat .n{{font-size:28px;font-weight:700}} .stat .l{{font-size:11px;color:#666;margin-top:2px}}
      table{{width:100%;border-collapse:collapse}}
      th{{background:#163240;color:#fff;padding:8px 10px;text-align:left;font-size:11px}}
      td{{padding:7px 10px;border-bottom:1px solid #eee;font-size:11px}}
      tr:nth-child(even) td{{background:#fafafa}}
      @media print{{
        @page{{margin:16mm}}
        button{{display:none}}
        body{{margin:0}}
      }}
    </style></head><body>
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;padding:12px 0">
      <a href="/admin/" style="padding:8px 18px;background:#f0f0f0;border:1px solid #ccc;border-radius:6px;font-size:13px;color:#333;text-decoration:none;font-weight:600">Back to Dashboard</a>
      <button onclick="window.print()" style="padding:8px 20px;background:#163240;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600">Print / Save as PDF</button>
    </div>
    <h1>Noble Wings Academy — Enquiries Report</h1>
    <p>Generated: {datetime.now().strftime('%d %B %Y, %H:%M')} &nbsp;|&nbsp; Total: {stats['total']} enquiries</p>
    <div class="stats">
      <div class="stat"><div class="n" style="color:#4a7eff">{stats['new']}</div><div class="l">New</div></div>
      <div class="stat"><div class="n" style="color:#f5a623">{stats['contacted']}</div><div class="l">Contacted</div></div>
      <div class="stat"><div class="n" style="color:#3ecf8e">{stats['enrolled']}</div><div class="l">Enrolled</div></div>
      <div class="stat"><div class="n" style="color:#ff5c5c">{stats['dropped']}</div><div class="l">Dropped</div></div>
    </div>
    <table><thead><tr><th>#</th><th>Name</th><th>Email</th><th>Phone</th><th>Programme</th><th>Status</th><th>Date</th><th>Notes</th></tr></thead>
    <tbody>{rows}</tbody></table>
    </body></html>"""
    return html

@admin.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('admin.login'))

# ============================================================
# COUNSELLING PAGE
# ============================================================

COUNSELLING_HTML = """<!DOCTYPE html><html><head><title>NWA — Counselling</title>{{ css|safe }}</head>
<body><div class="layout">{{ sidebar|safe }}
<div class="content">
  <div class="topbar">
    <div class="topbar-title">Counselling Bookings</div>
    <div class="topbar-right">
      <span class="week-badge" id="weekLabel"></span>
    </div>
  </div>

  <!-- CALENDAR -->
  <div class="week-nav">
    <h2 id="calTitle">This Week</h2>
  </div>
  <div class="cal-wrap">
    <div class="cal-grid" id="calGrid"></div>
  </div>

  <!-- BOOKING LIST -->
  <div class="section-header" style="margin-bottom:16px;">
    <h2 style="font-size:16px;font-weight:700;color:var(--dark-cocoa);">All Bookings</h2>
    <span class="badge">{{ bookings|length }}</span>
  </div>
  <div class="booking-list" id="bookingList">
    {% if bookings %}
      {% for b in bookings|reverse %}
      <div class="booking-row">
        <div><div class="br-name">{{ b.name }}</div><div class="br-contact">{{ b.email }}</div></div>
        <div class="br-contact">{{ b.phone }}</div>
        <div class="br-slot">{{ b.day }} {{ b.time }}</div>
        <div class="br-date">{{ b.date or '' }} · Booked {{ b.bookedAt[:10] }}</div>
        <form method="POST" action="/admin/delete-booking/{{ b.id }}" onsubmit="return confirm('Delete this booking?')">
          <button class="del-btn">Delete</button>
        </form>
      </div>
      {% endfor %}
    {% else %}
      <div class="empty-cal">No counselling bookings yet.</div>
    {% endif %}
  </div>
</div></div>

<script>
const bookings = {{ bookings_json|safe }};
const TIMES = ['10:00','12:30','15:00','17:30'];
const DAYS  = ['Mon','Tue','Wed','Thu','Fri'];
const MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'];

function getWeekDates() {
  const now = new Date();
  const dow = now.getDay();
  const mon = new Date(now);
  mon.setDate(now.getDate() - (dow === 0 ? 6 : dow - 1));
  mon.setHours(0,0,0,0);
  const dates = [];
  for(let i=0;i<5;i++){
    const d = new Date(mon);
    d.setDate(mon.getDate()+i);
    dates.push(d);
  }
  return dates;
}

function fmt(d){ return d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0'); }

function buildCalendar() {
  const dates = getWeekDates();
  const grid  = document.getElementById('calGrid');
  const now   = new Date();
  today = new Date(now.getFullYear(),now.getMonth(),now.getDate());

  // Week label
  const s = dates[0], e = dates[4];
  document.getElementById('weekLabel').textContent =
    s.getDate()+' '+MONTHS[s.getMonth()]+' – '+e.getDate()+' '+MONTHS[e.getMonth()]+' '+e.getFullYear();
  document.getElementById('calTitle').textContent =
    'Week of '+s.getDate()+' '+MONTHS[s.getMonth()]+' '+s.getFullYear();

  // Build booking lookup: "YYYY-MM-DD|HH:MM" -> booking
  const lookup = {};
  bookings.forEach(b => {
    if(b.date && b.time) lookup[b.date+'|'+b.time] = b;
  });

  // Header row
  grid.innerHTML = '<div class="cal-head"></div>';
  dates.forEach((d,i) => {
    grid.innerHTML += `<div class="cal-head">${DAYS[i]}<br><span style="font-weight:400;opacity:.8">${d.getDate()} ${MONTHS[d.getMonth()]}</span></div>`;
  });

  // Time rows
  TIMES.forEach(t => {
    grid.innerHTML += `<div class="cal-time">${t}</div>`;
    dates.forEach(d => {
      const key  = fmt(d)+'|'+t;
      const bk   = lookup[key];
      const past = d < today;
      if(bk){
        grid.innerHTML += `<div class="cal-cell booked"><div><div class="b-name">${bk.name}</div><div class="b-phone">${bk.phone}</div></div></div>`;
      } else if(past){
        grid.innerHTML += `<div class="cal-cell past">—</div>`;
      } else {
        grid.innerHTML += `<div class="cal-cell free">Available</div>`;
      }
    });
  });
}
buildCalendar();
</script>
</body></html>"""

@admin.route('/counselling')
def counselling():
    if not logged_in(): return redirect(url_for('admin.login'))
    bks = load_bookings()
    return render_template_string(COUNSELLING_HTML,
        css=BASE_CSS, sidebar=sidebar_html('counselling'),
        bookings=bks, bookings_json=json.dumps(bks))

@admin.route('/delete-booking/<bid>', methods=['POST'])
def delete_booking(bid):
    if not logged_in(): return redirect(url_for('admin.login'))
    bks = [b for b in load_bookings() if b.get('id') != bid]
    save_bookings(bks)
    return redirect(url_for('admin.counselling'))